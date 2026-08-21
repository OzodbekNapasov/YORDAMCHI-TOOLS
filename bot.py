import telebot
from telebot import apihelper
import openpyxl
from datetime import datetime, timedelta
import os
import time
import tempfile
import html
import re
import json
import uuid
from fuzzywuzzy import fuzz
from flask import Flask, request, jsonify
from PIL import Image, ImageDraw, ImageFont

# Docbot integratsiyasi uchun modullar
from docbot_config import TEMPLATES as DOCBOT_TEMPLATES, find_template_file
from services.image_builder import render_docx_template_to_image
from services.docx_filler import fill_template

# ATLAS Platformasi modullari
from services.atlas_db import init_db, track_user_activity, track_group_activity, log_audit, log_generated_document
from services.atlas_api import atlas_api
from services.lead_service import process_and_send_lead

# Instagram AutoPoster modullari
from services.insta_poster_service import (
    init_insta_tables,
    get_all_settings as get_insta_settings,
    get_setting as get_insta_setting,
    set_setting as set_insta_setting,
    scan_in_background as scan_insta_background,
    scan_and_enqueue_posts as scan_insta_posts,
    post_next_queued_item as post_next_insta_item,
    post_next_youtube_video,
    get_youtube_schedule_times,
    add_youtube_schedule_time,
    remove_youtube_schedule_time,
    reset_youtube_schedule_times,
    get_queue_stats as get_insta_queue_stats,
    reset_queue_status as reset_insta_queue,
    clear_all_queue as clear_insta_queue,
    toggle_post_like,
    get_post_inline_keyboard
)
from services.insta_scheduler import start_insta_scheduler, stop_insta_scheduler
from services.insta_bot_listener import start_insta_bot_listener, stop_insta_bot_listener

try:
    from meta_ads_bot.facebook_api import MetaAdsManager
    from meta_ads_bot.scheduler import BotScheduler, load_settings as load_meta_settings, save_settings as save_meta_settings
    meta_api = MetaAdsManager()
except Exception as _meta_e:
    print(f"[Meta Ads Init Warn]: {_meta_e}")
    meta_api = None
    load_meta_settings = None
    save_meta_settings = None

TOKEN = os.environ.get("BOT_TOKEN") or os.environ.get("TOKEN") or "8937819411:AAHrCwLyr_Ob3bM0ypwNFYP-SKb1weL97fs"
BOT_VERSION = "2.2.0"
PRIMARY_ADMIN_ID = 8135594558  # Sizning yagona rasmiy Telegram ID ingiz

def get_main_keyboard():
    """1-DARAJALI BOSH MENYU (Kategoriyalar / Papkalar)"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn_kontrakt = telebot.types.KeyboardButton("📁 Kontraktlar va Hisobotlar")
    btn_amaliyot = telebot.types.KeyboardButton("🏥 Malakaviy Amaliyot")
    btn_docs = telebot.types.KeyboardButton("📁 Ma'lumotnomalar")
    btn_buyruq = telebot.types.KeyboardButton("📁 Buyruqlar")
    btn_meta = telebot.types.KeyboardButton("🎯 Meta Ads Manager")
    btn_insta = telebot.types.KeyboardButton("📸 Instagram AutoPoster")
    btn_stats = telebot.types.KeyboardButton("📊 Tizim Statistikasi")
    markup.add(btn_kontrakt, btn_amaliyot)
    markup.add(btn_docs, btn_buyruq)
    markup.add(btn_meta, btn_insta)
    markup.add(btn_stats)
    return markup

def get_insta_poster_keyboard():
    """2-DARAJALI PAPKA: Instagram AutoPoster boshqaruvi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = telebot.types.KeyboardButton("🚀 Hozir yuborish (Keyingi 1 ta)")
    btn2 = telebot.types.KeyboardButton("📥 Instagramdan skanerlash")
    btn3 = telebot.types.KeyboardButton("📊 Navbat va Statistika")
    btn4 = telebot.types.KeyboardButton("⏰ Avto-jadval (Yoqish/O'chirish)")
    btn5 = telebot.types.KeyboardButton("⚙️ Insta Sozlamalar")
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn1, btn2)
    markup.add(btn3, btn4)
    markup.add(btn5, btn_back)
    return markup

def get_meta_ads_keyboard():
    """2-DARAJALI PAPKA: Meta Ads va Lidlar boshqaruvi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = telebot.types.KeyboardButton("💰 Hisob va Balans")
    btn2 = telebot.types.KeyboardButton("🎯 Kampaniyalar")
    btn3 = telebot.types.KeyboardButton("📈 Statistika (Hisobot)")
    btn4 = telebot.types.KeyboardButton("⏰ Avtomatlashtirish")
    btn5 = telebot.types.KeyboardButton("🔄 Meta Yangilash")
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn1, btn2)
    markup.add(btn3, btn4)
    markup.add(btn5, btn_back)
    return markup

def get_meta_insights_inline():
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        telebot.types.InlineKeyboardButton("📅 Bugun", callback_data="meta_ins_today"),
        telebot.types.InlineKeyboardButton("📆 Kecha", callback_data="meta_ins_yesterday"),
        telebot.types.InlineKeyboardButton("📊 Oxirgi 7 kun", callback_data="meta_ins_last_7d"),
        telebot.types.InlineKeyboardButton("🗓 Shu oy", callback_data="meta_ins_this_month")
    )
    return markup

def get_kontrakt_folder_keyboard(suggested_date=None):
    """2-DARAJALI PAPKA: Kontraktlar va Hisobotlar bo'limi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    if suggested_date:
        btn_date = telebot.types.KeyboardButton(suggested_date)
        markup.add(btn_date)
    btn1 = telebot.types.KeyboardButton("📝 Kontraktni yangilash")
    btn2 = telebot.types.KeyboardButton("📸 Guruh screenshotlarini olish")
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn1, btn2)
    markup.add(btn_back)
    return markup

def get_amaliyot_folder_keyboard():
    """2-DARAJALI PAPKA: Malakaviy Amaliyot bo'limi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = telebot.types.KeyboardButton("📁 Yo'nalishlar & Buyruq Yaratish")
    btn2 = telebot.types.KeyboardButton("📥 Oxirgi Buyruqlar Arxivini Ko'rish")
    btn3 = telebot.types.KeyboardButton("📑 Namunaviy So'rovnoma Excel")
    btn4 = telebot.types.KeyboardButton("🌐 ATLAS Web Platformasi Linki")
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn1, btn2)
    markup.add(btn3, btn4)
    markup.add(btn_back)
    return markup

def get_docs_folder_keyboard():
    """2-DARAJALI PAPKA: Ma'lumotnomalar bo'limi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    for tpl in DOCBOT_TEMPLATES:
        if tpl.get("category") == "malumotnoma":
            markup.add(telebot.types.KeyboardButton(tpl["name"]))
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn_back)
    return markup

def get_buyruqlar_folder_keyboard():
    """2-DARAJALI PAPKA: Buyruqlar bo'limi"""
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    for tpl in DOCBOT_TEMPLATES:
        if tpl.get("category") == "buyruq":
            markup.add(telebot.types.KeyboardButton(tpl["name"]))
    btn_back = telebot.types.KeyboardButton("🔙 Asosiy menyuga qaytish")
    markup.add(btn_back)
    return markup

def is_user_allowed(message):
    """Faqat siz (ID: 8135594558) botdan foydalana olishingizni ta'minlash"""
    user_id = message.from_user.id
    username = (message.from_user.username or "").strip().lower()

    # Environment variable orqali qo'shimcha foydalanuvchilar belgilangan bo'lsa
    env_users = os.environ.get("ALLOWED_USERS", "").strip()
    if env_users:
        allowed_list = [i.strip().lower() for i in env_users.split(",") if i.strip()]
        if str(user_id) in allowed_list or (username and f"@{username}" in allowed_list) or (username and username in allowed_list):
            return True

    return user_id == PRIMARY_ADMIN_ID

def send_access_denied(chat_id, user_id):
    """Begonalarga samimiy rad etish xabari (Tugmalarsiz)"""
    msg = f"😅 <b>Voy, shoshmang! Adashib qoldingiz shekilli...</b>\n\n" \
          f"Kechirasiz! Ushbu bot faqat Bosh Buxgalter uchun maxsus yaratilgan.\n\n" \
          f"🔑 Sizning ID: <code>{user_id}</code>\n" \
          f"<i>Agar sizga ham ma'lumot kerak bo'lsa, Bosh Buxgalterimizdan ruxsat so'rab ko'ring! 🤝</i>"
    bot.send_message(chat_id, msg, parse_mode="HTML", reply_markup=telebot.types.ReplyKeyboardRemove())

def save_user_chat_id(chat_id):
    """Foydalanuvchi chat ID sini saqlash"""
    try:
        chats_file = os.path.join(tempfile.gettempdir(), "user_chats.json")
        chats = set()
        if os.path.exists(chats_file):
            with open(chats_file, "r") as f:
                chats = set(json.load(f))
        chats.add(chat_id)
        with open(chats_file, "w") as f:
            json.dump(list(chats), f)
    except Exception:
        pass

def check_and_notify_updates():
    """Yangi versiya chiqqanda faqat ma'mur (buxgalter)ga avtomatik yangilanish xabarini yuborish"""
    try:
        ver_file = os.path.join(tempfile.gettempdir(), "last_notified_version.txt")
        last_ver = ""
        if os.path.exists(ver_file):
            with open(ver_file, "r") as f:
                last_ver = f.read().strip()

        if last_ver != BOT_VERSION:
            update_msg = f"📂 <b>KATEGORIYALAR BO'YICHA TARTIBLANGAN MENYU! (v{BOT_VERSION})</b>\n\n" \
                         f"✨ <b>Barcha xizmatlar mavzular bo'yicha papkalarga ajratildi:</b>\n" \
                         f"• 📁 <b>Kontraktlar va Hisobotlar:</b> Kontrakt yangilash va Guruh screenshotlarini olish xizmatlari.\n" \
                         f"• 📁 <b>Ma'lumotnomalar va Hujjatlar:</b> 1-kursga qabul ma'lumotnomalari va rasmiy shablonlar.\n" \
                         f"• 🔙 <b>Asosiy menyuga qaytish:</b> Istalgan vaqtda papkalar menyusiga bir tugma bilan qaytish.\n\n" \
                         f"<i>Kelajakda yangi mavzu va bo'limlarni ham avtomatik qo'shib ketaveramiz! 🚀</i>"

            try:
                bot.send_message(PRIMARY_ADMIN_ID, update_msg, parse_mode="HTML", reply_markup=get_main_keyboard())
            except Exception:
                pass

            with open(ver_file, "w") as f:
                f.write(BOT_VERSION)
    except Exception as e:
        print(f"NOTIFY ERR: {e}")

# PythonAnywhere bepul tarifida Telegram API proksi orqali ishlaydi
if os.path.exists('/var/www') or 'PYTHONANYWHERE_DOMAIN' in os.environ or 'PYTHONANYWHERE_SITE' in os.environ or 'pythonanywhere' in os.environ.get('HOME', ''):
    os.environ['http_proxy'] = 'http://proxy.server:3128'
    os.environ['https_proxy'] = 'http://proxy.server:3128'
    apihelper.proxy = {'http': 'http://proxy.server:3128', 'https': 'http://proxy.server:3128'}

bot = telebot.TeleBot(TOKEN, threaded=False)
app = Flask(__name__, static_folder="static", template_folder="templates")
init_db()
init_insta_tables()
try:
    start_insta_scheduler()
    start_insta_bot_listener()
except Exception as _sched_err:
    print(f"[Insta Scheduler/Listener Startup Warn]: {_sched_err}")
app.register_blueprint(atlas_api)
user_data = {}

class TelegramProgress:
    """Uzoq davom etadigan jarayonlar uchun universal progress/loading bar tizimi"""
    def __init__(self, bot, chat_id, initial_status="⏳ Jarayon boshlandi..."):
        self.bot = bot
        self.chat_id = chat_id
        self.message_id = None
        self.current_percent = 0
        self.current_status = initial_status
        self.last_update_time = 0
        self.min_interval = 0.6  # Telegram edit limitlarini hisobga olish
        self._send_initial()

    def _render_bar(self, percent):
        blocks = int(round(percent / 10))
        filled = "🟩" * blocks
        empty = "⬜" * (10 - blocks)
        return f"{filled}{empty} {percent}%"

    def _send_initial(self):
        text = f"{self.current_status}\n{self._render_bar(0)}"
        try:
            msg = self.bot.send_message(self.chat_id, text, parse_mode="HTML")
            self.message_id = msg.message_id
            self.last_update_time = time.time()
        except Exception:
            try:
                msg = self.bot.send_message(self.chat_id, text)
                self.message_id = msg.message_id
                self.last_update_time = time.time()
            except Exception:
                pass

    def update(self, status, percent, force=False):
        self.current_status = status
        self.current_percent = min(max(int(percent), 0), 100)
        now = time.time()

        if not force and (now - self.last_update_time < self.min_interval):
            return

        text = f"{self.current_status}\n{self._render_bar(self.current_percent)}"
        if self.message_id:
            try:
                self.bot.edit_message_text(text, self.chat_id, self.message_id, parse_mode="HTML")
                self.last_update_time = now
            except Exception:
                try:
                    self.bot.edit_message_text(text, self.chat_id, self.message_id)
                    self.last_update_time = now
                except Exception:
                    pass

    def success(self, status="✅ Tayyor!"):
        self.update(status, 100, force=True)

    def error(self, err_msg="❌ Jarayonni bajarishda xatolik yuz berdi.\nQayta urinib ko‘ring."):
        if self.message_id:
            try:
                self.bot.edit_message_text(f"{err_msg}", self.chat_id, self.message_id)
            except Exception:
                pass

    def cancel(self, cancel_msg="⛔ Jarayon bekor qilindi."):
        if self.message_id:
            try:
                self.bot.edit_message_text(f"{cancel_msg}", self.chat_id, self.message_id)
            except Exception:
                pass



def make_step_keyboard(button_rows):
    if not button_rows:
        return telebot.types.ReplyKeyboardRemove()
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    for row in button_rows:
        btns = [telebot.types.KeyboardButton(b) for b in row]
        markup.add(*btns)
    return markup

def start_docbot_wizard(chat_id, tpl_index):
    tpl = DOCBOT_TEMPLATES[tpl_index]
    first_step = tpl["steps"][0]
    user_data[chat_id] = {
        "holat": "docbot_step",
        "tpl_index": tpl_index,
        "step": 0,
        "answers": {}
    }
    
    markup = make_step_keyboard(first_step.get("buttons"))
    send_safe_message(
        chat_id,
        f"✅ <b>{escape_html_text(tpl['name'])}</b> tanlandi.\n\n"
        f"<b>(1/{len(tpl['steps'])})</b> {first_step['question']}",
        reply_markup=markup
    )

def process_docbot_generation(chat_id, tpl, answers):
    uid = uuid.uuid4().hex[:8]
    filename = tpl.get("filename", "malumotnoma.docx")
    tpl_id = tpl.get("id", "")
    is_buyruq = tpl.get("category") == "buyruq" or "buyruq" in tpl_id

    # Safidan chiqarish asosiga qarab shablonni aniqlash
    if tpl_id == "buyruq_safidan_chiqarish":
        asos = str(answers.get("asos_turi", "Talaba arizasi")).strip()
        if "bildirgi" in asos.lower() or "rahbar" in asos.lower():
            filename = "Talabalar safidan chiqarish — 2-asos.docx"
        else:
            filename = "Talabalar safidan chiqarish - 1-asos.docx"

    fio = str(answers.get("FIO") or answers.get("IFO") or "Talaba").strip()
    safe_fio = "".join(c for c in fio if c.isalnum() or c in (' ', '_', '-', "'", "’", "‘", "ʼ")).strip()

    output_png = os.path.join(tempfile.gettempdir(), f"doc_{uid}_{safe_fio}.png")
    output_docx = os.path.join(tempfile.gettempdir(), f"doc_{uid}_{safe_fio}.docx")

    status_name = "Buyruq" if is_buyruq else "Ma'lumotnoma"
    progress = TelegramProgress(bot, chat_id, f"⏳ {status_name} bo'yicha 300 DPI Ultra HD rasm va Word fayl tayyorlanmoqda...")

    try:
        progress.update(f"⚙️ {status_name} to'ldirilmoqda...", 40)
        
        # 1. Word (.docx) faylini to'ldirish
        tpl_path = find_template_file(filename)
        fill_template(tpl_path, output_docx, answers)

        # 2. 300 DPI Ultra HD rasm (.png) yaratish
        progress.update("📸 Yuqori sifatli 300 DPI rasm chizilmoqda...", 70)
        render_success = render_docx_template_to_image(filename, output_png, answers, tempfile.gettempdir())

        tpl_clean = tpl.get("name", "").replace("🎓", "").replace("📖", "").replace("📝", "").strip()
        suffix = "buyrug'i" if is_buyruq else "ma'lumotnomasi"
        custom_file_base = f"{fio} — {tpl_clean} {suffix}"

        # 3. Rasmni yuborish
        if render_success and os.path.exists(output_png):
            caption_lines = [
                f"✅ <b>{escape_html_text(fio)}</b> uchun <b>{escape_html_text(tpl_clean)} {suffix}</b> tayyor!\n"
            ]
            if answers.get("buyruq_raqami"):
                caption_lines.append(f"🔢 <b>Buyruq №:</b> {escape_html_text(answers['buyruq_raqami'])}")
            if answers.get("sanasi") or answers.get("SANA"):
                caption_lines.append(f"📆 <b>Sana:</b> {escape_html_text(answers.get('sanasi') or answers.get('SANA'))}")
            if answers.get("YONALISH") or answers.get("yonalishi"):
                caption_lines.append(f"📚 <b>Yo'nalish:</b> {escape_html_text(answers.get('YONALISH') or answers.get('yonalishi'))}")
            if answers.get("KURSI") or answers.get("kursi"):
                caption_lines.append(f"🎯 <b>Kursi:</b> {escape_html_text(answers.get('KURSI') or answers.get('kursi'))}-bosqich")
            if answers.get("GURUHI") or answers.get("guruhi") or answers.get("avvalgi_guruhi"):
                caption_lines.append(f"👥 <b>Guruhi:</b> {escape_html_text(answers.get('GURUHI') or answers.get('guruhi') or answers.get('avvalgi_guruhi'))}")
            if answers.get("yangi_guruhi"):
                caption_lines.append(f"➡️ <b>Yangi guruhi:</b> {escape_html_text(answers['yangi_guruhi'])}")
            if answers.get("asos_turi"):
                caption_lines.append(f"⚖️ <b>Asos:</b> {escape_html_text(answers['asos_turi'])}")

            caption_lines.append("\n<i>Asl shablon formati va rasmiy muhr/imzolar bilan tasdiqlangan.</i>")

            folder_markup = get_buyruqlar_folder_keyboard() if is_buyruq else get_docs_folder_keyboard()

            with open(output_png, "rb") as pf:
                bot.send_photo(
                    chat_id,
                    photo=pf,
                    caption="\n".join(caption_lines),
                    parse_mode="HTML"
                )

        # 4. Word (.docx) faylini ham yuborish
        if os.path.exists(output_docx):
            with open(output_docx, "rb") as df:
                bot.send_document(
                    chat_id,
                    df,
                    visible_file_name=f"{custom_file_base}.docx",
                    caption=f"📄 <b>Word formati (.docx):</b>\n<code>{escape_html_text(custom_file_base)}.docx</code>",
                    parse_mode="HTML",
                    reply_markup=folder_markup
                )

        # 5. Doimiy arxivga saqlash va Supabase Storage bulutiga yuklash
        is_serverless = os.environ.get("VERCEL") or os.environ.get("AWS_LAMBDA_FUNCTION_NAME") or os.path.exists("/tmp")
        saved_dir = "/tmp/saved_documents" if is_serverless else os.path.join(os.path.dirname(os.path.abspath(__file__)), "saved_documents")
        try:
            os.makedirs(saved_dir, exist_ok=True)
        except Exception:
            pass

        permanent_png = os.path.join(saved_dir, f"{uid}_{safe_fio}.png")
        permanent_docx = os.path.join(saved_dir, f"{uid}_{safe_fio}.docx")

        try:
            import shutil
            if os.path.exists(output_png): shutil.copy2(output_png, permanent_png)
            if os.path.exists(output_docx): shutil.copy2(output_docx, permanent_docx)
        except Exception:
            pass

        try:
            from services.supabase_storage import upload_document_to_supabase
            if os.path.exists(permanent_png):
                upload_document_to_supabase(permanent_png, f"{uid}_{safe_fio}.png")
            if os.path.exists(permanent_docx):
                upload_document_to_supabase(permanent_docx, f"{uid}_{safe_fio}.docx")
        except Exception:
            pass

        log_generated_document(
            template_id=tpl.get("id", "hujjat"),
            template_name=tpl.get("name", "Hujjat"),
            recipient_fio=fio,
            data=answers,
            file_type="png",
            file_path=permanent_png,
            created_by=f"bot_user_{chat_id}"
        )
        log_audit(
            actor=str(chat_id),
            module="documents",
            action="generate_document_bot",
            status="success",
            details={"template": tpl.get("name"), "fio": fio}
        )

        progress.success(f"✅ {status_name} muvaffaqiyatli tayyorlandi!")
    except Exception as e:
        progress.error(f"❌ Xatolik yuz berdi:\n<code>{escape_html_text(str(e))}</code>")
    finally:
        for f in [output_png, output_docx]:
            if os.path.exists(f):
                try: os.remove(f)
                except Exception: pass

def escape_html_text(text):
    """HTML rejimida maxsus belgilarni xavfsiz qilish"""
    if not text:
        return ""
    return html.escape(str(text))

def cyrillic_to_latin(text):
    """Kirill alifbosidagi ismlarni Lotin alifbosiga o'tkazish"""
    if not text: return ""
    text = str(text)
    trans = {
        'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'j','з':'z','и':'i',
        'й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
        'у':'u','ф':'f','х':'h','ц':'c','ч':'ch','ш':'sh','щ':'sh','ъ':'','ы':'y','ь':'',
        'э':'e','ю':'yu','я':'ya','ў':'o','ғ':'g','қ':'q','ҳ':'h',
        'А':'a','Б':'b','В':'v','Г':'g','Д':'d','Е':'e','Ё':'yo','Ж':'j','З':'z','И':'i',
        'Й':'y','К':'k','Л':'l','М':'m','Н':'n','О':'o','П':'p','Р':'r','С':'s','Т':'t',
        'У':'u','Ф':'f','Х':'h','Ц':'c','Ч':'ch','Ш':'sh','Щ':'sh','Ъ':'','Ы':'y','Ь':'',
        'Э':'e','Ю':'yu','Я':'ya','Ў':'o','Ғ':'g','Қ':'q','Ҳ':'h'
    }
    for k, v in trans.items():
        text = text.replace(k, v)
    return text

def ismlarni_standartlash(ism):
    """Ismlarni taqqoslash uchun tozalash va standartlash"""
    if not ism: return ""
    ism = cyrillic_to_latin(ism).strip().lower()
    ism = ism.replace("`", "").replace("ʻ", "").replace("‘", "").replace("’", "").replace("'", "")
    ism = ism.replace("о‘", "o").replace("o‘", "o").replace("o'", "o").replace("о'", "o")
    ism = ism.replace("g‘", "g").replace("g'", "g").replace("г‘", "g")
    ism = ism.replace("ch", "c").replace("sh", "s").replace("x", "h").replace("ya", "a").replace("yu", "u")
    return "".join(c for c in ism if c.isalpha() or c.isspace())

def send_safe_message(chat_id, text, reply_markup=None):
    """HTML rejimida xabarni HTML taglarini buzmasdan aqlli bo'lib yuborish"""
    if reply_markup is None:
        reply_markup = get_main_keyboard()

    if len(text) <= 3800:
        try:
            bot.send_message(chat_id, text, parse_mode="HTML", reply_markup=reply_markup)
            return
        except Exception:
            try:
                bot.send_message(chat_id, text, reply_markup=reply_markup)
                return
            except Exception:
                pass

    blocks = text.split("<blockquote>")
    current_chunk = blocks[0]

    for block in blocks[1:]:
        formatted_block = "<blockquote>" + block
        if len(current_chunk) + len(formatted_block) > 3700:
            try:
                bot.send_message(chat_id, current_chunk.strip(), parse_mode="HTML", reply_markup=reply_markup)
            except Exception:
                bot.send_message(chat_id, current_chunk.strip(), reply_markup=reply_markup)
            current_chunk = formatted_block
        else:
            current_chunk += formatted_block

    if current_chunk.strip():
        try:
            bot.send_message(chat_id, current_chunk.strip(), parse_mode="HTML", reply_markup=reply_markup)
        except Exception:
            bot.send_message(chat_id, current_chunk.strip(), reply_markup=reply_markup)

def fmt_num(val):
    if val is None or val == "" or val == "-":
        return "-"
    try:
        fval = float(val)
        if fval == 0:
            return "0"
        return f"{int(round(fval)):,}".replace(",", " ")
    except Exception:
        return str(val)

def get_font(size, bold=True):
    bundled_tnr = os.path.join(os.path.dirname(__file__), 'fonts', 'TimesNewRomanBold.ttf')
    if os.path.exists(bundled_tnr):
        try:
            return ImageFont.truetype(bundled_tnr, size)
        except Exception:
            pass

    bundled_font = os.path.join(os.path.dirname(__file__), 'fonts', 'AppBoldFont.ttf')
    if os.path.exists(bundled_font):
        try:
            return ImageFont.truetype(bundled_font, size)
        except Exception:
            pass

    font_paths = [
        '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf' if bold else '/usr/share/fonts/truetype/freefont/FreeSerif.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        r'C:\Windows\Fonts\timesbd.ttf' if bold else r'C:\Windows\Fonts\times.ttf',
        r'C:\Windows\Fonts\arialbd.ttf' if bold else r'C:\Windows\Fonts\arial.ttf',
    ]
    for path in font_paths:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                pass
    try:
        return ImageFont.load_default(size=size)
    except Exception:
        return ImageFont.load_default()

def generate_group_table_image(group_name, date_str, rows_data, output_path, header_bg_color=(0, 112, 192)):
    """Times New Roman shriftida pixel-perfect HD screenshot hosil qilish (v1.4.0)"""
    S = 3 # 3x Ultra HD Resolution
    col_w = [int(w * S) for w in [110, 60, 520, 310, 230, 250]]
    
    headers = [
        'GURUHI',
        '№',
        'Familiiyasi Ismi va Sharfi',
        'Shu vaqtgacha bo\'lishi\nkerak bo\'lgan to\'lov',
        'Jami',
        'Shu vaqtgacha\nqarzi'
    ]
    
    table_w = sum(col_w)
    title_h = int(55 * S)
    header_h = int(75 * S)
    row_h = int(42 * S)
    summary_h = int(50 * S)
    
    num_rows = len(rows_data)
    table_h = title_h + header_h + (num_rows * row_h) + summary_h
    
    margin = int(24 * S)
    img_w = table_w + 2 * margin
    img_h = table_h + 2 * margin
    
    img = Image.new('RGB', (img_w, img_h), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    font_cell = get_font(int(24 * S), bold=True)
    font_title = get_font(int(30 * S), bold=True)
    font_header = get_font(int(24 * S), bold=True)
    font_summary = get_font(int(25 * S), bold=True)
    
    grid_col = (0, 0, 0)
    border_w = int(3 * S // 2)
    
    ox, oy = margin, margin
    
    # 1. Title Row
    draw.rectangle([ox, oy, ox + table_w, oy + title_h], fill=(255, 255, 255), outline=grid_col, width=border_w)
    title_str = f'Yangilangan sanasi:   {date_str}'
    bbox = font_title.getbbox(title_str)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((ox + (table_w - tw) // 2, oy + (title_h - th) // 2 - bbox[1]), title_str, fill=(0, 0, 0), font=font_title)
    
    # 2. Header Row
    curr_y = oy + title_h
    curr_x = ox
    for idx, (h_text, w) in enumerate(zip(headers, col_w)):
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + header_h], fill=header_bg_color, outline=grid_col, width=border_w)
        lines = h_text.split('\n')
        total_lines_h = len(lines) * int(28 * S)
        line_y = curr_y + (header_h - total_lines_h) // 2
        for line in lines:
            bbox = font_header.getbbox(line)
            tw = bbox[2] - bbox[0]
            tx = curr_x + (w - tw) // 2
            draw.text((tx, line_y - bbox[1]), line, fill=(255, 255, 255), font=font_header)
            line_y += int(28 * S)
        curr_x += w

    # 3. Data Rows
    curr_y += header_h
    tot_kerak, tot_jami, tot_qarzi_musbat = 0.0, 0.0, 0.0
    
    for row_idx, rdata in enumerate(rows_data):
        curr_x = ox
        no_val = str(rdata.get('no', row_idx + 1))
        fio_val = str(rdata.get('fio', ''))
        kerak_num = rdata.get('kerak', 0.0)
        jami_num = rdata.get('jami', 0.0)
        qarzi_num = rdata.get('qarzi', 0.0)
        
        tot_kerak += kerak_num
        tot_jami += jami_num
        if qarzi_num > 0:
            tot_qarzi_musbat += qarzi_num
        
        cells_info = [
            (group_name, 'center', (255,255,255), (0,0,0)),
            (no_val, 'center', (255,255,255), (0,0,0)),
            (fio_val, 'left', (255,255,255), (0,0,0)),
            (fmt_num(kerak_num), 'right', (255,255,255), (0,0,0)),
            (fmt_num(jami_num), 'right', (255,255,255), (0,0,0)),
        ]
        
        if qarzi_num > 0:
            debt_bg = (255, 199, 206)
            debt_fg = (156, 0, 6)
        else:
            debt_bg = (198, 239, 206)
            debt_fg = (0, 97, 0)
            
        cells_info.append((fmt_num(qarzi_num), 'right', debt_bg, debt_fg))
        
        for (c_text, align, bg_col, fg_col), w in zip(cells_info, col_w):
            draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + row_h], fill=bg_col, outline=grid_col, width=border_w)
            bbox = font_cell.getbbox(c_text)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            
            if align == 'center':
                tx = curr_x + (w - tw) // 2
            elif align == 'right':
                tx = curr_x + w - tw - int(16 * S)
            else:
                tx = curr_x + int(16 * S)
                
            ty = curr_y + (row_h - th) // 2 - bbox[1]
            draw.text((tx, ty), c_text, fill=fg_col, font=font_cell)
            curr_x += w
            
        curr_y += row_h

    # 4. Summary Row (JAMI)
    curr_x = ox
    jami_w = col_w[0] + col_w[1] + col_w[2]
    draw.rectangle([curr_x, curr_y, curr_x + jami_w, curr_y + summary_h], fill=header_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox('JAMI')
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (jami_w - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), 'JAMI', fill=(255, 255, 255), font=font_summary)
    curr_x += jami_w
    
    summary_cols = [
        (fmt_num(tot_kerak), col_w[3]),
        (fmt_num(tot_jami), col_w[4]),
        (fmt_num(tot_qarzi_musbat), col_w[5]),
    ]
    for c_text, w in summary_cols:
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + summary_h], fill=header_bg_color, outline=grid_col, width=border_w)
        bbox = font_summary.getbbox(c_text)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        tx = curr_x + w - tw - int(16 * S)
        ty = curr_y + (summary_h - th) // 2 - bbox[1]
        draw.text((tx, ty), c_text, fill=(255, 255, 255), font=font_summary)
        curr_x += w
        
    img.save(output_path, 'PNG', quality=100)
    return output_path

def generate_xulosa_table_image(xulosa_rows, output_path):
    """Guruh rahbarlari va umumiy jamlanma XULOSA rasm jadvalini yaratish"""
    S = 3 # 3x Ultra HD Resolution
    col_w = [int(w * S) for w in [280, 160, 180, 240]]
    headers = ['Guruh rahbari', 'Guruh', 'Talabalar soni', 'Qarzdorligi']
    
    table_w = sum(col_w)
    header_h = int(65 * S)
    row_h = int(42 * S)
    summary_h = int(50 * S)
    
    num_rows = len(xulosa_rows)
    table_h = header_h + (num_rows * row_h) + summary_h
    
    margin = int(24 * S)
    img_w = table_w + 2 * margin
    img_h = table_h + 2 * margin
    
    img = Image.new('RGB', (img_w, img_h), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    font_cell = get_font(int(21 * S), bold=True)
    font_header = get_font(int(22 * S), bold=True)
    font_summary = get_font(int(23 * S), bold=True)
    
    grid_col = (0, 0, 0)
    border_w = int(3 * S // 2)
    ox, oy = margin, margin
    
    # 1. Header Row (To'q olovrang fon: #ED7D31)
    header_bg_color = (237, 125, 49)
    curr_x = ox
    curr_y = oy
    for idx, (h_text, w) in enumerate(zip(headers, col_w)):
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + header_h], fill=header_bg_color, outline=grid_col, width=border_w)
        bbox = font_header.getbbox(h_text)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        tx = curr_x + (w - tw) // 2
        ty = curr_y + (header_h - th) // 2 - bbox[1]
        draw.text((tx, ty), h_text, fill=(255, 255, 255), font=font_header)
        curr_x += w

    # 2. Data Rows
    curr_y += header_h
    tot_talabalar = 0
    tot_qarzdorlik = 0.0
    
    for rdata in xulosa_rows:
        curr_x = ox
        rahbar = str(rdata.get('rahbar', ''))
        guruh = str(rdata.get('guruh', ''))
        soni = int(rdata.get('soni', 0))
        qarz = float(rdata.get('qarz', 0.0))
        
        tot_talabalar += soni
        tot_qarzdorlik += qarz
        
        qarz_bg = (198, 239, 206) if qarz <= 0 else (255, 255, 255)
        
        cells_info = [
            (rahbar, 'center', (255,255,255), (0,0,0)),
            (guruh, 'center', (255,255,255), (0,0,0)),
            (str(soni), 'center', (255,255,255), (0,0,0)),
            (fmt_num(qarz), 'right', qarz_bg, (0,0,0)),
        ]
        
        for (c_text, align, bg_col, fg_col), w in zip(cells_info, col_w):
            draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + row_h], fill=bg_col, outline=grid_col, width=border_w)
            bbox = font_cell.getbbox(c_text)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            
            if align == 'center':
                tx = curr_x + (w - tw) // 2
            elif align == 'right':
                tx = curr_x + w - tw - int(16 * S)
            else:
                tx = curr_x + int(16 * S)
                
            ty = curr_y + (row_h - th) // 2 - bbox[1]
            draw.text((tx, ty), c_text, fill=fg_col, font=font_cell)
            curr_x += w
            
        curr_y += row_h

    # 3. Bottom Summary Row (JAMI - Qizil fon)
    curr_x = ox
    jami_w = col_w[0] + col_w[1]
    summary_bg_color = (255, 0, 0)
    
    draw.rectangle([curr_x, curr_y, curr_x + jami_w, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox('Jami')
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (jami_w - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), 'Jami', fill=(0, 0, 0), font=font_summary)
    curr_x += jami_w
    
    w_soni = col_w[2]
    draw.rectangle([curr_x, curr_y, curr_x + w_soni, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox(str(tot_talabalar))
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (w_soni - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), str(tot_talabalar), fill=(0, 0, 0), font=font_summary)
    curr_x += w_soni
    
    w_qarz = col_w[3]
    qarz_str = fmt_num(tot_qarzdorlik)
    draw.rectangle([curr_x, curr_y, curr_x + w_qarz, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox(qarz_str)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    tx = curr_x + w_qarz - tw - int(16 * S)
    ty = curr_y + (summary_h - th) // 2 - bbox[1]
    draw.text((tx, ty), qarz_str, fill=(0, 0, 0), font=font_summary)
    
    img.save(output_path, 'PNG', quality=100)
    return output_path

# Webhook Marshrutlari
@app.route('/' + TOKEN, methods=['POST'])
def getMessage():
    import sys, json as _json, traceback as _tb
    try:
        raw = request.get_data().decode('utf-8')
        data = _json.loads(raw)
        msg = data.get('message') or {}
        cid = msg.get('chat',{}).get('id')
        if cid:
            save_user_chat_id(cid)
            
        print(f"WEBHOOK: chat={cid} text={msg.get('text','')}", file=sys.stderr, flush=True)
        update = telebot.types.Update.de_json(raw)
        bot.process_new_updates([update])
        
        check_and_notify_updates()
        
        print("WEBHOOK: done", file=sys.stderr, flush=True)
    except Exception as e:
        print(f"WEBHOOK ERR: {e}", file=sys.stderr, flush=True)
        _tb.print_exc(file=sys.stderr)
    return "!", 200

@app.route("/")
@app.route("/atlas")
@app.route("/login")
def index_atlas():
    from flask import render_template
    return render_template("atlas.html")

@app.route("/bot_status")
def webhook_status_check():
    check_and_notify_updates()
    return f"Bot PythonAnywhere/Vercel bulutida 24/7 faol! (v{BOT_VERSION})", 200

@app.route("/set_webhook", methods=['GET'])
def set_webhook_route():
    host_url = request.host_url.rstrip('/')
    webhook_url = f"{host_url}/{TOKEN}"
    try:
        bot.remove_webhook()
        success = bot.set_webhook(url=webhook_url)
        if success:
            check_and_notify_updates()
            return f"<h3>✅ Webhook muvaffaqiyatli o'rnatildi!</h3><p>URL: <b>{webhook_url}</b></p><p>Versiya: <b>v{BOT_VERSION}</b></p><p>Endi Telegram botingizga /start yuborib tekshirishingiz mumkin.</p>", 200
        else:
            return "<h3>❌ Webhook o'rnatilmadi!</h3>", 500
    except Exception as e:
        return f"<h3>❌ Xatolik: {str(e)}</h3>", 500

@app.route("/delete_webhook", methods=['GET'])
def delete_webhook_route():
    try:
        success = bot.remove_webhook(drop_pending_updates=True)
        if success:
            return "<h3>✅ Webhook muvaffaqiyatli uzildi va o'chirildi!</h3>", 200
        else:
            return "<h3>❌ Webhook uzib bo'lmadi!</h3>", 500
    except Exception as e:
        return f"<h3>❌ Xatolik: {str(e)}</h3>", 500

@app.route("/webhook_info", methods=['GET'])
def webhook_info():
    try:
        info = bot.get_webhook_info()
        return jsonify({
            "version": BOT_VERSION,
            "url": info.url,
            "has_custom_certificate": info.has_custom_certificate,
            "pending_update_count": info.pending_update_count,
            "last_error_date": info.last_error_date,
            "last_error_message": info.last_error_message,
            "max_connections": info.max_connections,
            "ip_address": info.ip_address
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/lead", methods=['GET', 'POST', 'OPTIONS'])
@app.route("/api/leads", methods=['GET', 'POST', 'OPTIONS'])
def handle_lead_submission():
    if request.method == 'OPTIONS':
        resp = jsonify({"status": "ok"})
        resp.headers.add("Access-Control-Allow-Origin", "*")
        resp.headers.add("Access-Control-Allow-Headers", "Content-Type,Authorization")
        resp.headers.add("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        return resp, 200

    if request.method == 'GET':
        resp = jsonify({"status": "active", "service": "Meta Ads & Website Lead Receiver"})
        resp.headers.add("Access-Control-Allow-Origin", "*")
        return resp, 200

    try:
        data = request.get_json(silent=True) or request.form.to_dict() or {}
        res = process_and_send_lead(data)
        resp = jsonify(res)
        resp.headers.add("Access-Control-Allow-Origin", "*")
        return resp, 200 if res.get("success") else 400
    except Exception as e:
        resp = jsonify({"success": False, "error": str(e)})
        resp.headers.add("Access-Control-Allow-Origin", "*")
        return resp, 500

# Telegram Bot Handlerlari
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    chat_id = message.chat.id
    if message.from_user:
        track_user_activity(
            telegram_id=message.from_user.id,
            username=message.from_user.username or "",
            first_name=message.from_user.first_name or "",
            last_name=message.from_user.last_name or ""
        )
        log_audit(
            actor=str(message.from_user.id),
            module="bot",
            action="command_start",
            status="success",
            details={"chat_id": chat_id, "text": message.text}
        )
    if not is_user_allowed(message):
        send_access_denied(chat_id, message.from_user.id)
        return
    save_user_chat_id(chat_id)
    user_data[chat_id] = {}
    send_safe_message(chat_id, f"🚀 <b>Salom! Aqlli kontrakt va hujjatlar platformasi faol.</b>\n"
                               f"📌 <b>Tizim versiyasi:</b> <code>v{BOT_VERSION}</code>\n"
                               f"🔑 <b>Foydalanuvchi ID:</b> <code>{message.from_user.id}</code> (Ruxsat berilgan)\n\n"
                               f"📂 <b>ASOSIY KATEGORIYALAR MENYUSI</b>\n"
                               f"Kerakli bo'lim papkasini tanlang:",
                               reply_markup=get_main_keyboard())

@bot.message_handler(func=lambda message: message.text == "📝 Kontraktni yangilash")
def start_kontrakt_yangilash(message):
    chat_id = message.chat.id
    if not is_user_allowed(message):
        send_access_denied(chat_id, message.from_user.id)
        return
    save_user_chat_id(chat_id)
    user_data[chat_id] = {"holat": "baza_kutish"}
    send_safe_message(chat_id, "📑 <b>1-BOSQICH: ASOSIY BAZANI YUKLASH</b>\n\n"
                               "Iltimos, kontraktlar kiritilgan asosiy <b>.xlsx</b> faylingizni yuboring:")

@bot.message_handler(func=lambda message: message.text == "📸 Guruh screenshotlarini olish")
def start_guruh_screenshot(message):
    chat_id = message.chat.id
    if not is_user_allowed(message):
        send_access_denied(chat_id, message.from_user.id)
        return
    save_user_chat_id(chat_id)
    user_data[chat_id] = {"holat": "guruh_fayl_kutish"}
    send_safe_message(chat_id, "📸 <b>GURUH SCREENSHOTLARINI OLISH</b>\n\n"
                               "Guruhlar screenshotlarini olish uchun tayyorlangan Excel (<b>.xlsx</b>) faylingizni yuboring:")

# Matnli xabarlar handler
@bot.message_handler(func=lambda message: message.content_type == 'text' and not message.text.startswith('/'))
def handle_text_messages(message):
    chat_id = message.chat.id
    if not is_user_allowed(message):
        send_access_denied(chat_id, message.from_user.id)
        return
    save_user_chat_id(chat_id)
    holat = user_data.get(chat_id, {}).get("holat")
    user_text = message.text.strip()

    # 1-Darajali Papkalar (Kategoriyalar) va Ortga qaytish
    if user_text == "📸 Instagram AutoPoster":
        user_data[chat_id] = {}
        send_safe_message(
            chat_id,
            "📸 <b>INSTAGRAM AUTOPOSTER MENYUSI</b>\n\n"
            "Instagram sahifasidagi postlarni xronologik tartibda kanalga/botga avtomatik joylab borish bo‘limi:",
            reply_markup=get_insta_poster_keyboard()
        )
        handle_insta_dashboard(chat_id)
        return

    if user_text == "🚀 Hozir yuborish (Keyingi 1 ta)":
        handle_insta_post_one(chat_id)
        return

    if user_text == "📥 Instagramdan skanerlash":
        handle_insta_scan(chat_id)
        return

    if user_text == "📊 Navbat va Statistika":
        handle_insta_dashboard(chat_id)
        return

    if user_text == "⏰ Avto-jadval (Yoqish/O'chirish)":
        enabled = get_insta_setting("auto_schedule_enabled", "0") == "1"
        new_val = "0" if enabled else "1"
        set_insta_setting("auto_schedule_enabled", new_val)
        status_word = "faollashtirildi (yoqildi)" if new_val == "1" else "to‘xtatildi (o‘chirildi)"
        send_safe_message(
            chat_id,
            f"✅ <b>Avto-jadval {status_word}!</b>\n\n"
            f"Endi bot har {get_insta_setting('interval_minutes', '60')} daqiqada navbatdagi 1 ta postni avtomatik yuboradi.",
            reply_markup=get_insta_poster_keyboard()
        )
        return

    if user_text == "⚙️ Insta Sozlamalar":
        handle_insta_settings_menu(chat_id)
        return

    if user_text == "🎯 Meta Ads Manager":
        user_data[chat_id] = {}
        send_safe_message(
            chat_id,
            "🎯 <b>META ADS MANAGER BOSHQARUVI</b>\n\n"
            "Reklamalarni yoqish/o'chirish, byudjetni o'zgartirish, hisob balansi va statistika hisobotlarini olish uchun quyidagi bo'limni tanlang:",
            reply_markup=get_meta_ads_keyboard()
        )
        return

    if user_text == "💰 Hisob va Balans":
        handle_meta_account_info(chat_id)
        return

    if user_text == "🎯 Kampaniyalar":
        handle_meta_campaigns_list(chat_id)
        return

    if user_text == "📈 Statistika (Hisobot)":
        handle_meta_insights_menu(chat_id)
        return

    if user_text == "⏰ Avtomatlashtirish":
        handle_meta_automation_menu(chat_id)
        return

    if user_text == "🔄 Meta Yangilash":
        send_safe_message(chat_id, "🔄 Ma'lumotlar yangilandi!", reply_markup=get_meta_ads_keyboard())
        handle_meta_account_info(chat_id)
        return

    # Instagram State Input tekshiruvi
    user_id_val = message.from_user.id
    if user_id_val in INSTA_USER_STATE:
        insta_state = INSTA_USER_STATE.pop(user_id_val, None)
        if insta_state:
            act = insta_state.get("action")
            if act == "set_insta_user":
                val = user_text.replace("@", "").strip()
                set_insta_setting("insta_username", val)
                bot.reply_to(message, f"✅ <b>Instagram profili <code>@{val}</code> ga o‘zgartirildi!</b>", reply_markup=get_insta_poster_keyboard(), parse_mode="HTML")
                return
            elif act == "set_insta_chat":
                val = user_text.strip()
                set_insta_setting("target_chat_id", val)
                bot.reply_to(message, f"✅ <b>Target Chat / Kanal ID <code>{val}</code> ga o‘zgartirildi!</b>", reply_markup=get_insta_poster_keyboard(), parse_mode="HTML")
                return
            elif act == "set_insta_interval":
                val = user_text.strip()
                try:
                    num = int(val)
                    if num < 1:
                        bot.reply_to(message, "❌ Oraliq vaqti kamida 1 daqiqa bo‘lishi kerak.", parse_mode="HTML")
                        return
                    set_insta_setting("interval_minutes", str(num))
                    bot.reply_to(message, f"✅ <b>Yuborish oralig‘i har <code>{num}</code> daqiqa qilib belgilandi!</b>", reply_markup=get_insta_poster_keyboard(), parse_mode="HTML")
                    return
                except ValueError:
                    bot.reply_to(message, "❌ Noto‘g‘ri son kiritildi. Masalan: <code>60</code> deb yozing.", parse_mode="HTML")
                    return
            elif act == "add_yt_time":
                val = user_text.strip()
                ok, res = add_youtube_schedule_time(val)
                if ok:
                    bot.reply_to(message, f"✅ <b>YouTube Shorts yuklash vaqti qo‘shildi: <code>{res}</code></b>", parse_mode="HTML")
                else:
                    bot.reply_to(message, f"❌ <b>Xatolik:</b> {res}", parse_mode="HTML")
                handle_insta_youtube_schedule_menu(message.chat.id)
                return

    # Meta Ads State Input tekshiruvi
    user_id_val = message.from_user.id
    if user_id_val in META_USER_STATE:
        state_obj = META_USER_STATE.pop(user_id_val, None)
        if state_obj:
            action = state_obj.get("action")
            if action == "set_custom_budget_limit":
                text_clean = user_text.replace("$", "").replace(",", ".").strip()
                try:
                    val = float(text_clean)
                    if val <= 0:
                        bot.reply_to(message, "❌ Byudjet 0 dan katta bo‘lishi kerak.", parse_mode="HTML")
                        return
                    bal_info = meta_api.get_balance_details() if meta_api else {}
                    current_spent = float(bal_info.get("amount_spent", 0))
                    settings = load_meta_settings() if load_meta_settings else {}
                    settings["custom_budget_limit"] = val
                    settings["initial_spent_base"] = current_spent
                    settings["alert_threshold_sent"] = False
                    if save_meta_settings: save_meta_settings(settings)
                    bot.reply_to(
                        message,
                        f"✅ <b>Byudjet limiti muvaffaqiyatli ${val:.2f} qilib belgilandi!</b>\n\n"
                        f"📊 Hozirgi qoldiq: <b>${val:.2f}</b>\n"
                        f"🚨 Ushbu byudjet sarflanib <b>0.00 $</b> ga yetganda, bot sizga darhol bildirishnoma yuboradi. <i>(Reklamalaringiz to‘xtatilmaydi, qarzga ishlashda davom etadi).</i>",
                        reply_markup=get_meta_ads_keyboard(),
                        parse_mode="HTML"
                    )
                except ValueError:
                    bot.reply_to(message, "❌ Noto‘g‘ri raqam kiritildi. Iltimos, masalan: <code>50</code> yoki <code>100</code> deb yozing.", parse_mode="HTML")
                return

            elif action == "change_budget":
                campaign_id = state_obj.get("campaign_id")
                text_clean = user_text.replace("$", "").replace(",", ".").strip()
                try:
                    val = float(text_clean)
                    if val <= 0:
                        bot.reply_to(message, "❌ Byudjet 0 dan katta bo‘lishi kerak.", parse_mode="HTML")
                        return
                    res = meta_api.set_campaign_budget(campaign_id, val) if meta_api else {}
                    if "error" in res:
                        bot.reply_to(message, f"❌ Xatolik yuz berdi:\n{res['error'].get('message')}", parse_mode="HTML")
                    else:
                        bot.reply_to(message, f"✅ <b>Kampaniya kunlik byudjeti muvaffaqiyatli ${val:.2f} ga o‘zgartirildi!</b>", reply_markup=get_meta_ads_keyboard(), parse_mode="HTML")
                except ValueError:
                    bot.reply_to(message, "❌ Noto‘g‘ri raqam kiritildi. Iltimos, faqat raqam yuboring (masalan: <code>20</code>).", parse_mode="HTML")
                return

            elif action in ["set_pause_time", "set_resume_time"]:
                val = user_text.strip()
                if not re.match(r"^([01]?[0-9]|2[0-3]):[0-5][0-9]$", val):
                    bot.reply_to(message, "❌ Noto‘g‘ri vaqt formati. Iltimos, <code>23:00</code> yoki <code>07:00</code> kabi formatda yozing.", parse_mode="HTML")
                    return
                settings = load_meta_settings() if load_meta_settings else {}
                key = "pause_time" if action == "set_pause_time" else "resume_time"
                settings[key] = val
                if save_meta_settings: save_meta_settings(settings)
                label = "Tungi to‘xtatish" if key == "pause_time" else "Ertalabki yoqish"
                bot.reply_to(message, f"✅ <b>{label} vaqti <code>{val}</code> ga o‘rnatildi!</b>", reply_markup=get_meta_ads_keyboard(), parse_mode="HTML")
                return

    if user_text == "📁 Kontraktlar va Hisobotlar":
        user_data[chat_id] = {}
        send_safe_message(chat_id, "📁 <b>KONTRAKTLAR VA HISOBOTLAR BO'LIMI</b>\n\nKerakli xizmatni tanlang:", reply_markup=get_kontrakt_folder_keyboard())
        return

    if user_text in ["📁 Ma'lumotnomalar", "📁 Ma'lumotnomalar va Hujjatlar"]:
        user_data[chat_id] = {}
        send_safe_message(chat_id, "📁 <b>MA'LUMOTNOMALAR BO'LIMI</b>\n\nQaysi rasmiy ma'lumotnomani tayyorlamoqchisiz?", reply_markup=get_docs_folder_keyboard())
        return

    if user_text == "📁 Buyruqlar":
        user_data[chat_id] = {}
        send_safe_message(chat_id, "📁 <b>RASMIY BUYRUQLAR BO'LIMI</b>\n\nQaysi buyruq turini shakllantirmoqchisiz?", reply_markup=get_buyruqlar_folder_keyboard())
        return

    if user_text == "📊 Tizim Statistikasi":
        from services.atlas_db import get_dashboard_summary_data
        try:
            summary = get_dashboard_summary_data()
            k_stat = summary.get("kontrakt_stats", {})
            d_stat = summary.get("documents_stats", {})
            a_stat = summary.get("amaliyot_stats", {})

            stats_msg = (
                "📊 <b>ATLAS PLATFORMASI VA BOT STATISTIKASI</b>\n\n"
                "💰 <b>KONTRAKTLAR VA MOLIYA:</b>\n"
                f"• Jami talabalar: <b>{k_stat.get('total_students', 0)} nafar</b>\n"
                f"• Shartnoma summasi: <b>{k_stat.get('total_contract_sum', 0):,.0f} so'm</b>\n"
                f"• To'langan summa: <b>{k_stat.get('total_paid_sum', 0):,.0f} so'm</b>\n"
                f"• Qolgan qarzdorlik: <b>{k_stat.get('total_debt_sum', 0):,.0f} so'm</b>\n\n"
                "📄 <b>HUJJATLAR VA BUYRUQLAR:</b>\n"
                f"• Jami shakllantirilgan: <b>{d_stat.get('total_docs', 0)} ta</b>\n"
                f"• Oxirgi 7 kundagi hujjatlar: <b>{d_stat.get('recent_docs_count', 0)} ta</b>\n\n"
                "🏥 <b>MALAKAVIY AMALIYOT:</b>\n"
                f"• Jami yo'nalishlar: <b>{a_stat.get('total_directions', 0)} ta</b>\n"
                f"• So'rovnomadagi talabalar: <b>{a_stat.get('total_survey_students', 0)} nafar</b>\n"
                f"• Yaratilgan amaliyot buyruqlari: <b>{a_stat.get('total_orders', 0)} ta</b>\n\n"
                f"📅 <i>Ma'lumotlar yangilangan vaqti: {datetime.now().strftime('%d.%m.%Y %H:%M')}</i>"
            )
            send_safe_message(chat_id, stats_msg, reply_markup=get_main_keyboard())
        except Exception as e:
            send_safe_message(chat_id, f"❌ Statistika olishda xatolik: {str(e)}", reply_markup=get_main_keyboard())
        return

    if user_text == "📑 Namunaviy So'rovnoma Excel":
        from services.amaliyot_service import generate_sample_survey_excel
        try:
            excel_bytes = generate_sample_survey_excel()
            temp_sample = os.path.join(tempfile.gettempdir(), "Amaliyot So'rovnoma Namuna.xlsx")
            with open(temp_sample, "wb") as f_s:
                f_s.write(excel_bytes)

            caption = (
                "📑 <b>Namunaviy So'rovnoma Excel Fayli</b>\n\n"
                "Ushbu faylga talabalar guruhi, F.I.SH va amaliyot tumanlarini kiritib, botga yoki platformaga import qilishingiz mumkin."
            )
            with open(temp_sample, "rb") as f_s:
                bot.send_document(chat_id, document=f_s, caption=caption, parse_mode="HTML")
            if os.path.exists(temp_sample): os.remove(temp_sample)
        except Exception as e:
            send_safe_message(chat_id, f"❌ Namuna yaratishda xatolik: {str(e)}")
        return

    if user_text in ["📁 Yo'nalishlar & Buyruq Yaratish", "📋 Yo'nalishlar va Guruhlar Ro'yxati"]:
        from services.atlas_db import get_amaliyot_folders_hierarchy
        res = get_amaliyot_folders_hierarchy()
        folders = res.get("folders", [])
        if not folders:
            send_safe_message(chat_id, "ℹ️ Hozircha bazada amaliyot papkalari mavjud emas.", reply_markup=get_amaliyot_folder_keyboard())
            return

        markup = telebot.types.InlineKeyboardMarkup(row_width=1)
        for y_item in folders:
            for dir_item in y_item.get("children", []):
                dur = dir_item.get("extra_data", {}).get("duration", "")
                dur_txt = f" ({dur})" if dur else ""
                markup.add(telebot.types.InlineKeyboardButton(f"📚 {dir_item['name']}{dur_txt}", callback_data=f"am_dir:{dir_item['id']}"))

        send_safe_message(
            chat_id,
            "🏥 <b>MALAKAVIY AMALIYOT: YO'NALISHNI TANLANG:</b>\n\n"
            "Kerakli ta'lim yo'nalishini tanlang, so'ng semestr buyruqlarini yaratishingiz mumkin:",
            reply_markup=markup
        )
        return

    if user_text == "🏥 Malakaviy Amaliyot":
        user_data[chat_id] = {}
        msg_text = (
            "🏥 <b>MALAKAVIY AMALIYOT BUYRUQLARI VA REJALARI</b>\n\n"
            "✨ Ushbu bo'lim orqali barcha ta'lim yo'nalishlari, semestrlar va talabalar amaliyot buyruqlarini boshqarishingiz mumkin.\n\n"
            "Kerakli xizmatni tanlang:"
        )
        send_safe_message(chat_id, msg_text, reply_markup=get_amaliyot_folder_keyboard())
        return

    if user_text == "📥 Oxirgi Buyruqlar Arxivini Ko'rish":
        from services.atlas_db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM generated_docs ORDER BY id DESC LIMIT 5")
        recent_docs = [dict(r) for r in cursor.fetchall()]
        conn.close()

        if not recent_docs:
            send_safe_message(chat_id, "ℹ️ Hozircha shakllantirilgan buyruqlar arxivi bo'sh.", reply_markup=get_amaliyot_folder_keyboard())
            return

        send_safe_message(chat_id, f"📥 <b>OXIRGI SHAKLLANTIRILGAN {len(recent_docs)} TA HUJJAT:</b>\n\nFayllar yuborilmoqda...")
        for d in recent_docs:
            fpath = d.get("file_path", "")
            if fpath and os.path.exists(fpath):
                fio = d.get("recipient_fio") or "Talaba"
                tpl_name = d.get("template_name") or "Hujjat"
                caption = f"📄 <b>{tpl_name}</b>\n👤 <b>Talaba / Qabul qiluvchi:</b> {fio}\n📅 <b>Vaqti:</b> {d.get('created_at', '')}"
                with open(fpath, "rb") as f_obj:
                    ext = os.path.splitext(fpath)[1].lower()
                    if ext in [".png", ".jpg", ".jpeg"]:
                        bot.send_photo(chat_id, photo=f_obj, caption=caption, parse_mode="HTML")
                    else:
                        bot.send_document(chat_id, document=f_obj, caption=caption, parse_mode="HTML")
        return

    if user_text == "🌐 ATLAS Web Platformasi Linki":
        markup = telebot.types.InlineKeyboardMarkup()
        btn_web = telebot.types.InlineKeyboardButton("🌐 ATLAS Web Platformasiga O'tish", url="https://atlas-my-tools.vercel.app")
        markup.add(btn_web)
        send_safe_message(
            chat_id,
            "🌐 <b>ATLAS Universal Platformasi:</b>\n\n"
            "Kompyuter yoki telefon brauzeridan barcha amaliyot buyruqlari, kontraktlar, ma'lumotnomalar va tahlillarni boshqarish uchun pastdagi tugmani bosing:",
            reply_markup=markup
        )
        return

    if user_text == "🔙 Asosiy menyuga qaytish":
        user_data[chat_id] = {}
        send_safe_message(chat_id, "📂 <b>ASOSIY KATEGORIYALAR MENYUSI</b>\n\nKerakli bo'lim papkasini tanlang:", reply_markup=get_main_keyboard())
        return

    # Docbot shablonlarini tanlashni tekshirish
    for idx, tpl in enumerate(DOCBOT_TEMPLATES):
        if user_text == tpl["name"].strip():
            start_docbot_wizard(chat_id, idx)
            return

    # Docbot savol-javob zanjiri (FSM)
    if holat == "docbot_step":
        u_info = user_data[chat_id]
        tpl_index = u_info["tpl_index"]
        step = u_info["step"]
        answers = u_info["answers"]
        tpl = DOCBOT_TEMPLATES[tpl_index]
        current_step_info = tpl["steps"][step]

        answers[current_step_info["field"]] = user_text
        step += 1

        if step < len(tpl["steps"]):
            next_step_info = tpl["steps"][step]
            u_info["step"] = step
            u_info["answers"] = answers
            markup = make_step_keyboard(next_step_info.get("buttons"))
            send_safe_message(
                chat_id,
                f"<b>({step + 1}/{len(tpl['steps'])})</b> {next_step_info['question']}",
                reply_markup=markup
            )
        else:
            user_data[chat_id] = {}
            process_docbot_generation(chat_id, tpl, answers)
        return

    if holat == "sana_kutish":
        sana_matni = user_text.replace("📅", "").strip()

        try:
            cheklov_sanasi = datetime.strptime(sana_matni, "%d.%m.%Y")
            user_data[chat_id]["sana"] = cheklov_sanasi
            
            baza_path = user_data[chat_id].get("baza_path")
            deb_path = user_data[chat_id].get("deb_path")

            if not baza_path or not os.path.exists(baza_path) or not deb_path or not os.path.exists(deb_path):
                send_safe_message(chat_id, "❌ Yuklangan fayllar topilmadi. Qayta urinib ko'ring.")
                user_data[chat_id] = {}
                return

            process_kontrakt_update(chat_id, baza_path, deb_path, cheklov_sanasi)
            user_data[chat_id] = {}

        except ValueError:
            suggested = user_data.get(chat_id, {}).get("taklif_sana_str")
            markup = get_main_keyboard(suggested)
            send_safe_message(chat_id, "❌ Noto'g'ri sana formati. Nuqtalar bilan kiriting (Masalan: <code>27.06.2026</code>):", reply_markup=markup)

# ============================================================
# MALAKAVIY AMALIYOT CALLBACK QUERY HANDLER (INLINE BUTTONS)
# ============================================================
@bot.callback_query_handler(func=lambda call: call.data.startswith("am_"))
def handle_amaliyot_callbacks(call):
    chat_id = call.message.chat.id
    if not is_user_allowed(call.message):
        bot.answer_callback_query(call.id, "Ruxsat etilmagan!")
        return

    data = call.data

    try:
        from services.atlas_db import get_amaliyot_folder, get_amaliyot_surveys, get_amaliyot_folder_path, get_db_connection, save_amaliyot_surveys
        from services.amaliyot_service import generate_all_district_orders, fill_amaliyot_template, find_matching_amaliyot_template, DISTRICT_DOCTORS

        if data.startswith("am_dir:"):
            dir_id = int(data.split(":")[1])
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM amaliyot_folders WHERE parent_id = ? ORDER BY order_num ASC, id ASC", (dir_id,))
            groups = [dict(r) for r in cur.fetchall()]
            conn.close()

            dir_info = get_amaliyot_folder(dir_id) or {}
            markup = telebot.types.InlineKeyboardMarkup(row_width=1)
            for g in groups:
                markup.add(telebot.types.InlineKeyboardButton(f"👥 {g['name']}", callback_data=f"am_grp:{g['id']}"))
            markup.add(telebot.types.InlineKeyboardButton("🔙 Yo'nalishlarga qaytish", callback_data="am_back_dirs"))

            bot.edit_message_text(
                f"📚 <b>{dir_info.get('name', 'Yo‘nalish')}</b>\n\nGuruhlar to'plamini tanlang:",
                chat_id=chat_id,
                message_id=call.message.message_id,
                parse_mode="HTML",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id)
            return

        elif data == "am_back_dirs":
            from services.atlas_db import get_amaliyot_folders_hierarchy
            res = get_amaliyot_folders_hierarchy()
            folders = res.get("folders", [])
            markup = telebot.types.InlineKeyboardMarkup(row_width=1)
            for y_item in folders:
                for dir_item in y_item.get("children", []):
                    dur = dir_item.get("extra_data", {}).get("duration", "")
                    dur_txt = f" ({dur})" if dur else ""
                    markup.add(telebot.types.InlineKeyboardButton(f"📚 {dir_item['name']}{dur_txt}", callback_data=f"am_dir:{dir_item['id']}"))

            bot.edit_message_text(
                "🏥 <b>MALAKAVIY AMALIYOT: YO'NALISHNI TANLANG:</b>\n\nKerakli ta'lim yo'nalishini tanlang:",
                chat_id=chat_id,
                message_id=call.message.message_id,
                parse_mode="HTML",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id)
            return

        elif data.startswith("am_grp:"):
            grp_id = int(data.split(":")[1])
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT * FROM amaliyot_folders WHERE parent_id = ? ORDER BY order_num ASC, id ASC", (grp_id,))
            semesters = [dict(r) for r in cur.fetchall()]
            conn.close()

            grp_info = get_amaliyot_folder(grp_id) or {}
            markup = telebot.types.InlineKeyboardMarkup(row_width=1)
            for s in semesters:
                markup.add(telebot.types.InlineKeyboardButton(f"🔖 {s['name']}", callback_data=f"am_sem:{s['id']}"))
            markup.add(telebot.types.InlineKeyboardButton("🔙 Orqaga", callback_data=f"am_dir:{grp_info.get('parent_id', 1)}"))

            bot.edit_message_text(
                f"👥 <b>{grp_info.get('name', 'Guruhlar')}</b>\n\nSemestrni tanlang:",
                chat_id=chat_id,
                message_id=call.message.message_id,
                parse_mode="HTML",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id)
            return

        elif data.startswith("am_sem:"):
            sem_id = int(data.split(":")[1])
            sem_info = get_amaliyot_folder(sem_id) or {}
            surveys_res = get_amaliyot_surveys(sem_id)
            students = surveys_res.get("surveys", [])
            extra = sem_info.get("extra_data", {})

            dist_counts = {}
            for st in students:
                t = st.get("tumani", "").strip() or "Shahrisabz shahar"
                dist_counts[t] = dist_counts.get(t, 0) + 1

            text = (
                f"🔖 <b>{sem_info.get('name')}</b>\n\n"
                f"👥 <b>Talabalar soni:</b> {len(students)} nafar\n"
                f"📅 <b>Muddat:</b> {extra.get('start_date', '08.06.2026')} — {extra.get('end_date', '06.07.2026')}\n"
                f"🏛️ <b>Tumanlar:</b> {', '.join([f'{k} ({v})' for k, v in dist_counts.items()]) if dist_counts else 'Hali talabalar yuklanmagan'}\n\n"
                f"Quyidagi amallardan birini tanlang:"
            )

            markup = telebot.types.InlineKeyboardMarkup(row_width=1)
            if students:
                markup.add(telebot.types.InlineKeyboardButton("⚡ Barcha Tumanlar Buyruqlari (ZIP yuklab olish)", callback_data=f"am_gen_zip:{sem_id}"))
                markup.add(telebot.types.InlineKeyboardButton("📄 Tumanlar Bo'yicha Alohida Word Olish", callback_data=f"am_dist_menu:{sem_id}"))
            markup.add(telebot.types.InlineKeyboardButton("📤 So'rovnoma Excel Faylini Yuklash", callback_data=f"am_upload_excel:{sem_id}"))
            markup.add(telebot.types.InlineKeyboardButton("🔙 Guruhlarga qaytish", callback_data=f"am_grp:{sem_info.get('parent_id', 1)}"))

            bot.edit_message_text(
                text,
                chat_id=chat_id,
                message_id=call.message.message_id,
                parse_mode="HTML",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id)
            return

        elif data.startswith("am_gen_zip:"):
            sem_id = int(data.split(":")[1])
            bot.answer_callback_query(call.id, "ZIP paket shakllantirilmoqda...")
            
            surveys_res = get_amaliyot_surveys(sem_id)
            students = surveys_res.get("surveys", [])
            if not students:
                send_safe_message(chat_id, "❌ Ushbu semestrda talabalar topilmadi.")
                return

            sem_info = get_amaliyot_folder(sem_id) or {}
            extra = sem_info.get("extra_data", {})
            path_res = get_amaliyot_folder_path(sem_id)
            folder_path = path_res.get("path", [])

            direction_name = folder_path[1]["name"] if len(folder_path) > 1 else ""
            duration = extra.get("duration") or (folder_path[1].get("extra_data", {}).get("duration", "") if len(folder_path) > 1 else "")
            semester_name = sem_info.get("name", "")
            custom_tpl = extra.get("template_file")

            tpl_path = find_matching_amaliyot_template(direction_name, duration, semester_name, custom_tpl)

            start_date = extra.get("start_date", "08.06.2026")
            end_date = extra.get("end_date", "06.07.2026")
            amaliyot_muddati = extra.get("amaliyot_muddati", "")

            semester_data = {
                "buyruq_raqami": "____",
                "buyruq_sanasi": datetime.now().strftime("%d.%m.%Y"),
                "oquv_yili": extra.get("oquv_yili", "2025/2026"),
                "kursi": str(extra.get("kursi", "1")),
                "amaliyot_muddati": amaliyot_muddati,
                "start_date": start_date,
                "end_date": end_date
            }

            output_dir = os.path.join(tempfile.gettempdir(), f"amaliyot_batch_{uuid.uuid4().hex[:8]}")
            res = generate_all_district_orders(tpl_path, semester_data, students, output_dir)

            caption = (
                f"📦 <b>Malakaviy Amaliyot Buyruqlari (ZIP)</b>\n\n"
                f"🔖 <b>Semestr:</b> {semester_name}\n"
                f"👥 <b>Jami talabalar:</b> {res['total_students']} ta\n"
                f"🏛️ <b>Jami tumanlar:</b> {res['total_districts']} ta"
            )
            with open(res["zip_path"], "rb") as z_f:
                bot.send_document(chat_id, document=z_f, caption=caption, parse_mode="HTML")
            return

        elif data.startswith("am_dist_menu:"):
            sem_id = int(data.split(":")[1])
            surveys_res = get_amaliyot_surveys(sem_id)
            students = surveys_res.get("surveys", [])
            districts = sorted(list(set(s.get("tumani", "").strip() or "Shahrisabz shahar" for s in students)))

            markup = telebot.types.InlineKeyboardMarkup(row_width=1)
            for d in districts:
                d_cnt = len([s for s in students if (s.get("tumani", "").strip() or "Shahrisabz shahar") == d])
                markup.add(telebot.types.InlineKeyboardButton(f"📄 {d} ({d_cnt} ta talaba)", callback_data=f"am_gen_dist:{sem_id}:{d}"))
            markup.add(telebot.types.InlineKeyboardButton("🔙 Semestrga qaytish", callback_data=f"am_sem:{sem_id}"))

            bot.edit_message_text(
                "🏛️ <b>Qaysi tuman buyrug'ini yuklab olmoqchisiz?</b>",
                chat_id=chat_id,
                message_id=call.message.message_id,
                parse_mode="HTML",
                reply_markup=markup
            )
            bot.answer_callback_query(call.id)
            return

        elif data.startswith("am_gen_dist:"):
            parts = data.split(":", 2)
            sem_id = int(parts[1])
            tumani = parts[2]
            bot.answer_callback_query(call.id, f"{tumani} buyrug'i tayyorlanmoqda...")

            surveys_res = get_amaliyot_surveys(sem_id)
            students = [s for s in surveys_res.get("surveys", []) if (s.get("tumani", "").strip() or "Shahrisabz shahar") == tumani]

            sem_info = get_amaliyot_folder(sem_id) or {}
            extra = sem_info.get("extra_data", {})
            path_res = get_amaliyot_folder_path(sem_id)
            folder_path = path_res.get("path", [])

            direction_name = folder_path[1]["name"] if len(folder_path) > 1 else ""
            duration = extra.get("duration") or (folder_path[1].get("extra_data", {}).get("duration", "") if len(folder_path) > 1 else "")
            semester_name = sem_info.get("name", "")
            custom_tpl = extra.get("template_file")

            tpl_path = find_matching_amaliyot_template(direction_name, duration, semester_name, custom_tpl)

            start_date = extra.get("start_date", "08.06.2026")
            end_date = extra.get("end_date", "06.07.2026")
            dist_groups = sorted(list(set(s.get("guruhi", "").strip() for s in students if s.get("guruhi", "").strip())))

            order_data = {
                "buyruq_raqami": "____",
                "buyruq_sanasi": datetime.now().strftime("%d.%m.%Y"),
                "tumani": tumani,
                "shu_tuman_shifokori": DISTRICT_DOCTORS.get(tumani, "Bosh shifokor"),
                "oquv_yili": extra.get("oquv_yili", "2025/2026"),
                "kursi": str(extra.get("kursi", "1")),
                "guruhlar": dist_groups,
                "amaliyot_muddati": extra.get("amaliyot_muddati", ""),
                "start_date": start_date,
                "end_date": end_date,
                "students": students
            }

            clean_tumani = re.sub(r'[\\/*?:"<>|]', "", tumani).strip()
            clean_grp = re.sub(r'[\\/*?:"<>|]', "", ", ".join(dist_groups) if dist_groups else "Guruh").strip()
            docx_filename = f"{clean_tumani} — {clean_grp} — {len(students)} ta talaba.docx"
            temp_docx = os.path.join(tempfile.gettempdir(), docx_filename)
            fill_amaliyot_template(tpl_path, order_data, temp_docx)

            caption = (
                f"📄 <b>Amaliyot Buyrug'i: {tumani}</b>\n\n"
                f"🔖 <b>Semestr:</b> {semester_name}\n"
                f"👥 <b>Talabalar:</b> {len(students)} nafar\n"
                f"📅 <b>Muddat:</b> {start_date} — {end_date}"
            )
            with open(temp_docx, "rb") as d_f:
                bot.send_document(chat_id, document=d_f, caption=caption, parse_mode="HTML")
            if os.path.exists(temp_docx): os.remove(temp_docx)
            return

        elif data.startswith("am_upload_excel:"):
            sem_id = int(data.split(":")[1])
            user_data[chat_id] = {
                "holat": "amaliyot_excel_kutish",
                "folder_id": sem_id
            }
            bot.answer_callback_query(call.id)
            send_safe_message(
                chat_id,
                "📤 <b>SO'ROVNOMA EXCEL FAYLINI YUBORING:</b>\n\n"
                "Talabalar ro'yxati yozilgan `.xlsx` faylni botga yuboring. Tizim uni avtomatik o'qib, ushbu semestrga saqlaydi."
            )
            return

    except Exception as e:
        bot.answer_callback_query(call.id, f"Xatolik: {str(e)}")
        send_safe_message(chat_id, f"❌ Xatolik yuz berdi: {str(e)}")


# Fayl yuklanganda ishlovchi handler
@bot.message_handler(content_types=['document'])
def handle_docs(message):
    chat_id = message.chat.id
    if not is_user_allowed(message):
        send_access_denied(chat_id, message.from_user.id)
        return
    save_user_chat_id(chat_id)
    holat = user_data.get(chat_id, {}).get("holat")

    if not holat:
        send_safe_message(chat_id, "Iltimos, avval menyudan kerakli tugmani tanlang:")
        return

    # AMALIYOT SO'ROVNOMA EXCEL YUKLASH
    if holat == "amaliyot_excel_kutish":
        folder_id = user_data[chat_id].get("folder_id")
        try:
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)

            from services.amaliyot_service import parse_survey_excel
            from services.atlas_db import save_amaliyot_surveys, get_amaliyot_folder

            parsed_students = parse_survey_excel(downloaded_file)
            if not parsed_students:
                send_safe_message(chat_id, "❌ Excel faylidan talabalar topilmadi. Fayl ustunlarini tekshiring.")
                user_data[chat_id] = {}
                return

            res = save_amaliyot_surveys(folder_id, parsed_students, replace_all=True)
            sem_info = get_amaliyot_folder(folder_id) or {}

            send_safe_message(
                chat_id,
                f"✅ <b>{sem_info.get('name', 'Semestr')}</b> uchun <b>{len(parsed_students)} ta talaba</b> muvaffaqiyatli qabul qilindi va Supabase bazasiga saqlandi!\n\n"
                f"Endi bemalol <b>«📁 Yo'nalishlar & Buyruq Yaratish»</b> orqali buyruqlarni yuklab olishingiz mumkin.",
                reply_markup=get_amaliyot_folder_keyboard()
            )
            user_data[chat_id] = {}
            return
        except Exception as e:
            send_safe_message(chat_id, f"❌ Excel faylni o'qishda xatolik: {str(e)}")
            user_data[chat_id] = {}
            return

    # GURUH SCREENSHOTLARINI OLISH
    if holat == "guruh_fayl_kutish":
        process_group_screenshots(chat_id, message)
        return

    # KONTRAKTNI YANGILASH: 1-BOSQICH
    if holat == "baza_kutish":
        try:
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            
            baza_nomi = os.path.join(tempfile.gettempdir(), f"baza_{chat_id}.xlsx")
            with open(baza_nomi, 'wb') as f:
                f.write(downloaded_file)

            taklif_sana_dt = None
            try:
                wb_check = openpyxl.load_workbook(baza_nomi, data_only=True)
                sheet_check = wb_check.active
                for r in range(1, 30):
                    for c in range(1, 10):
                        val = str(sheet_check.cell(row=r, column=c).value or "")
                        if 'yangilangan sanasi' in val.lower():
                            cell_val = sheet_check.cell(row=r, column=c+1).value or sheet_check.cell(row=r, column=c+2).value
                            if cell_val:
                                if isinstance(cell_val, datetime):
                                    taklif_sana_dt = cell_val + timedelta(days=1)
                                elif isinstance(cell_val, str):
                                    try:
                                        dt = datetime.strptime(cell_val.strip(), "%d.%m.%Y")
                                        taklif_sana_dt = dt + timedelta(days=1)
                                    except ValueError:
                                        pass
                wb_check.close()
            except Exception:
                pass

            if not taklif_sana_dt and message.document.file_name:
                match = re.search(r'(\d{2}\.\d{2}\.\d{4})', message.document.file_name)
                if match:
                    try:
                        dt = datetime.strptime(match.group(1), "%d.%m.%Y")
                        taklif_sana_dt = dt + timedelta(days=1)
                    except ValueError:
                        pass

            taklif_sana_str = taklif_sana_dt.strftime("%d.%m.%Y") if taklif_sana_dt else None

            user_data[chat_id]["baza_path"] = baza_nomi
            user_data[chat_id]["taklif_sana_str"] = taklif_sana_str
            user_data[chat_id]["holat"] = "deb_kutish"

            send_safe_message(chat_id, "✅ <b>Asosiy baza qabul qilindi!</b>\n\n"
                                       "📥 <b>2-BOSQICH: BANK DEBITORKASINI YUKLASH</b>\n\n"
                                       "Endi bankdan kelgan yangi <b>Debitorka (.xlsx)</b> faylini yuboring:")
        except Exception as e:
            send_safe_message(chat_id, f"❌ Faylni yuklashda xatolik: <code>{escape_html_text(str(e))}</code>")
            user_data[chat_id] = {}
        return

    # KONTRAKTNI YANGILASH: 2-BOSQICH
    if holat == "deb_kutish":
        try:
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            
            deb_nomi = os.path.join(tempfile.gettempdir(), f"deb_{chat_id}.xlsx")
            with open(deb_nomi, 'wb') as f:
                f.write(downloaded_file)

            user_data[chat_id]["deb_path"] = deb_nomi
            user_data[chat_id]["holat"] = "sana_kutish"

            taklif_sana_str = user_data[chat_id].get("taklif_sana_str")
            
            msg_text = "✅ <b>Bank debitorkasi qabul qilindi!</b>\n\n" \
                       "📅 <b>3-BOSQICH: BOSHLANISH SANASI</b>\n\n"
            
            if taklif_sana_str:
                msg_text += f"💡 Bazangizdagi oxirgi sana bo'yicha tavsiya etilgan boshlanish sanasi: <code>{taklif_sana_str}</code>\n\n"
            
            msg_text += "To'lovlarni <b>qaysi sanadan boshlab</b> hisoblayin?\n"
            if taklif_sana_str:
                msg_text += f"<i>(Pastdagi <code>{taklif_sana_str}</code> tugmasini bosing yoki o'zingiz sana kiriting):</i>"
            else:
                msg_text += "Format: <code>27.06.2026</code> shaklida yozing:"

            markup = get_main_keyboard(suggested_date=taklif_sana_str)
            send_safe_message(chat_id, msg_text, reply_markup=markup)

        except Exception as e:
            send_safe_message(chat_id, f"❌ Faylni yuklashda xatolik: <code>{escape_html_text(str(e))}</code>")
            user_data[chat_id] = {}
        return

def process_kontrakt_update(chat_id, baza_path, deb_nomi, cheklov_sanasi):
    """Kontraktlarni yangilash, hisobot va XULOSA rasm tayyorlash"""
    progress = TelegramProgress(bot, chat_id, "⏳ Jarayon boshlandi...")

    try:
        progress.update("📥 Ma'lumotlar olinmoqda...", 10)
        
        wb_baza_write = openpyxl.load_workbook(baza_path, data_only=False)
        wb_baza_read = openpyxl.load_workbook(baza_path, data_only=True)
        wb_deb = openpyxl.load_workbook(deb_nomi, data_only=True)

        varoq_nomi = 'KONTRAKTLAR' if 'KONTRAKTLAR' in wb_baza_write.sheetnames else wb_baza_write.sheetnames[0]
        sheet_write = wb_baza_write[varoq_nomi]
        sheet_read = wb_baza_read[varoq_nomi]
        sheet_deb = wb_deb['bank'] if 'bank' in wb_deb.sheetnames else wb_deb.active

        progress.update("🧠 Ma'lumotlar tahlil qilinmoqda...", 25)

        ism_ustun = 3
        kerak_ustun = 4
        tolov_ustun = 5
        guruh_ustun = 1
        boshlanish_row = 23

        for r in range(1, 30):
            for c in range(1, 15):
                val = str(sheet_read.cell(row=r, column=c).value or "").lower()
                if 'guruh' in val and 'rahbar' not in val and 'soni' not in val:
                    guruh_ustun = c
                if any(x in val for x in ['familiya', 'f.i.sh', 'ism', 'sharfi']):
                    ism_ustun = c
                    boshlanish_row = r + 1
                if 'bo\'lishi' in val or 'kerak' in val:
                    kerak_ustun = c
                if any(x in val for x in ['jami', 'to\'lagan summasi', 'to\'lov']):
                    tolov_ustun = c

        baza_talabalari = []
        for row in range(boshlanish_row, sheet_read.max_row + 1):
            fio = sheet_read.cell(row=row, column=ism_ustun).value
            if fio and str(fio).strip() and not str(fio).lower().startswith(('familiya', 'f.i.sh', 'итого', 'jami', 'guruh')):
                guruh_val = sheet_read.cell(row=row, column=guruh_ustun).value
                guruh_str = str(guruh_val).strip() if guruh_val else "Noma'lum"
                if guruh_str.endswith('.0'): guruh_str = guruh_str[:-2]

                eski_val = sheet_read.cell(row=row, column=tolov_ustun).value
                kerak_val = sheet_read.cell(row=row, column=kerak_ustun).value
                try: eski_sum = float(eski_val) if eski_val else 0.0
                except (ValueError, TypeError): eski_sum = 0.0

                try: kerak_sum = float(kerak_val) if kerak_val else 0.0
                except (ValueError, TypeError): kerak_sum = 0.0

                baza_talabalari.append({
                    "row": row,
                    "original_name": str(fio).strip(),
                    "clean_name": ismlarni_standartlash(fio),
                    "guruh": guruh_str,
                    "kerak_summa": kerak_sum,
                    "boshlangich_summa": eski_sum,
                    "joriy_summa": eski_sum
                })

        yangilanish_tarixi = []
        topilmaganlar = []
        jami_tushgan_pul = 0.0
        oxirgi_to_lov_sanasi = None
        yangilangan_talabalar_set = set()
        
        max_deb_rows = sheet_deb.max_row
        for row_idx, row in enumerate(range(2, max_deb_rows + 1), start=1):
            current_pct = 25 + int((row_idx / max(max_deb_rows - 1, 1)) * 50)
            progress.update("🧠 Ma'lumotlar tahlil qilinmoqda...", current_pct)

            sana_val = sheet_deb.cell(row=row, column=1).value
            if not sana_val: continue

            try:
                if isinstance(sana_val, str):
                    if '.' in sana_val:
                        to_lov_sanasi = datetime.strptime(sana_val.strip(), "%d.%m.%y")
                    else:
                        to_lov_sanasi = datetime.strptime(sana_val.strip(), "%Y-%m-%d %H:%M:%S")
                elif isinstance(sana_val, datetime):
                    to_lov_sanasi = sana_val
                else:
                    continue
            except ValueError:
                continue

            if to_lov_sanasi >= cheklov_sanasi:
                summa_val = sheet_deb.cell(row=row, column=7).value
                h_val = sheet_deb.cell(row=row, column=8).value
                deb_fio = sheet_deb.cell(row=row, column=9).value

                deb_fio_str = str(deb_fio).strip() if deb_fio else ""
                h_str = str(h_val).strip() if h_val else ""

                if not deb_fio_str and not h_str: continue
                if not summa_val: continue

                try:
                    yangi_summa = float(summa_val)
                except (ValueError, TypeError):
                    continue

                if oxirgi_to_lov_sanasi is None or to_lov_sanasi > oxirgi_to_lov_sanasi:
                    oxirgi_to_lov_sanasi = to_lov_sanasi

                deb_fio_clean = ismlarni_standartlash(deb_fio_str or h_str)
                eng_yaxshi_moslik = None
                eng_yuqori_ball = 0

                for talaba in baza_talabalari:
                    s_set = fuzz.token_set_ratio(deb_fio_clean, talaba["clean_name"])
                    s_partial = fuzz.partial_ratio(talaba["clean_name"], deb_fio_clean)
                    s_sort = fuzz.token_sort_ratio(deb_fio_clean, talaba["clean_name"])
                    ball = max(s_set, s_partial, s_sort)

                    if ball > eng_yuqori_ball:
                        eng_yuqori_ball = ball
                        eng_yaxshi_moslik = talaba

                if eng_yaxshi_moslik and eng_yuqori_ball >= 70:
                    target_row = eng_yaxshi_moslik["row"]
                    
                    eski_summa = eng_yaxshi_moslik["joriy_summa"]
                    jami_yangi = eski_summa + yangi_summa
                    
                    sheet_write.cell(row=target_row, column=tolov_ustun).value = jami_yangi
                    eng_yaxshi_moslik["joriy_summa"] = jami_yangi
                    
                    jami_tushgan_pul += yangi_summa
                    yangilangan_talabalar_set.add(eng_yaxshi_moslik['original_name'])

                    safe_orig_name = escape_html_text(eng_yaxshi_moslik['original_name'])
                    safe_guruh = escape_html_text(eng_yaxshi_moslik['guruh'])
                    safe_deb_fio = escape_html_text(deb_fio_str or h_str)
                    safe_sana = to_lov_sanasi.strftime('%d.%m.%Y')
                    
                    yangilanish_tarixi.append(
                        f"<blockquote>👤 <b>{safe_orig_name}</b>\n"
                        f"├ 🏦 Debitorkada: <code>{safe_deb_fio}</code>\n"
                        f"├ 🏫 Guruh: <code>{safe_guruh}</code>\n"
                        f"├ 📅 Toʻlov sanasi: <code>{safe_sana}</code>\n"
                        f"├ ➕ Tushgan pul: <code>{yangi_summa:,.0f} so'm</code>\n"
                        f"└ 📊 Jami toʻladi: <code>{jami_yangi:,.0f} so'm</code></blockquote>"
                    )
                else:
                    if h_str and deb_fio_str and h_str != deb_fio_str and deb_fio_str != "?":
                        disp_name = f"{deb_fio_str} | {h_str}"
                    elif h_str:
                        disp_name = h_str
                    else:
                        disp_name = deb_fio_str or "Noma'lum"

                    safe_disp_name = escape_html_text(disp_name)
                    safe_sana = to_lov_sanasi.strftime('%d.%m.%Y')
                    topilmaganlar.append(f"<blockquote>❓ <code>{safe_disp_name}</code> — <code>{yangi_summa:,.0f} so'm</code> (Sana: {safe_sana})</blockquote>")

        progress.update("⚙️ Natija tayyorlanmoqda...", 85)

        oxirgi_sana_str = oxirgi_to_lov_sanasi.strftime('%d.%m.%Y') if oxirgi_to_lov_sanasi else cheklov_sanasi.strftime('%d.%m.%Y')
        for r in range(1, 30):
            for c in range(1, 10):
                val = str(sheet_write.cell(row=r, column=c).value or "")
                if 'yangilangan sanasi' in val.lower():
                    target_c = c + 1
                    if not sheet_write.cell(row=r, column=target_c).value:
                        target_c = c + 2
                    sheet_write.cell(row=r, column=target_c).value = oxirgi_sana_str

        # Guruhlar bo'yicha umumiy XULOSA ma'lumotlarini bazadan avtomatik shakllantirish
        xulosa_rows = []
        for r in range(1, 20):
            rahbar = sheet_read.cell(row=r, column=3).value
            guruh = sheet_read.cell(row=r, column=4).value
            if rahbar and guruh and str(rahbar).strip() and str(guruh).strip():
                if str(rahbar).lower().startswith(('jami', 'итого', 'guruh rahbari')): continue
                g_str = str(guruh).strip()
                if g_str.endswith('.0'): g_str = g_str[:-2]

                g_students = [t for t in baza_talabalari if t['guruh'] == g_str]
                soni = len(g_students)
                qarz_sum = sum(max(0.0, t['kerak_summa'] - t['joriy_summa']) for t in g_students)

                xulosa_rows.append({
                    'rahbar': str(rahbar).strip(),
                    'guruh': g_str,
                    'soni': soni if soni > 0 else int(sheet_read.cell(row=r, column=5).value or 0),
                    'qarz': qarz_sum
                })

        natija_nomi = os.path.join(tempfile.gettempdir(), f"{oxirgi_sana_str}_GACHA_KONTRAKTLAR.xlsx")
        wb_baza_write.save(natija_nomi)
        wb_baza_write.close()
        wb_baza_read.close()
        wb_deb.close()

        progress.update("🔍 Yakuniy tekshiruv...", 95)

        keyingi_sana_dt = (oxirgi_to_lov_sanasi or cheklov_sanasi) + timedelta(days=1)
        keyingi_sana_str = keyingi_sana_dt.strftime('%d.%m.%Y')

        hisobot_matni = f"📊 <b>KONTRAKT YANGILANISH HISOBOTI</b>\n"
        hisobot_matni += f"📅 Filtr sanasi: {cheklov_sanasi.strftime('%d.%m.%Y')} dan {oxirgi_sana_str} gacha\n"
        hisobot_matni += f"━━━━━━━━━━━━━━━━━━━━\n"
        hisobot_matni += f"💰 <b>Jami tushgan pul:</b> <code>{jami_tushgan_pul:,.0f} so'm</code>\n"
        hisobot_matni += f"👥 <b>Muvaffaqiyatli yangilandi:</b> {len(yangilangan_talabalar_set)} kishi\n"
        hisobot_matni += f"━━━━━━━━━━━━━━━━━━━━\n\n"
        
        if yangilanish_tarixi:
            hisobot_matni += f"✅ <b>Yangilangan talabalar ({len(yangilanish_tarixi)} ta):</b>\n\n"
            for t in yangilanish_tarixi:
                hisobot_matni += t + "\n\n"
        else:
            hisobot_matni += "❌ Yangi to'lovlar topilmadi.\n\n"

        if topilmaganlar:
            hisobot_matni += f"❓ <b>Umuman topilmagan ismlar ({len(topilmaganlar)} ta):</b>\n"
            for top in topilmaganlar:
                hisobot_matni += top + "\n"
            hisobot_matni += "\n"

        hisobot_matni += f"━━━━━━━━━━━━━━━━━━━━\n"
        hisobot_matni += f"💡 <b>Keyingi safar adashmasligingiz uchun eslatma:</b>\n"
        hisobot_matni += f"Navbatdagi debitorkani yuklaganingizda botga boshlanish sanasi sifatida <code>{keyingi_sana_str}</code> sanasini kiriting."

        with open(natija_nomi, 'rb') as f_send:
            bot.send_document(chat_id, f_send, caption=f"📄 Formulalari buzilmagan tayyor Excel faylingiz: <code>{os.path.basename(natija_nomi)}</code>", parse_mode="HTML", reply_markup=get_main_keyboard())
        
        send_safe_message(chat_id, hisobot_matni)

        # ENG OXIRIDA GURUHLAR BO'YICHA XULOSA RASM JADVALINI YUBORISH
        if xulosa_rows:
            xulosa_img = os.path.join(tempfile.gettempdir(), f"xulosa_{chat_id}.png")
            generate_xulosa_table_image(xulosa_rows, xulosa_img)
            with open(xulosa_img, 'rb') as xf:
                bot.send_photo(chat_id, photo=xf, caption=f"📊 <b>Guruh rahbarlari bo'yicha umumiy XULOSA hisoboti (v{BOT_VERSION})</b>", parse_mode="HTML")
            if os.path.exists(xulosa_img): os.remove(xulosa_img)

        progress.success("✅ Tayyor!")

        if os.path.exists(baza_path): os.remove(baza_path)
        if os.path.exists(deb_nomi): os.remove(deb_nomi)
        if os.path.exists(natija_nomi): os.remove(natija_nomi)

    except Exception as e:
        progress.error(f"❌ Jarayonni bajarishda xatolik yuz berdi:\n<code>{escape_html_text(str(e))}</code>")
        if os.path.exists(baza_path): os.remove(baza_path)
        if os.path.exists(deb_nomi): os.remove(deb_nomi)
        if os.path.exists(natija_nomi): os.remove(natija_nomi)

def process_group_screenshots(chat_id, message):
    """Guruhlar screenshotlarini olish jarayoni (Real Progress bilan)"""
    progress = TelegramProgress(bot, chat_id, "⏳ Jarayon boshlandi...")
    temp_excel = os.path.join(tempfile.gettempdir(), f"temp_guruh_{chat_id}.xlsx")

    try:
        progress.update("📥 Ma'lumotlar olinmoqda...", 10)

        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        with open(temp_excel, 'wb') as f:
            f.write(downloaded_file)

        wb = openpyxl.load_workbook(temp_excel, data_only=True)
        sheet = wb.active

        progress.update("🧠 Guruhlar tahlil qilinmoqda...", 25)

        header_row = 22
        date_str = datetime.now().strftime("%d.%m.%Y")

        for r in range(1, 30):
            val_a = str(sheet.cell(row=r, column=1).value or "").lower()
            val_c = str(sheet.cell(row=r, column=3).value or "").lower()
            if "guruh" in val_a or "familiy" in val_c:
                header_row = r
                break
            for c in range(1, 10):
                cell_val = str(sheet.cell(row=r, column=c).value or "")
                if "yangilangan sanasi" in cell_val.lower():
                    next_cell = sheet.cell(row=r, column=c+1).value or sheet.cell(row=r, column=c+2).value
                    if next_cell:
                        if isinstance(next_cell, datetime):
                            date_str = next_cell.strftime("%d.%m.%Y")
                        else:
                            date_str = str(next_cell).strip()

        guruhlar = {}
        for r in range(header_row + 1, sheet.max_row + 1):
            guruh_val = sheet.cell(row=r, column=1).value
            fio_val = sheet.cell(row=r, column=3).value

            if not guruh_val or not fio_val: continue
            
            g_name = str(guruh_val).strip()
            if g_name.endswith('.0'): g_name = g_name[:-2]
            if g_name.lower().startswith(('jami', 'итого', 'guruh')): continue

            no_val = sheet.cell(row=r, column=2).value or len(guruhlar.get(g_name, [])) + 1
            kerak_val = sheet.cell(row=r, column=4).value or 0
            jami_val = sheet.cell(row=r, column=5).value or 0
            qarzi_val = sheet.cell(row=r, column=6).value or 0

            try: kerak_num = float(kerak_val)
            except: kerak_num = 0.0

            try: jami_num = float(jami_val)
            except: jami_num = 0.0

            try: qarzi_num = float(qarzi_val)
            except: qarzi_num = 0.0

            if g_name not in guruhlar:
                guruhlar[g_name] = []

            guruhlar[g_name].append({
                "no": no_val,
                "fio": str(fio_val).strip(),
                "kerak": kerak_num,
                "jami": jami_num,
                "qarzi": qarzi_num
            })

        wb.close()
        if os.path.exists(temp_excel): os.remove(temp_excel)

        if not guruhlar:
            progress.error("❌ Excel faylidan guruhlar ma'lumotlari topilmadi.")
            user_data[chat_id] = {}
            return

        guruh_nomlari = sorted(guruhlar.keys())
        total_g = len(guruh_nomlari)

        progress.update("⚙️ Natija tayyorlanmoqda...", 40)

        for idx, g_name in enumerate(guruh_nomlari, start=1):
            pct = 40 + int((idx / total_g) * 50)
            progress.update(f"⚙️ {g_name} guruhi tayyorlanmoqda...", pct)

            rows_data = guruhlar[g_name]
            img_path = os.path.join(tempfile.gettempdir(), f"screenshot_{chat_id}_{g_name}.png")
            
            generate_group_table_image(g_name, date_str, rows_data, img_path)
            
            with open(img_path, 'rb') as photo_f:
                bot.send_photo(chat_id, photo=photo_f, caption=f"📊 <b>Guruh: {escape_html_text(g_name)}</b>", parse_mode="HTML")
            
            if os.path.exists(img_path): os.remove(img_path)

        progress.update("🔍 Yakuniy tekshiruv...", 95)
        send_safe_message(chat_id, f"✅ <b>Barcha {len(guruh_nomlari)} ta guruh screenshotlari muvaffaqiyatli yuborildi!</b>")
        progress.success("✅ Tayyor!")
        user_data[chat_id] = {}

    except Exception as e:
        progress.error(f"❌ Xatolik yuz berdi:\n<code>{escape_html_text(str(e))}</code>")
        if os.path.exists(temp_excel): os.remove(temp_excel)
        user_data[chat_id] = {}

# ==================== META ADS INTEGRATSIYASI FUNKSIYALARI ====================
META_USER_STATE = {}

def handle_meta_account_info(chat_id):
    if not meta_api:
        send_safe_message(chat_id, "❌ Meta Ads API moduli ulanmagan.", reply_markup=get_meta_ads_keyboard())
        return
    bot.send_chat_action(chat_id, "typing")
    bal_info = meta_api.get_balance_details()
    if "error" in bal_info:
        err_msg = bal_info['error'].get('message', 'Nomaʼlum xatolik')
        send_safe_message(chat_id, f"❌ <b>Xatolik yuz berdi:</b>\n{err_msg}", reply_markup=get_meta_ads_keyboard())
        return

    settings = load_meta_settings() if load_meta_settings else {}
    custom_limit = float(settings.get("custom_budget_limit", 0))
    base_spent = float(settings.get("initial_spent_base", 0))
    current_spent = float(bal_info.get("amount_spent", 0))

    if custom_limit > 0:
        spent_since_limit = max(0.0, current_spent - base_spent)
        remaining = custom_limit - spent_since_limit
        if remaining > 0:
            budget_text = (
                f"💵 <b>Kiritilgan mablag‘ (Funds):</b> <code>${custom_limit:.2f}</code>\n"
                f"💸 <b>Ushbu mablag‘dan sarflandi:</b> <code>${spent_since_limit:.2f}</code>\n"
                f"🟢 <b>QOLGAN MABLAG' (Mavjud qoldiq):</b> <b>${remaining:.2f}</b>"
            )
        else:
            budget_text = (
                f"💵 <b>Kiritilgan mablag‘ (Funds):</b> <code>${custom_limit:.2f}</code>\n"
                f"💸 <b>Ushbu mablag‘dan sarflandi:</b> <code>${spent_since_limit:.2f}</code>\n"
                f"🔴 <b>QOLGAN MABLAG':</b> <b>$0.00 (Qarzda: ${abs(remaining):.2f})</b>"
            )
    else:
        budget_text = "ℹ️ <i>Maxsus byudjet limiti belgilanmagan.\n(Qolgan pulni hisoblash uchun quyidagi '⚙️ Byudjet limitini o‘rnatish' tugmasini bosing)</i>"

    status_map = {1: "🟢 Faol (Active)", 2: "🔴 O‘chirilgan (Disabled)", 3: "🟡 To‘lov kutilmoqda (Unsettled)"}
    status_text = status_map.get(bal_info.get("account_status"), "Nomaʼlum")
    card_info = bal_info.get("card", "Karta")

    today_ins = meta_api.get_insights("today")
    from meta_ads_bot.config import AD_ACCOUNT_ID

    text = (
        f"👤 <b>Reklama hisobi:</b> {bal_info.get('account_name')}\n"
        f"🆔 <b>Ad Account ID:</b> <code>{AD_ACCOUNT_ID}</code>\n"
        f"⚡️ <b>Holat:</b> {status_text}\n"
        f"💳 <b>To‘lov usuli:</b> {card_info}\n"
        f"────────────────────\n"
        f"📊 <b>Bugun sarflandi:</b> ${today_ins.get('spend', '0.00')}\n"
        f"🎯 <b>Bugungi lidlar:</b> {today_ins.get('leads', '0')} ta ({today_ins.get('cpl', '—')}/lid)\n"
        f"💳 <b>Jami hisob xarajati:</b> ${current_spent:,.2f}\n"
        f"────────────────────\n"
        f"{budget_text}\n"
    )

    markup = telebot.types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        telebot.types.InlineKeyboardButton("⚙️ Byudjet limitini o‘rnatish ($)", callback_data="meta_set_budget_limit"),
        telebot.types.InlineKeyboardButton("🎯 Kampaniyalarni ko‘rish", callback_data="meta_show_campaigns"),
        telebot.types.InlineKeyboardButton("📈 To‘liq hisobot", callback_data="meta_ins_today")
    )
    bot.send_message(chat_id, text, reply_markup=markup, parse_mode="HTML")

def handle_meta_campaigns_list(chat_id):
    if not meta_api:
        send_safe_message(chat_id, "❌ Meta Ads API moduli ulanmagan.", reply_markup=get_meta_ads_keyboard())
        return
    bot.send_chat_action(chat_id, "typing")
    campaigns = meta_api.get_campaigns()
    if not campaigns:
        send_safe_message(chat_id, "📭 Hech qanday kampaniya topilmadi yoki xatolik yuz berdi.", reply_markup=get_meta_ads_keyboard())
        return

    markup = telebot.types.InlineKeyboardMarkup(row_width=1)
    text = "📋 <b>Mavjud reklama kampaniyalari:</b>\n\nBatafsil ma'lumot yoki boshqarish uchun kampaniya ustiga bosing:\n"

    for c in campaigns:
        status_icon = "🟢" if c.get("status") == "ACTIVE" else "🔴"
        budget = float(c.get("daily_budget", 0)) / 100 if c.get("daily_budget") else 0
        budget_str = f" (${budget:.0f}/kun)" if budget > 0 else ""
        btn_text = f"{status_icon} {c.get('name')}{budget_str}"
        markup.add(telebot.types.InlineKeyboardButton(btn_text, callback_data=f"meta_camp_{c['id']}"))

    bot.send_message(chat_id, text, reply_markup=markup, parse_mode="HTML")

def handle_meta_insights_menu(chat_id):
    bot.send_message(
        chat_id,
        "📈 <b>Qaysi davr uchun statistikani ko‘rmoqchisiz?</b>\nQuyidagi tugmalardan birini tanlang:",
        reply_markup=get_meta_insights_inline(),
        parse_mode="HTML"
    )

def handle_meta_automation_menu(chat_id):
    settings = load_meta_settings() if load_meta_settings else {}
    status_str = "🟢 Yoqilgan" if settings.get("auto_schedule_enabled") else "🔴 O‘chirilgan"
    report_status = "🟢 Yoqilgan" if settings.get("daily_report_enabled") else "🔴 O‘chirilgan"

    text = (
        f"⏰ <b>Avtomatlashtirish va Xavfsizlik sozlamalari</b>\n\n"
        f"🌙 <b>Tungi rejim (Auto-Pause):</b> {status_str}\n"
        f"  └ ⏸ O‘chirish: <code>{settings.get('pause_time')}</code>\n"
        f"  └ ▶️ Qayta yoqish: <code>{settings.get('resume_time')}</code>\n\n"
        f"📊 <b>Kunlik avtomat hisobot:</b> {report_status}\n"
        f"  └ ⏰ Yuborish vaqti: <code>{settings.get('daily_report_time')}</code>\n\n"
        f"ℹ️ <i>Eslatma: Byudjet 0 ga tushganda faqat ogohlantirish xabari keladi, reklamalar to‘xtatilmaydi.</i>"
    )

    toggle_btn_text = "🔴 Tungi rejimni o‘chirish" if settings.get("auto_schedule_enabled") else "🟢 Tungi rejimni yoqish"
    report_toggle_text = "🔴 Hisobotni o‘chirish" if settings.get("daily_report_enabled") else "🟢 Hisobotni yoqish"

    markup = telebot.types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        telebot.types.InlineKeyboardButton(toggle_btn_text, callback_data="meta_toggle_auto_schedule"),
        telebot.types.InlineKeyboardButton("⏱ Tungi o‘chirish vaqtini o‘zgartirish", callback_data="meta_set_pause_time"),
        telebot.types.InlineKeyboardButton("⏱ Ertalabki yoqish vaqtini o‘zgartirish", callback_data="meta_set_resume_time"),
        telebot.types.InlineKeyboardButton(report_toggle_text, callback_data="meta_toggle_daily_report")
    )
    bot.send_message(chat_id, text, reply_markup=markup, parse_mode="HTML")

# Meta Ads Callback Handlers
@bot.callback_query_handler(func=lambda call: call.data.startswith("meta_") or call.data.startswith("ins_") or call.data.startswith("camp_"))
def handle_meta_callbacks(call):
    if not is_user_allowed(call):
        bot.answer_callback_query(call.id, "⛔️ Ruxsat berilmagan!", show_alert=True)
        return

    data = call.data
    chat_id = call.message.chat.id
    user_id = call.from_user.id

    if data in ["meta_set_budget_limit", "set_budget_limit"]:
        META_USER_STATE[user_id] = {"action": "set_custom_budget_limit"}
        bot.send_message(
            chat_id,
            "💰 <b>Reklama uchun ajratgan byudjetingizni ($ dollarda) yozing:</b>\n\n"
            "Misol uchun: <code>50</code> yoki <code>100</code> yoki <code>250</code>\n\n"
            "<i>(Bot hozirdan boshlab xarajatni hisoblaydi va ushbu summa tugab, 0 $ bo‘lganda sizga darhol bildirishnoma yuboradi. Reklamalar to‘xtatilmaydi).</i>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data in ["meta_show_campaigns", "show_campaigns"]:
        handle_meta_campaigns_list(chat_id)
        bot.answer_callback_query(call.id)

    elif data.startswith("meta_ins_") or data.startswith("ins_"):
        period = data.replace("meta_ins_", "").replace("ins_", "")
        period_names = {
            "today": "Bugungi",
            "yesterday": "Kechagi",
            "last_7d": "Oxirgi 7 kunlik",
            "this_month": "Shu oylik"
        }
        bot.answer_callback_query(call.id, "Statistika yuklanmoqda...")
        bot.send_chat_action(chat_id, "typing")

        ins = meta_api.get_insights(period) if meta_api else {}
        acc = meta_api.get_account_info() if meta_api else {}

        title = period_names.get(period, "Statistika")
        text = (
            f"📊 <b>{title} hisobot</b>\n"
            f"👤 <b>Hisob:</b> {acc.get('name', 'Ads Account')}\n"
            f"📅 <b>Sana:</b> {ins.get('date_start', '')} — {ins.get('date_stop', '')}\n"
            f"────────────────────\n"
            f"💵 <b>Xarajat (Spend):</b> ${ins.get('spend', '0')}\n"
            f"🎯 <b>Lidlar soni (Leads):</b> {ins.get('leads', '0')} ta\n"
            f"📉 <b>1 ta lid narxi (CPL):</b> {ins.get('cpl', '—')}\n"
            f"👁 <b>Ko‘rishlar (Impressions):</b> {ins.get('impressions', '0')}\n"
            f"🖱 <b>Kliklar (Clicks):</b> {ins.get('clicks', '0')}\n"
            f"🎯 <b>CTR:</b> {ins.get('ctr', '0.00%')}\n"
            f"⚡️ <b>CPC:</b> {ins.get('cpc', '$0.00')}\n"
            f"📈 <b>CPM:</b> {ins.get('cpm', '$0.00')}\n"
        )

        markup = telebot.types.InlineKeyboardMarkup(row_width=2)
        markup.add(
            telebot.types.InlineKeyboardButton("🔄 Yangilash", callback_data=f"meta_ins_{period}"),
            telebot.types.InlineKeyboardButton("⬅️ Boshqa davr", callback_data="meta_back_to_insights")
        )
        try:
            bot.edit_message_text(text, chat_id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
        except Exception:
            bot.send_message(chat_id, text, reply_markup=markup, parse_mode="HTML")

    elif data in ["meta_back_to_insights", "back_to_insights"]:
        bot.edit_message_text(
            "📈 <b>Qaysi davr uchun statistikani ko‘rmoqchisiz?</b>\nQuyidagi tugmalardan birini tanlang:",
            chat_id,
            call.message.message_id,
            reply_markup=get_meta_insights_inline(),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data.startswith("meta_camp_") or data.startswith("camp_"):
        campaign_id = data.replace("meta_camp_", "").replace("camp_", "")
        bot.answer_callback_query(call.id, "Yuklanmoqda...")
        bot.send_chat_action(chat_id, "typing")

        c = meta_api.get_campaign(campaign_id) if meta_api else {}
        if "error" in c:
            bot.send_message(chat_id, f"❌ Xatolik: {c['error'].get('message')}", parse_mode="HTML")
            return

        ins = meta_api.get_insights("today", campaign_id=campaign_id) if meta_api else {}
        status = c.get("status")
        status_icon = "🟢 Faol (ACTIVE)" if status == "ACTIVE" else "🔴 To‘xtatilgan (PAUSED)"
        budget = float(c.get("daily_budget", 0)) / 100 if c.get("daily_budget") else 0

        text = (
            f"🎯 <b>Kampaniya:</b> {c.get('name')}\n"
            f"🆔 <b>ID:</b> <code>{c.get('id')}</code>\n"
            f"⚡️ <b>Holat:</b> {status_icon}\n"
            f"🎯 <b>Maqsad:</b> <code>{c.get('objective', 'Nomaʼlum')}</code>\n"
            f"💰 <b>Kunlik byudjet:</b> ${budget:.2f}\n"
            f"────────────────────\n"
            f"📊 <b>Bugungi statistika:</b>\n"
            f"  • Sarflandi: ${ins.get('spend', '0')}\n"
            f"  • Lidlar: {ins.get('leads', '0')} ta\n"
            f"  • Lid narxi: {ins.get('cpl', '—')}\n"
            f"  • Kliklar: {ins.get('clicks', '0')} (CTR: {ins.get('ctr', '0')})\n"
        )

        markup = telebot.types.InlineKeyboardMarkup(row_width=2)
        if status == "ACTIVE":
            markup.add(telebot.types.InlineKeyboardButton("⏸ To‘xtatish (Pause)", callback_data=f"meta_toggle_camp_{campaign_id}_PAUSED"))
        else:
            markup.add(telebot.types.InlineKeyboardButton("▶️ Yoqish (Active)", callback_data=f"meta_toggle_camp_{campaign_id}_ACTIVE"))

        markup.add(
            telebot.types.InlineKeyboardButton("💵 Byudjetni o‘zgartirish", callback_data=f"meta_set_budget_{campaign_id}"),
            telebot.types.InlineKeyboardButton("⬅️ Barcha kampaniyalar", callback_data="meta_show_campaigns")
        )

        try:
            bot.edit_message_text(text, chat_id, call.message.message_id, reply_markup=markup, parse_mode="HTML")
        except Exception:
            bot.send_message(chat_id, text, reply_markup=markup, parse_mode="HTML")

    elif data.startswith("meta_toggle_camp_"):
        parts = data.split("_")
        target_status = parts[-1]
        campaign_id = parts[3]

        bot.answer_callback_query(call.id, "Holat o'zgartirilmoqda...")
        res = meta_api.set_campaign_status(campaign_id, target_status) if meta_api else {}

        if "error" in res:
            bot.send_message(chat_id, f"❌ Xatolik yuz berdi: {res['error'].get('message')}", parse_mode="HTML")
        else:
            status_word = "yoqildi (ACTIVE)" if target_status == "ACTIVE" else "to‘xtatildi (PAUSED)"
            bot.answer_callback_query(call.id, f"✅ Kampaniya {status_word}!", show_alert=True)
            call.data = f"meta_camp_{campaign_id}"
            handle_meta_callbacks(call)

    elif data.startswith("meta_set_budget_"):
        campaign_id = data.replace("meta_set_budget_", "")
        META_USER_STATE[user_id] = {"action": "change_budget", "campaign_id": campaign_id}
        bot.send_message(
            chat_id,
            "💵 <b>Yangi kunlik byudjet miqdorini dollarda yuboring:</b>\n\nMisol uchun: <code>15</code> yoki <code>25.5</code>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data in ["meta_toggle_auto_schedule", "toggle_auto_schedule"]:
        settings = load_meta_settings() if load_meta_settings else {}
        settings["auto_schedule_enabled"] = not settings.get("auto_schedule_enabled", False)
        if save_meta_settings: save_meta_settings(settings)

        state_word = "faollashtirildi" if settings["auto_schedule_enabled"] else "o‘chirildi"
        bot.answer_callback_query(call.id, f"✅ Tungi rejim {state_word}!", show_alert=True)
        handle_meta_automation_menu(chat_id)

    elif data in ["meta_toggle_daily_report", "toggle_daily_report"]:
        settings = load_meta_settings() if load_meta_settings else {}
        settings["daily_report_enabled"] = not settings.get("daily_report_enabled", True)
        if save_meta_settings: save_meta_settings(settings)

        state_word = "yoqildi" if settings["daily_report_enabled"] else "o‘chirildi"
        bot.answer_callback_query(call.id, f"✅ Kunlik hisobot {state_word}!", show_alert=True)
        handle_meta_automation_menu(chat_id)

    elif data in ["meta_set_pause_time", "meta_set_resume_time", "set_pause_time", "set_resume_time"]:
        action = data.replace("meta_", "")
        META_USER_STATE[user_id] = {"action": action}
        label = "o‘chirish" if "pause" in action else "qayta yoqish"
        bot.send_message(
            chat_id,
            f"⏱ <b>Reklamalarni avtomatik {label} vaqtini yozing (HH:MM formatida):</b>\n\nMisol uchun: <code>23:00</code> yoki <code>07:30</code>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

# ==================== INSTAGRAM AUTOPOSTER INTEGRATSIYASI FUNKSIYALARI ====================
INSTA_USER_STATE = {}

def handle_insta_dashboard(chat_id):
    """Instagram AutoPoster asosiy ma'lumotlar paneli"""
    stats = get_insta_queue_stats()
    settings = stats.get("settings", {})
    
    username = settings.get("insta_username", "shahrisabz_t_t_uz")
    sched_enabled = settings.get("auto_schedule_enabled") == "1"
    sched_str = f"🟢 Yoqilgan (Har {settings.get('interval_minutes', '60')} daqiqada)" if sched_enabled else "🔴 O‘chirilgan"
    yt_enabled = settings.get("youtube_auto_upload", "1") == "1"
    yt_str = "🟢 Yoqilgan (Shorts)" if yt_enabled else "🔴 O‘chirilgan"
    target_chat = settings.get("target_chat_id", "8135594558")
    last_post = settings.get("last_post_time") or "Hali yuborilmadi"
    last_scan = settings.get("last_scan_time") or "Skanerlanmagan"
    
    text = (
        f"📸 <b>INSTAGRAM AUTOPOSTER BOSHQARUVI</b>\n\n"
        f"👤 <b>Instagram Profil:</b> <code>@{username}</code>\n"
        f"🎯 <b>Maqsadli Chat/Kanal ID:</b> <code>{target_chat}</code>\n"
        f"⏰ <b>Avto-jadval holati:</b> {sched_str}\n"
        f"📺 <b>YouTube Shorts yuklash:</b> {yt_str}\n"
        f"🕒 <b>Oxirgi yuborilgan:</b> <code>{last_post}</code>\n"
        f"🔍 <b>Oxirgi skanerlash:</b> <code>{last_scan}</code>\n"
        f"────────────────────\n"
        f"📊 <b>NAVBAT STATISTIKASI:</b>\n"
        f"  • 📁 Jami postlar bazada: <b>{stats['total']} ta</b>\n"
        f"  • ⏳ Yuborishga tayyor (Navbatda): <b>{stats['pending']} ta</b>\n"
        f"  • ✅ Kanal/Botga yuborildi: <b>{stats['sent']} ta</b>\n"
        f"  • ❌ Xatolik berganlar: <b>{stats['failed']} ta</b>\n"
    )
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        telebot.types.InlineKeyboardButton("🚀 Keyingi 1 tani yuborish", callback_data="insta_post_now"),
        telebot.types.InlineKeyboardButton("📥 Qayta skanerlash", callback_data="insta_scan_now")
    )
    
    sched_btn_text = "🔴 Jadvalni to‘xtatish" if sched_enabled else "🟢 Avto-jadvalni yoqish"
    markup.add(
        telebot.types.InlineKeyboardButton(sched_btn_text, callback_data="insta_toggle_sched"),
        telebot.types.InlineKeyboardButton("🔄 Xatolarni qayta qo‘yish", callback_data="insta_reset_failed")
    )
    markup.add(
        telebot.types.InlineKeyboardButton("⚙️ Sozlamalar", callback_data="insta_settings_menu"),
        telebot.types.InlineKeyboardButton("🔄 Yangilash", callback_data="insta_refresh_dash")
    )
    
    send_safe_message(chat_id, text, reply_markup=markup)


def handle_insta_post_one(chat_id):
    """Navbatdagi 1 ta postni yuborish"""
    bot.send_chat_action(chat_id, "upload_video")
    res = post_next_insta_item()
    if res.get("success"):
        stats = get_insta_queue_stats()
        send_safe_message(
            chat_id,
            f"✅ <b>Post muvaffaqiyatli Telegramga yuborildi!</b>\n\n"
            f"🔗 <b>Post:</b> <a href='{res.get('post_url')}'>{res.get('shortcode')}</a>\n"
            f"📝 <b>Matn:</b> <i>{escape_html_text(res.get('caption', ''))}</i>\n\n"
            f"⏳ <b>Navbatda qoldi:</b> {stats['pending']} ta post",
            reply_markup=get_insta_poster_keyboard()
        )
    elif res.get("empty"):
        send_safe_message(
            chat_id,
            "🎉 <b>Barcha postlar yuklab bo‘lingan!</b>\nNavbatda boshqa yuborilmagan post qolmadi. Yangi postlar uchun '📥 Instagramdan skanerlash' tugmasini bosing.",
            reply_markup=get_insta_poster_keyboard()
        )
    else:
        send_safe_message(
            chat_id,
            f"❌ <b>Postni yuborishda xatolik:</b>\n<code>{escape_html_text(res.get('error', 'Nomaʼlum'))}</code>",
            reply_markup=get_insta_poster_keyboard()
        )


def handle_insta_scan(chat_id):
    """Profilni fonda skanerlash"""
    is_scanning = get_insta_setting("is_scanning", "0") == "1"
    if is_scanning:
        send_safe_message(chat_id, "⏳ <b>Skanerlash jarayoni hozir fonda davom etmoqda...</b>\nIltimos, biroz kuting.", reply_markup=get_insta_poster_keyboard())
        return
        
    send_safe_message(
        chat_id,
        "🔍 <b>Instagram profil skanerlanishi boshlandi...</b>\n\n"
        "Playwright fon rejimida barcha postlarni xronologik tartibda yig‘ib bazaga saqlaydi. "
        "Tugagach sizga bildirishnoma yuboriladi.",
        reply_markup=get_insta_poster_keyboard()
    )
    
    def _on_finish(result):
        if result.get("success"):
            stats = get_insta_queue_stats()
            bot.send_message(
                chat_id,
                f"✅ <b>Instagram skanerlash muvaffaqiyatli yakunlandi!</b>\n\n"
                f"👤 <b>Profil:</b> @{result.get('username')}\n"
                f"🔍 <b>Topilgan postlar:</b> {result.get('total_found')} ta\n"
                f"📥 <b>Navbatga yangi qo‘shildi:</b> {result.get('new_added')} ta\n"
                f"⏳ <b>Jami navbatda kutmoqda:</b> {stats['pending']} ta",
                parse_mode="HTML"
            )
        else:
            bot.send_message(
                chat_id,
                f"❌ <b>Skanerlashda xatolik:</b>\n<code>{escape_html_text(result.get('error', ''))}</code>",
                parse_mode="HTML"
            )
            
    scan_insta_background(callback_notify=_on_finish)


def handle_insta_settings_menu(chat_id):
    """Instagram AutoPoster sozlamalari"""
    settings = get_insta_settings()
    username = settings.get("insta_username", "shahrisabz_t_t_uz")
    chat_target = settings.get("target_chat_id", "8135594558")
    interval = settings.get("interval_minutes", "60")
    yt_enabled = settings.get("youtube_auto_upload", "1") == "1"
    yt_btn_text = "🔴 YouTube Shorts yuklashni o‘chirish" if yt_enabled else "🟢 YouTube Shorts yuklashni yoqish"
    
    text = (
        f"⚙️ <b>INSTAGRAM AUTOPOSTER SOZLAMALARI</b>\n\n"
        f"1️⃣ <b>Instagram Username:</b> <code>@{username}</code>\n"
        f"2️⃣ <b>Yuborish manzili (Chat/Kanal ID):</b> <code>{chat_target}</code>\n"
        f"3️⃣ <b>Avto-yuborish oralig‘i:</b> <code>{interval} daqiqa (har {int(interval)//60 if int(interval)>=60 else interval} soat)</code>\n"
        f"4️⃣ <b>YouTube Shorts Auto-Upload:</b> {'🟢 Yoqilgan' if yt_enabled else '🔴 O‘chirilgan'}\n\n"
        f"O‘zgartirmoqchi bo‘lgan parametrni tanlang:"
    )
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=1)
    markup.add(
        telebot.types.InlineKeyboardButton("👤 Instagram profil nomini o‘zgartirish", callback_data="insta_set_user"),
        telebot.types.InlineKeyboardButton("🎯 Target Chat / Kanal ID sini o‘zgartirish", callback_data="insta_set_chat"),
        telebot.types.InlineKeyboardButton("⏱ Telegram oraliq vaqtini o‘zgartirish", callback_data="insta_set_interval_prompt"),
        telebot.types.InlineKeyboardButton("⏰ YouTube Rek Vaqtlari Jadvali (+/-)", callback_data="insta_yt_sched_menu"),
        telebot.types.InlineKeyboardButton(yt_btn_text, callback_data="insta_toggle_youtube"),
        telebot.types.InlineKeyboardButton("⬅️ Boshqaruv paneliga qaytish", callback_data="insta_refresh_dash")
    )
    send_safe_message(chat_id, text, reply_markup=markup)


def handle_insta_youtube_schedule_menu(chat_id):
    """YouTube Shorts Rek Jadvali menyusi"""
    times = get_youtube_schedule_times()
    sched_enabled = get_insta_setting("youtube_schedule_enabled", "1") == "1"
    yt_auto = get_insta_setting("youtube_auto_upload", "1") == "1"
    
    times_list = [f"  {idx+1}️⃣ 🕒 <b>{t}</b>" for idx, t in enumerate(times)]
    times_str = "\n".join(times_list) if times else "  <i>Hozircha vaqtlar belgilanmagan</i>"
    
    status_str = "🟢 Faol (Yoqilgan)" if (sched_enabled and yt_auto) else "🔴 O‘chirilgan"
    
    text = (
        f"📺 <b>YOUTUBE SHORTS REK VAQTLARI JADVALI</b>\n\n"
        f"⚡️ Bot har kuni quyida belgilangan vaqtlarda Instagramdagi navbatdagi videoni avtomatik <b>YouTube Shorts</b> ga yuklab boradi:\n\n"
        f"📋 <b>Joriy belgilangan vaqtlar:</b>\n{times_str}\n\n"
        f"📊 <b>Jadval holati:</b> {status_str}\n\n"
        f"Yangi vaqt qo‘shish (<code>+</code>) yoki o‘chirish (<code>-</code>) uchun tugmalardan foydalaning:"
    )
    
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        telebot.types.InlineKeyboardButton("➕ Yangi vaqt qo‘shish", callback_data="insta_yt_add_time"),
        telebot.types.InlineKeyboardButton("🗑 Vaqtni o‘chirish", callback_data="insta_yt_del_menu")
    )
    sched_toggle_txt = "🔴 Jadvalni to‘xtatish" if sched_enabled else "🟢 Jadvalni yoqish"
    markup.add(
        telebot.types.InlineKeyboardButton(sched_toggle_txt, callback_data="insta_yt_toggle_sched"),
        telebot.types.InlineKeyboardButton("🔄 Standart (3 ta vaqt)", callback_data="insta_yt_reset_times")
    )
    markup.add(
        telebot.types.InlineKeyboardButton("🚀 Hozir YouTubega 1 ta yuklash", callback_data="insta_yt_post_now")
    )
    markup.add(
        telebot.types.InlineKeyboardButton("⬅️ Sozlamalarga qaytish", callback_data="insta_settings_menu")
    )
    send_safe_message(chat_id, text, reply_markup=markup)


def handle_insta_youtube_delete_menu(chat_id):
    """YouTube jadvalidan vaqtni o'chirish tanlash menyusi"""
    times = get_youtube_schedule_times()
    if not times:
        send_safe_message(chat_id, "Jadvalda o‘chirish uchun vaqtlar mavjud emas.")
        handle_insta_youtube_schedule_menu(chat_id)
        return
        
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    for t in times:
        markup.add(telebot.types.InlineKeyboardButton(f"🗑 {t} ni o‘chirish", callback_data=f"insta_yt_del_val_{t}"))
    markup.add(telebot.types.InlineKeyboardButton("⬅️ Bekor qilish", callback_data="insta_yt_sched_menu"))
    
    send_safe_message(chat_id, "🗑 <b>O‘chirmoqchi bo‘lgan vaqtingizni tanlang:</b>", reply_markup=markup)


# Instagram AutoPoster Callback Handlers
@bot.callback_query_handler(func=lambda call: call.data.startswith("insta_"))
def handle_insta_callbacks(call):
    if not is_user_allowed(call):
        bot.answer_callback_query(call.id, "⛔️ Ruxsat berilmagan!", show_alert=True)
        return

    data = call.data
    chat_id = call.message.chat.id
    user_id = call.from_user.id

    if data in ["insta_refresh_dash", "insta_dash"]:
        bot.answer_callback_query(call.id, "Yangilanmoqda...")
        handle_insta_dashboard(chat_id)

    elif data == "insta_post_now":
        bot.answer_callback_query(call.id, "Post yuborilmoqda...")
        handle_insta_post_one(chat_id)

    elif data == "insta_scan_now":
        bot.answer_callback_query(call.id, "Skanerlash boshlandi...")
        handle_insta_scan(chat_id)

    elif data == "insta_toggle_sched":
        enabled = get_insta_setting("auto_schedule_enabled", "0") == "1"
        new_val = "0" if enabled else "1"
        set_insta_setting("auto_schedule_enabled", new_val)
        status_word = "yoqildi" if new_val == "1" else "to‘xtatildi"
        bot.answer_callback_query(call.id, f"✅ Telegram avto-jadval {status_word}!", show_alert=True)
        handle_insta_dashboard(chat_id)

    elif data == "insta_reset_failed":
        count = reset_insta_queue()
        bot.answer_callback_query(call.id, f"✅ {count} ta xatolik bergan post qayta navbatga qo‘yildi!", show_alert=True)
        handle_insta_dashboard(chat_id)

    elif data == "insta_settings_menu":
        bot.answer_callback_query(call.id)
        handle_insta_settings_menu(chat_id)

    elif data == "insta_toggle_youtube":
        current = get_insta_setting("youtube_auto_upload", "1") == "1"
        new_val = "0" if current else "1"
        set_insta_setting("youtube_auto_upload", new_val)
        status_word = "yoqildi" if new_val == "1" else "to‘xtatildi"
        bot.answer_callback_query(call.id, f"✅ YouTube avto-yuklash {status_word}!", show_alert=True)
        handle_insta_settings_menu(chat_id)

    elif data == "insta_yt_sched_menu":
        bot.answer_callback_query(call.id)
        handle_insta_youtube_schedule_menu(chat_id)

    elif data == "insta_yt_add_time":
        INSTA_USER_STATE[user_id] = {"action": "add_yt_time"}
        bot.send_message(
            chat_id,
            "⏰ <b>Yangi YouTube Shorts yuklash vaqtini kiriting:</b>\n\n"
            "Format: <code>HH:MM</code> (Masalan: <code>15:30</code>, <code>20:00</code> yoki <code>11:45</code>)",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data == "insta_yt_del_menu":
        bot.answer_callback_query(call.id)
        handle_insta_youtube_delete_menu(chat_id)

    elif data.startswith("insta_yt_del_val_"):
        t_val = data.replace("insta_yt_del_val_", "")
        remove_youtube_schedule_time(t_val)
        bot.answer_callback_query(call.id, f"🗑 {t_val} vaqti jadvaldan o‘chirildi!", show_alert=True)
        handle_insta_youtube_schedule_menu(chat_id)

    elif data == "insta_yt_reset_times":
        reset_youtube_schedule_times()
        bot.answer_callback_query(call.id, "🔄 Standart 3 ta vaqtga qaytarildi (09:00, 13:00, 19:30)!", show_alert=True)
        handle_insta_youtube_schedule_menu(chat_id)

    elif data == "insta_yt_toggle_sched":
        current = get_insta_setting("youtube_schedule_enabled", "1") == "1"
        new_val = "0" if current else "1"
        set_insta_setting("youtube_schedule_enabled", new_val)
        status_word = "yoqildi" if new_val == "1" else "to‘xtatildi"
        bot.answer_callback_query(call.id, f"✅ YouTube jadvali {status_word}!", show_alert=True)
        handle_insta_youtube_schedule_menu(chat_id)

    elif data == "insta_yt_post_now":
        bot.answer_callback_query(call.id, "YouTubega yuklanmoqda...")
        bot.send_message(chat_id, "⏳ <b>Navbatdagi video YouTube Shorts ga yuklanmoqda...</b>", parse_mode="HTML")
        yt_res = post_next_youtube_video()
        if yt_res.get("success"):
            bot.send_message(
                chat_id,
                f"✅ <b>Video YouTube Shorts ga muvaffaqiyatli yuklandi!</b>\n\n"
                f"🎬 <b>Sarlavha:</b> {escape_html_text(yt_res.get('title', ''))}\n"
                f"🔗 <b>Havola:</b> {yt_res.get('url')}",
                parse_mode="HTML"
            )
        elif yt_res.get("empty"):
            bot.send_message(chat_id, "🎉 YouTubega yuklash uchun yangi videolar qolmadi.", parse_mode="HTML")
        else:
            bot.send_message(chat_id, f"❌ <b>Xatolik:</b> {escape_html_text(yt_res.get('error', 'Nomaʼlum'))}", parse_mode="HTML")

    elif data == "insta_set_user":
        INSTA_USER_STATE[user_id] = {"action": "set_insta_user"}
        bot.send_message(
            chat_id,
            "👤 <b>Yangi Instagram sahifa nomini (username) yuboring:</b>\n\nMisol uchun: <code>shahrisabz_t_t_uz</code>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data == "insta_set_chat":
        INSTA_USER_STATE[user_id] = {"action": "set_insta_chat"}
        bot.send_message(
            chat_id,
            "🎯 <b>Postlar yuboriladigan yangi Chat yoki Kanal ID sini yuboring:</b>\n\n"
            "Misol uchun:\n"
            "• Shaxsiy Telegram ID: <code>8135594558</code>\n"
            "• Telegram Kanal ID: <code>-1001234567890</code> yoki <code>@kanal_nomi</code>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

    elif data.startswith("insta_like_"):
        post_id_str = data.replace("insta_like_", "")
        try:
            post_id = int(post_id_str)
            res = toggle_post_like(post_id, user_id)
            new_kb = get_post_inline_keyboard(post_id, res["post_url"], res["likes_count"])
            try:
                bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=new_kb)
            except Exception:
                pass
            
            status_text = f"❤️ Sizga yoqdi! ({res['likes_count']} ta)" if res["is_liked"] else f"💔 Like bekor qilindi ({res['likes_count']} ta)"
            bot.answer_callback_query(call.id, status_text)
        except Exception as e:
            print(f"[Like Callback Error]: {e}")
            bot.answer_callback_query(call.id, "Like qayd etildi!")

    elif data == "insta_set_interval_prompt":
        INSTA_USER_STATE[user_id] = {"action": "set_insta_interval"}
        bot.send_message(
            chat_id,
            "⏱ <b>Avto-yuborish oralig‘ini daqiqalarda kiriting:</b>\n\n"
            "Misol uchun:\n"
            "• Har 30 daqiqada: <code>30</code>\n"
            "• Har 1 soatda (60 daqiqa): <code>60</code>\n"
            "• Har 2 soatda: <code>120</code>",
            reply_markup=telebot.types.ForceReply(selective=True),
            parse_mode="HTML"
        )
        bot.answer_callback_query(call.id)

if __name__ == "__main__":
    bot.remove_webhook()
    app.run(host="0.0.0.0", port=int(os.environ.get('PORT', 5000)))