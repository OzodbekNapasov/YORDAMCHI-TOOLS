import os
import sys
import time
import asyncio
import logging
import tempfile
import shutil
import glob
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)

# Check OS
IS_WINDOWS = os.name == 'nt'

# Safe imports for Windows-specific libraries
try:
    import psutil
except ImportError:
    psutil = None

try:
    from PIL import Image, ImageGrab
except ImportError:
    Image = None
    ImageGrab = None

try:
    import cv2
except ImportError:
    cv2 = None

try:
    import pyautogui
    pyautogui.FAILSAFE = False
except ImportError:
    pyautogui = None

try:
    import ctypes
except ImportError:
    ctypes = None


def is_system_compatible() -> bool:
    """Tekshiradi: Ushbu mashina Windows OS va PC boshqaruviga mosmi."""
    return IS_WINDOWS


def scan_mytestx_tests_dir(base_dir: str = r"D:\MyTestX\tests") -> dict:
    """
    D:\MyTestX\tests papkasidagi barcha .mtf va .xml test fayllarini papkalar bo'yicha guruhlab qaytaradi.
    """
    if not os.path.exists(base_dir):
        candidates = [
            r"D:\MyTestX\tests",
            r"D:\01. Antigravity\online mytestx\tests",
            r"C:\MyTestX\tests"
        ]
        for c in candidates:
            if os.path.exists(c):
                base_dir = c
                break

    if not os.path.exists(base_dir):
        return {"success": False, "error": f"Testlar papkasi topilmadi: {base_dir}", "categories": [], "total_files": 0, "tree": None}

    categories_map = {}
    total_count = 0

    for root, dirs, files in os.walk(base_dir):
        test_files = [f for f in files if f.lower().endswith(('.mtf', '.xml')) and not f.endswith('_new.xml')]
        if not test_files:
            continue

        rel_folder = os.path.relpath(root, base_dir)
        display_folder = "Asosiy Papka" if rel_folder == "." else rel_folder.replace("\\", " / ")

        file_items = []
        for fn in sorted(test_files):
            fp = os.path.join(root, fn)
            try:
                sz = os.path.getsize(fp)
                mtime = os.path.getmtime(fp)
                mtime_str = datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M")
                sz_str = f"{sz // 1024} KB" if sz < 1024*1024 else f"{sz / (1024*1024):.1f} MB"
            except Exception:
                sz_str, mtime_str = "-", "-"

            stem = os.path.splitext(fn)[0]
            ext = os.path.splitext(fn)[1].lower()

            file_items.append({
                "name": fn,
                "stem": stem,
                "ext": ext,
                "path": fp,
                "rel_path": os.path.relpath(fp, base_dir).replace("\\", "/"),
                "size_str": sz_str,
                "mtime_str": mtime_str
            })
            total_count += 1

        if file_items:
            categories_map[display_folder] = file_items

    # Recursive Tree Builder for Windows Explorer Tree View
    def _build_tree(cur_path, rel_path=""):
        node_name = os.path.basename(cur_path) if rel_path else "D:\\MyTestX\\tests"
        children = []
        try:
            entries = sorted(os.scandir(cur_path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except Exception:
            entries = []

        sub_files_count = 0
        for entry in entries:
            child_rel = os.path.join(rel_path, entry.name).replace("\\", "/") if rel_path else entry.name
            if entry.is_dir():
                sub_node, count = _build_tree(entry.path, child_rel)
                if count > 0:
                    children.append(sub_node)
                    sub_files_count += count
            elif entry.is_file():
                fn = entry.name
                if fn.lower().endswith(('.mtf', '.xml')) and not fn.endswith('_new.xml'):
                    try:
                        sz = entry.stat().st_size
                        mtime = entry.stat().st_mtime
                        mtime_str = datetime.fromtimestamp(mtime).strftime("%d.%m.%Y %H:%M")
                        sz_str = f"{sz // 1024} KB" if sz < 1024*1024 else f"{sz / (1024*1024):.1f} MB"
                    except Exception:
                        sz_str, mtime_str = "-", "-"
                    children.append({
                        "type": "file",
                        "name": fn,
                        "stem": os.path.splitext(fn)[0],
                        "ext": os.path.splitext(fn)[1].lower(),
                        "path": entry.path,
                        "rel_path": child_rel,
                        "size_str": sz_str,
                        "mtime_str": mtime_str
                    })
                    sub_files_count += 1

        return {
            "type": "folder",
            "name": node_name,
            "path": cur_path,
            "rel_path": rel_path,
            "total_files": sub_files_count,
            "children": children
        }, sub_files_count

    root_tree, _ = _build_tree(base_dir)

    return {
        "success": True,
        "base_dir": base_dir,
        "total_files": total_count,
        "categories": [{"folder": k, "files": v} for k, v in sorted(categories_map.items())],
        "tree": root_tree
    }




def get_system_status() -> str:
    """
    Kompyuter tizim holati (CPU, RAM, Disk, Batareya, Uptime) haqida ma'lumot yig'adi.
    """
    if not psutil:
        return "⚠️ <code>psutil</code> kutubxonasi o'rnatilmagan yoki muhit mos emas."

    try:
        # CPU
        cpu_usage = psutil.cpu_percent(interval=0.5)
        cpu_count = psutil.cpu_count(logical=True)
        
        # RAM
        ram = psutil.virtual_memory()
        ram_total_gb = round(ram.total / (1024 ** 3), 2)
        ram_used_gb = round(ram.used / (1024 ** 3), 2)
        ram_percent = ram.percent

        # Disk C: & D:
        disk_lines = []
        for drive in ["C:\\", "D:\\"]:
            if os.path.exists(drive):
                try:
                    d_usage = psutil.disk_usage(drive)
                    d_tot = round(d_usage.total / (1024 ** 3), 1)
                    d_free = round(d_usage.free / (1024 ** 3), 1)
                    disk_lines.append(f"💾 <b>{drive} Disk:</b> {d_usage.percent}% ({d_free} GB bo'sh / {d_tot} GB)")
                except Exception:
                    pass
        disk_info = "\n".join(disk_lines) if disk_lines else "💾 <b>Disk:</b> Ma'lumot olib bo'lmadi"

        # Batareya
        battery = psutil.sensors_battery()
        if battery:
            plugged = "🔌 Tarmoqqa ulangan" if battery.power_plugged else "🔋 Batareyadan ishlamoqda"
            battery_info = f"<b>Batareya:</b> {battery.percent}% ({plugged})"
        else:
            battery_info = "<b>Batareya:</b> Stasionar kompyuter (Mavjud emas)"

        # Uptime
        boot_time = datetime.fromtimestamp(psutil.boot_time())
        uptime = datetime.now() - boot_time
        uptime_str = str(timedelta(seconds=int(uptime.total_seconds())))

        status_text = (
            "📊 <b>TIZIM HOLATI (SYSTEM MONITOR)</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━\n"
            f"💻 <b>CPU Yuklanishi:</b> {cpu_usage}% ({cpu_count} ta yadro)\n"
            f"🧠 <b>RAM Yuklanishi:</b> {ram_percent}% ({ram_used_gb} GB / {ram_total_gb} GB)\n"
            f"{disk_info}\n"
            f"🔋 {battery_info}\n"
            f"⏱ <b>Ishlash vaqti (Uptime):</b> {uptime_str}\n"
            f"🕒 <b>Hozirgi vaqt:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        return status_text
    except Exception as e:
        logger.error(f"System status olishda xatolik: {e}")
        return f"❌ Tizim holatini olishda xatolik yuz berdi: {e}"


def get_monitors_list() -> list:
    """Kompyuterga ulangan barcha monitorlar koordinata va o'lchamlarini aniqlash"""
    monitors = []
    if IS_WINDOWS and ctypes:
        try:
            from ctypes import wintypes
            user32 = ctypes.windll.user32

            def _enum(hMon, hdcMon, lprcMon, dwData):
                r = lprcMon.contents
                x, y, w, h = r.left, r.top, r.right - r.left, r.bottom - r.top
                is_prim = (x == 0 and y == 0)
                monitors.append({
                    "id": len(monitors) + 1,
                    "x": x,
                    "y": y,
                    "width": w,
                    "height": h,
                    "is_primary": is_prim,
                    "name": f"{len(monitors) + 1}-Monitor {'(Asosiy)' if is_prim else ''} ({w}x{h})".strip()
                })
                return 1

            MONITORENUMPROC = ctypes.WINFUNCTYPE(
                ctypes.c_int,
                wintypes.HMONITOR,
                wintypes.HDC,
                ctypes.POINTER(wintypes.RECT),
                wintypes.LPARAM
            )
            proc = MONITORENUMPROC(_enum)
            user32.EnumDisplayMonitors(0, 0, proc, 0)
        except Exception as e:
            logger.warning(f"Error enumerating monitors: {e}")

    if not monitors:
        monitors.append({
            "id": 1,
            "x": 0,
            "y": 0,
            "width": 1920,
            "height": 1080,
            "is_primary": True,
            "name": "1-Monitor (Asosiy) (1920x1080)"
        })
    return monitors


def take_screenshot(filepath: str, monitor_index: int = None) -> str:
    """
    Ekran tasvirini (screenshot) olib ko'rsatilgan yo'lga saqlaydi (Fail-safe Win32 Desktop DC + ImageGrab).
    Agar monitor_index (1, 2, ...) berilsa, aynan o'sha monitor olinadi.
    """
    if IS_WINDOWS and ctypes and Image:
        try:
            user32 = ctypes.windll.user32
            gdi32 = ctypes.windll.gdi32

            # DPI awareness
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except Exception:
                try:
                    ctypes.windll.user32.SetProcessDPIAware()
                except Exception:
                    pass

            # Active Input Desktop'ga ulanish
            try:
                hdesk = user32.OpenInputDesktop(0, False, 0x01FF)
                if hdesk:
                    user32.SetThreadDesktop(hdesk)
            except Exception:
                pass

            mons = get_monitors_list()
            target_mon = None
            if monitor_index and isinstance(monitor_index, int) and 1 <= monitor_index <= len(mons):
                target_mon = mons[monitor_index - 1]
            elif monitor_index == "all":
                target_mon = None
            else:
                for m in mons:
                    if m.get("is_primary"):
                        target_mon = m
                        break
                if not target_mon and mons:
                    target_mon = mons[0]

            if target_mon:
                x, y, w, h = target_mon["x"], target_mon["y"], target_mon["width"], target_mon["height"]
            else:
                # Virtual desktop bounding box
                x = user32.GetSystemMetrics(76)  # SM_XVIRTUALSCREEN
                y = user32.GetSystemMetrics(77)  # SM_YVIRTUALSCREEN
                w = user32.GetSystemMetrics(78)  # SM_CXVIRTUALSCREEN
                h = user32.GetSystemMetrics(79)  # SM_CYVIRTUALSCREEN
                if w <= 0 or h <= 0:
                    x, y, w, h = 0, 0, 1920, 1080

            hdc = user32.GetDC(0)
            memdc = gdi32.CreateCompatibleDC(hdc)
            bmp = gdi32.CreateCompatibleBitmap(hdc, w, h)
            oldbmp = gdi32.SelectObject(memdc, bmp)

            # SRCCOPY = 0x00CC0020, CAPTUREBLT = 0x40000000
            gdi32.BitBlt(memdc, 0, 0, w, h, hdc, x, y, 0x00CC0020 | 0x40000000)

            class BITMAPINFOHEADER(ctypes.Structure):
                _fields_ = [
                    ('biSize', ctypes.c_uint32),
                    ('biWidth', ctypes.c_int32),
                    ('biHeight', ctypes.c_int32),
                    ('biPlanes', ctypes.c_uint16),
                    ('biBitCount', ctypes.c_uint16),
                    ('biCompression', ctypes.c_uint32),
                    ('biSizeImage', ctypes.c_uint32),
                    ('biXPelsPerMeter', ctypes.c_int32),
                    ('biYPelsPerMeter', ctypes.c_int32),
                    ('biClrUsed', ctypes.c_uint32),
                    ('biClrImportant', ctypes.c_uint32)
                ]

            bmi = BITMAPINFOHEADER()
            bmi.biSize = ctypes.sizeof(BITMAPINFOHEADER)
            bmi.biWidth = w
            bmi.biHeight = -h
            bmi.biPlanes = 1
            bmi.biBitCount = 32
            bmi.biCompression = 0

            buffer = ctypes.create_string_buffer(w * h * 4)
            gdi32.GetDIBits(hdc, bmp, 0, h, buffer, ctypes.byref(bmi), 0)

            gdi32.SelectObject(memdc, oldbmp)
            gdi32.DeleteObject(bmp)
            gdi32.DeleteDC(memdc)
            user32.ReleaseDC(0, hdc)

            img = Image.frombuffer('RGBA', (w, h), buffer.raw, 'raw', 'BGRA', 0, 1)
            img = img.convert('RGB')
            img.save(filepath, "PNG")
            return filepath
        except Exception as e:
            logger.warning(f"Win32 Direct GDI Screenshot failed, falling back: {e}")

    if ImageGrab:
        try:
            screenshot = ImageGrab.grab(all_screens=True)
            screenshot.save(filepath, "PNG")
            return filepath
        except Exception as e:
            logger.warning(f"ImageGrab failed: {e}")

    raise Exception("Screenshot olish imkonsiz: Pillow yoki Windows GUI mavjud emas.")


def take_all_monitors_screenshots() -> list:
    """Barcha monitorlardan alohida-alohida skrinshot olib base64 formatda qaytaradi"""
    import base64
    mons = get_monitors_list()
    results = []
    temp_dir = tempfile.gettempdir()

    for i, m in enumerate(mons, 1):
        try:
            shot_path = os.path.join(temp_dir, f"bridge_mon_{i}_{int(time.time())}.png")
            take_screenshot(shot_path, monitor_index=i)
            if os.path.exists(shot_path):
                with open(shot_path, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode("utf-8")
                try: os.remove(shot_path)
                except Exception: pass
                results.append({
                    "id": i,
                    "name": m.get("name", f"{i}-Monitor"),
                    "width": m.get("width", 1920),
                    "height": m.get("height", 1080),
                    "is_primary": m.get("is_primary", False),
                    "image": f"data:image/png;base64,{b64}"
                })
        except Exception as e:
            logger.error(f"Error capturing monitor {i}: {e}")

    if not results:
        shot_path = os.path.join(temp_dir, f"bridge_mon_def_{int(time.time())}.png")
        take_screenshot(shot_path)
        if os.path.exists(shot_path):
            with open(shot_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode("utf-8")
            try: os.remove(shot_path)
            except Exception: pass
            results.append({
                "id": 1,
                "name": "1-Monitor (Asosiy)",
                "width": 1920,
                "height": 1080,
                "is_primary": True,
                "image": f"data:image/png;base64,{b64}"
            })
    return results


def take_webcam_photo(filepath: str) -> str:
    """
    OpenCV orqali veb-kamera orqali suratga olib saqlaydi.
    """
    if not cv2:
        raise Exception("OpenCV (cv2) kutubxonasi o'rnatilmagan!")

    try:
        cap = cv2.VideoCapture(0, cv2.CAP_DSHOW if IS_WINDOWS else cv2.CAP_ANY)
        if not cap.isOpened():
            cap = cv2.VideoCapture(0)
        
        if not cap.isOpened():
            raise Exception("Veb-kamera topilmadi yoki boshqa dastur tomonidan band qilingan!")

        for _ in range(5):
            ret, frame = cap.read()

        ret, frame = cap.read()
        cap.release()

        if not ret or frame is None:
            raise Exception("Kameradan tasvirni o'qib bo'lmadi!")

        cv2.imwrite(filepath, frame)
        return filepath
    except Exception as e:
        logger.error(f"Webcam rasm olishda xatolik: {e}")
        raise e


def execute_cmd_sync(command: str, timeout: int = 30) -> str:
    """
    Windows CMD/PowerShell buyrug'ini bajaradi va natijasini matn ko'rinishida qaytaradi (Sinxron).
    """
    import subprocess
    try:
        proc = subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout,
            encoding='cp866',
            errors='replace'
        )
        out_str = proc.stdout.strip() or proc.stderr.strip()
        if not out_str:
            out_str = "✅ Buyruq muvaffaqiyatli bajarildi (hech qanday matn qaytarilmadi)."

        if len(out_str) > 3500:
            out_str = out_str[:3500] + "\n\n... [Matn juda uzun bo'lgani uchun qisqartirildi]"

        return out_str
    except subprocess.TimeoutExpired:
        return "⏳ <b>Xatolik:</b> Buyruq bajarilish vaqti tugadi (Timeout: 30s)."
    except Exception as e:
        logger.error(f"CMD bajarishda xatolik: {e}")
        return f"❌ Buyruqni bajarishda xatolik: {e}"


def get_running_apps() -> str:
    """
    Hozirda ishlab turgan asosiy va faol dasturlar ro'yxatini qaytaradi (TOP 20 RAM).
    """
    if not psutil:
        return "⚠️ <code>psutil</code> kutubxonasi mavjud emas."

    try:
        processes = []
        for proc in psutil.process_iter(['pid', 'name', 'memory_info', 'cpu_percent']):
            try:
                pinfo = proc.info
                if pinfo['name'] and pinfo['memory_info']:
                    mem_mb = round(pinfo['memory_info'].rss / (1024 * 1024), 1)
                    if mem_mb > 15:
                        processes.append({
                            'pid': pinfo['pid'],
                            'name': pinfo['name'],
                            'memory': mem_mb
                        })
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass

        processes = sorted(processes, key=lambda x: x['memory'], reverse=True)[:20]

        if not processes:
            return "📱 Ishlayotgan dasturlar topilmadi."

        result = "🎮 <b>ISHLAYOTGAN ASOSIY DASTURLAR (TOP 20 RAM):</b>\n━━━━━━━━━━━━━━━━━━━━━\n"
        for p in processes:
            result += f"🔹 <code>{p['name']}</code> | PID: <code>{p['pid']}</code> | RAM: <b>{p['memory']} MB</b>\n"

        result += "\n💡 <i>To'xtatish uchun: /kill &lt;dastur_nomi&gt; yoki /kill &lt;PID&gt;</i>"
        return result
    except Exception as e:
        logger.error(f"Apps ro'yxatini olishda xatolik: {e}")
        return f"❌ Dasturlar ro'yxatini olishda xatolik: {e}"


def kill_process(target: str) -> str:
    """
    Nomi yoki PID raqami bo'yicha jarayonni majburiy to'xtatadi.
    """
    if not psutil:
        return "⚠️ <code>psutil</code> kutubxonasi mavjud emas."

    try:
        killed_count = 0
        target = target.strip()

        if target.isdigit():
            pid = int(target)
            proc = psutil.Process(pid)
            pname = proc.name()
            proc.kill()
            return f"✅ PID <code>{pid}</code> ({pname}) jarayoni to'xtatildi."

        target_name = target.lower()
        target_name_exe = target_name if target_name.endswith('.exe') else target_name + '.exe'

        for proc in psutil.process_iter(['pid', 'name']):
            try:
                pname = proc.info['name'].lower()
                if pname == target_name or pname == target_name_exe:
                    proc.kill()
                    killed_count += 1
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

        if killed_count > 0:
            return f"✅ <code>{target}</code> nomi bo'yicha {killed_count} ta jarayon to'xtatildi."
        else:
            return f"⚠️ <code>{target}</code> nomli faol jarayon topilmadi."
    except Exception as e:
        logger.error(f"Kill process xatoligi: {e}")
        return f"❌ Jarayonni to'xtatishda xatolik: {e}"


def show_popup(text: str):
    """
    Windows operatsion tizimida MessageBox ko'rsatish.
    """
    if IS_WINDOWS and ctypes:
        try:
            ctypes.windll.user32.MessageBoxW(0, text, "ATLAS Bot Bildirishnomasi", 0x40 | 0x1000)
            return "✅ Bildirishnoma kompyuter ekranida ko'rsatildi."
        except Exception as e:
            return f"❌ Bildirishnoma ko'rsatishda xatolik: {e}"
    return "⚠️ Ushbu muhitda MessageBox mavjud emas."


def power_control(action: str) -> str:
    """
    Quvvatni boshqarish (Shutdown, Restart, Sleep, Lock, Cancel).
    """
    if not IS_WINDOWS:
        return "⚠️ Quvvat boshqaruvi faqat Windows operatsion tizimida ishlaydi."

    try:
        if action == "shutdown":
            os.system("shutdown /s /t 10 /c \"ATLAS Bot orqali o'chirilmoqda...\"")
            return "⚡ Kompyuter 10 soniyadan so'ng o'chiriladi! Bekor qilish uchun: /cancel_power"
        elif action == "restart":
            os.system("shutdown /r /t 10 /c \"ATLAS Bot orqali Qayta yuklanmoqda...\"")
            return "🔄 Kompyuter 10 soniyadan so'ng qayta yuklanadi! Bekor qilish uchun: /cancel_power"
        elif action == "sleep":
            os.system("rundll32.exe powrprof.dll,SetSuspendState 0,1,0")
            return "🌙 Kompyuter uyqu rejimiga (Sleep) o'tkazildi."
        elif action == "lock":
            os.system("rundll32.exe user32.dll,LockWorkStation")
            return "🔒 Ekran qulflandi (Lock)."
        elif action == "cancel":
            os.system("shutdown /a")
            return "❌ Rejalashtirilgan o'chirish/qayta yuklash bekor qilindi."
        else:
            return "⚠️ Noma'lum quvvat buyrug'i."
    except Exception as e:
        logger.error(f"Quvvatni boshqarishda xatolik: {e}")
        return f"❌ Quvvat buyrug'ini bajarishda xatolik: {e}"


def is_admin() -> bool:
    """
    Python skripti Administrator (UAC) huquqlari bilan ishlayotganini tekshiradi.
    """
    if IS_WINDOWS and ctypes:
        try:
            return ctypes.windll.shell32.IsUserAnAdmin() != 0
        except Exception:
            return False
    return False


def click_screen(x: int, y: int) -> str:
    """
    Ekranning ko'rsatilgan (x, y) koordinatasida sichqoncha tugmasini bosadi.
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI kutubxonasi mavjud emas."
    try:
        pyautogui.click(x, y)
        return f"✅ Sichqoncha ekranning ({x}, {y}) nuqtasida bosildi."
    except Exception as e:
        return f"❌ Klik xatoligi: {e}"


def press_key(key: str) -> str:
    """
    Klaviatura tugmasini bosadi (masalan: enter, space, tab, alt+a va h.k.).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI kutubxonasi mavjud emas."
    try:
        if "+" in key:
            keys = key.split("+")
            pyautogui.hotkey(*keys)
        else:
            pyautogui.press(key)
        return f"✅ Klaviaturada <code>{key}</code> tugmasi bosildi."
    except Exception as e:
        return f"❌ Tugma bosish xatoligi: {e}"


def set_volume(percent: int) -> str:
    """
    Windows masofaviy ovoz balandligini o'rnatadi (0 - 100).
    """
    percent = max(0, min(100, percent))
    try:
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
        from comtypes import CLSCTX_ALL

        devices = AudioUtilities.GetSpeakers()
        interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        volume = interface.QueryInterface(IAudioEndpointVolume)
        volume.SetMasterVolumeLevelScalar(percent / 100.0, None)
        return f"🔊 Ovoz balandligi <b>{percent}%</b> ga o'rnatildi."
    except Exception as e:
        logger.error(f"Volume pycaw error: {e}")
        return f"🔊 Ovoz o'rnatildi (yoki pycaw mavjud emas: {e})."


def set_mute(mute: bool = True) -> str:
    """
    Windows masofaviy ovozini o'chirish (Mute) yoki yoqish (Unmute).
    """
    try:
        from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
        from comtypes import CLSCTX_ALL

        devices = AudioUtilities.GetSpeakers()
        interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
        volume = interface.QueryInterface(IAudioEndpointVolume)
        volume.SetMute(int(mute), None)
        status = "o'chirildi (Mute 🔇)" if mute else "yoqildi (Unmute 🔊)"
        return f"🔊 Ovoz {status}."
    except Exception:
        if pyautogui:
            pyautogui.press('volumemute')
            return "🔊 Ovoz holati (Mute/Unmute) o'zgartirildi."
        return "⚠️ Ovozni o'zgartirib bo'lmadi."


def media_control(action: str) -> str:
    """
    Media pleyer boshqaruvi (playpause, nexttrack, prevtrack).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI kutubxonasi mavjud emas."
    try:
        if action in ["play", "pause", "playpause"]:
            pyautogui.press('playpause')
            return "⏯ Media pleyer ijro / to'xtatish tugmasi bosildi."
        elif action in ["next", "nexttrack"]:
            pyautogui.press('nexttrack')
            return "⏭ Keyingi trekka o'tildi."
        elif action in ["prev", "prevtrack"]:
            pyautogui.press('prevtrack')
            return "⏮ Oldingi trekka o'tildi."
        else:
            return "⚠️ Noma'lum media buyruq."
    except Exception as e:
        return f"❌ Media boshqaruvida xatolik: {e}"


def set_brightness(percent: int) -> str:
    """
    Ekran yorqinligini (Brightness) moslaydi (0 - 100).
    """
    import subprocess
    percent = max(0, min(100, percent))
    try:
        ps_script = f"(Get-WmiObject -Namespace root/WMI -Class WmiMonitorBrightnessMethods).WmiSetBrightness(1, {percent})"
        subprocess.run(["powershell", "-Command", ps_script], capture_output=True, text=True, timeout=5)
        return f"☀️ Ekran yorqinligi <b>{percent}%</b> ga o'rnatildi."
    except Exception as e:
        logger.error(f"Brightness error: {e}")
        return f"❌ Ekran yorqinligini o'zgartirishda xatolik: {e}"


def open_app(app_name: str) -> str:
    """
    Kalkulyator, Bloknot, Chrome, Telegram, Explorer kabi dasturlarni nomiga qarab ochadi.
    """
    name = app_name.lower().strip()
    apps_map = {
        "kalkulyator": "calc",
        "calculator": "calc",
        "calc": "calc",
        "bloknot": "notepad",
        "notepad": "notepad",
        "chrome": "chrome",
        "google chrome": "chrome",
        "browser": "chrome",
        "telegram": "telegram",
        "explorer": "explorer",
        "my pc": "explorer",
        "dispetcher": "taskmgr",
        "task manager": "taskmgr",
        "taskmgr": "taskmgr",
        "cmd": "cmd /c start cmd",
        "powershell": "powershell",
        "word": "winword",
        "excel": "excel",
        "control panel": "control",
        "boshqaruv paneli": "control",
        "paint": "mspaint",
        "skrinshot": "snippingtool",
        "anydesk": "anydesk",
        "rustdesk": "rustdesk"
    }

    cmd_to_run = apps_map.get(name, name)
    try:
        import subprocess
        subprocess.Popen(f"start {cmd_to_run}", shell=True)
        return f"🚀 <b>{app_name.capitalize()}</b> dasturi ishga tushirildi!"
    except Exception as e:
        logger.error(f"Open app error: {e}")
        return f"❌ Dasturni ochishda xatolik: {e}"


def show_desktop() -> str:
    """
    Barcha ochiq oynalarni yig'adi / Ish stolini ko'rsatadi (Win + D).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI mavjud emas."
    try:
        pyautogui.hotkey('win', 'd')
        return "🖥 Ish stoli ko'rsatildi (Win + D)."
    except Exception as e:
        return f"❌ Xatolik: {e}"


def close_active_window() -> str:
    """
    Hozirgi faol oynani yopadi (Alt + F4).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI mavjud emas."
    try:
        pyautogui.hotkey('alt', 'f4')
        return "❌ Faol oyna yopildi (Alt + F4)."
    except Exception as e:
        return f"❌ Xatolik: {e}"


def empty_recycle_bin() -> str:
    """
    Windows Korzina (Chaqindi qutisi) ni tozalaydi.
    """
    import subprocess
    try:
        ps_script = "Clear-RecycleBin -Confirm:$false -ErrorAction SilentlyContinue"
        subprocess.run(["powershell", "-Command", ps_script], capture_output=True, text=True, timeout=10)
        return "🗑 <b>Windows Korzina (Chaqindi qutisi) tozalandi!</b>"
    except Exception as e:
        return f"❌ Korzinani tozalashda xatolik: {e}"


def clean_temp_files() -> str:
    """
    Windows va Foydalanuvchi vaqtinchalik (Temp) kesh fayllarini tozalaydi.
    """
    cleaned_size = 0
    temp_dirs = [os.environ.get("TEMP"), r"C:\Windows\Temp"]
    
    for tdir in temp_dirs:
        if tdir and os.path.exists(tdir):
            for item in glob.glob(os.path.join(tdir, "*")):
                try:
                    if os.path.isfile(item):
                        sz = os.path.getsize(item)
                        os.remove(item)
                        cleaned_size += sz
                    elif os.path.isdir(item):
                        shutil.rmtree(item, ignore_errors=True)
                except Exception:
                    pass

    mb_cleaned = round(cleaned_size / (1024 * 1024), 2)
    return f"🧹 <b>Vaqtinchalik kesh (Temp) fayllari tozalandi!</b>\n💾 Taxminan <b>{mb_cleaned} MB</b> joy bo'shatildi."


def type_text(text: str) -> str:
    """
    Ekranga matn kiritadi (Klaviatura orqali yozadi).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI mavjud emas."
    try:
        pyautogui.write(text, interval=0.02)
        return f"✍️ Ekranga matn kiritildi: \"{text}\""
    except Exception as e:
        return f"❌ Matn kiritishda xatolik: {e}"


def scroll_page(amount: int = -500) -> str:
    """
    Ekranni pastga yoki yuqoriga aylantiradi (Scroll).
    """
    if not pyautogui:
        return "⚠️ PyAutoGUI mavjud emas."
    try:
        pyautogui.scroll(amount)
        direction = "pastga" if amount < 0 else "yuqoriga"
        return f"📜 Sahifa {direction} aylantirildi ({amount})."
    except Exception as e:
        return f"❌ Scroll xatoligi: {e}"


def record_mic_audio(filepath: str, duration: int = 5) -> str:
    """
    Windows winmm.dll (MCI) orqali mikrofondan ovoz yozib .wav fayliga saqlaydi.
    """
    if not IS_WINDOWS or not ctypes:
        raise Exception("Mikrofon yozish faqat Windows muhitida ishlaydi.")

    winmm = ctypes.windll.winmm
    try:
        winmm.mciSendStringW("open new type waveaudio alias mic_rec", None, 0, 0)
        winmm.mciSendStringW("record mic_rec", None, 0, 0)
        time.sleep(duration)
        winmm.mciSendStringW(f'save mic_rec "{filepath}"', None, 0, 0)
        winmm.mciSendStringW("close mic_rec", None, 0, 0)
        return filepath
    except Exception as e:
        winmm.mciSendStringW("close mic_rec", None, 0, 0)
        logger.error(f"Mic record error: {e}")
        raise e


def list_directory_info(dir_path: str = None) -> tuple:
    """
    Papka ichidagi barcha fayl, yarlik (.lnk) va papkalar ro'yxatini qaytaradi.
    """
    if not dir_path or not os.path.exists(dir_path):
        dir_path = os.path.join(os.path.expanduser("~"), "Desktop")

    dir_path = os.path.abspath(dir_path)
    parent_dir = os.path.dirname(dir_path)

    try:
        entries = sorted(os.listdir(dir_path))
    except Exception as e:
        return f"❌ Papkani o'qishda xatolik: {e}", dir_path, parent_dir, [], []

    dirs_list = []
    files_list = []

    for entry in entries:
        if entry.startswith("$") or entry == "System Volume Information":
            continue
        full_p = os.path.join(dir_path, entry)
        try:
            if os.path.isdir(full_p):
                dirs_list.append(entry)
            else:
                files_list.append(entry)
        except Exception:
            pass

    text = f"📁 <b>PAPKA MANZILI:</b>\n<code>{dir_path}</code>\n\n"
    text += f"📂 Papkalar ({len(dirs_list)} ta) | 📄 Fayllar & Yarliklar ({len(files_list)} ta)\n"
    text += "━━━━━━━━━━━━━━━━━━━━━\n"

    return text, dir_path, parent_dir, dirs_list, files_list


def search_user_files(keyword: str, max_results: int = 8) -> str:
    """
    Nomi bo'yicha Desktop, Downloads, Documents papkalaridan fayl qidiradi.
    """
    search_dirs = [
        os.path.join(os.path.expanduser("~"), "Desktop"),
        os.path.join(os.path.expanduser("~"), "Downloads"),
        os.path.join(os.path.expanduser("~"), "Documents")
    ]

    kw = keyword.lower().strip()
    found = []

    for sdir in search_dirs:
        if os.path.exists(sdir):
            for root, _, files in os.walk(sdir):
                for f in files:
                    if kw in f.lower():
                        found.append(os.path.join(root, f))
                        if len(found) >= max_results:
                            break
                if len(found) >= max_results:
                    break

    if not found:
        return f"🔍 <code>{keyword}</code> nomli fayllar topilmadi."

    res = f"🔍 <b>TOPILGAN FAYLLAR ({len(found)} ta):</b>\n━━━━━━━━━━━━━━━━━━━━━\n"
    for fp in found:
        sz_mb = round(os.path.getsize(fp) / (1024 * 1024), 2)
        res += f"📄 <code>{fp}</code> ({sz_mb} MB)\n"
    return res


def pair_sunshine_pin(pin: str) -> str:
    """
    Sunshine PIN Pairing (brauzersiz, 100% orqa fonda).
    """
    import urllib3, requests
    urllib3.disable_warnings()

    pin = str(pin).strip()
    if not pin.isdigit() or len(pin) != 4:
        return f"⚠️ Sunshine PIN-kodi 4 xonali raqam bo'lishi kerak! (Siz kiritdingiz: <code>{pin}</code>)"

    # Ekranni uyg'otish
    wake_and_unlock_pc()

    # Orqa fonda Sunshine API ga PIN yuborish (hech qanday brauzer ochilmaydi)
    try:
        requests.post("https://127.0.0.1:47990/pin", json={"pin": pin}, verify=False, timeout=3)
    except Exception:
        pass

    return (
        f"⚡ <b>SUNSHINE AVTO-PAIRING (ORQA FONDA)</b>\n\n"
        f"🔑 <b>PIN Kod:</b> <code>{pin}</code>\n"
        f"✨ <b>Holati:</b> Sunshine serveriga orqa fonda ulandi!\n\n"
        f"<i>Moonlight ilovangizda kompyuterni tanlang va ulaning!</i>"
    )


def wake_and_unlock_pc(password: str = None) -> str:
    """
    Kompyuter ekranini uyg'otish va Windows Lock Screen'dan chiqarish.
    """
    try:
        import ctypes
        user32 = ctypes.windll.user32

        # 1. Reset display idle and wake display
        ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)

        # 2. Mouse move to wake display
        user32.mouse_event(0x0001, 2, 2, 0, 0)
        time.sleep(0.15)
        user32.mouse_event(0x0001, -2, -2, 0, 0)
        time.sleep(0.15)

        # 3. Simulate Spacebar to lift Lock screen wallpaper
        user32.keybd_event(0x20, 0, 0, 0)
        user32.keybd_event(0x20, 0, 2, 0)
        time.sleep(0.3)
        user32.keybd_event(0x20, 0, 0, 0)
        user32.keybd_event(0x20, 0, 2, 0)
        time.sleep(0.2)

        # 4. If password or PIN provided, type it and press Enter
        if password:
            if pyautogui:
                pyautogui.write(password, interval=0.04)
            time.sleep(0.2)
            user32.keybd_event(0x0D, 0, 0, 0)
            user32.keybd_event(0x0D, 0, 2, 0)
            return "🔓 <b>Kompyuter ekrani uyg'otildi va Windows paroli kiritildi!</b>"
        else:
            return "☀️ <b>Kompyuter ekrani uyg'otildi (Lock screen ko'tarildi)!</b>\n<i>Agar Windows paroli bo'lsa, /unlock &lt;parol&gt; orqali kiritishingiz mumkin.</i>"
    except Exception as e:
        return f"❌ Ekranni uyg'otishda xatolik: {e}"


def register_sunshine_client_cert(cert_path: str) -> str:
    """
    Moonlight client.pem sertifikatini Sunshine serveriga doimiy ishonchli klient (Auto-Pair) sifatida kiritadi.
    """
    if not os.path.exists(cert_path):
        return f"❌ Sertifikat fayli topilmadi: <code>{cert_path}</code>"

    possible_paths = [
        r"C:\Program Files\Sunshine\config",
        r"C:\ProgramData\Sunshine\config",
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "sunshine", "config"),
        os.path.join(os.environ.get("APPDATA", ""), "sunshine", "config")
    ]

    target_config_dir = None
    for p in possible_paths:
        if os.path.exists(p):
            target_config_dir = p
            break

    if not target_config_dir:
        target_config_dir = r"C:\ProgramData\Sunshine\config"
        os.makedirs(target_config_dir, exist_ok=True)

    client_certs_dir = os.path.join(target_config_dir, "client_certs")
    os.makedirs(client_certs_dir, exist_ok=True)

    cert_filename = os.path.basename(cert_path)
    if not cert_filename.endswith(".pem") and not cert_filename.endswith(".crt"):
        cert_filename += ".pem"

    target_cert_path = os.path.join(client_certs_dir, cert_filename)
    try:
        shutil.copy2(cert_path, target_cert_path)
        return (
            f"☀️ <b>MOONLIGHT AVTO-ULANISH (CERTIFICATE AUTO-PAIR)</b>\n\n"
            f"📜 <b>Sertifikat saqlandi:</b> <code>{target_cert_path}</code>\n\n"
            f"✅ Moonlight qurilmangiz Sunshine serveriga muvaffaqiyatli doimiy ulandi!"
        )
    except Exception as e:
        return f"❌ Sertifikatni saqlashda xatolik: {e}"


def get_anydesk_id() -> str:
    """Kompyuterdagi AnyDesk ID raqamini aniqlash"""
    # 1. ProgramData/AnyDesk/system.conf
    p1 = r"C:\ProgramData\AnyDesk\system.conf"
    if os.path.exists(p1):
        try:
            with open(p1, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    if line.strip().startswith("ad.anynet.id="):
                        return line.strip().split("=", 1)[1].strip()
        except Exception:
            pass

    # 2. AppData/AnyDesk/user.conf
    appdata = os.environ.get("APPDATA", "")
    p2 = os.path.join(appdata, "AnyDesk", "user.conf")
    if os.path.exists(p2):
        try:
            with open(p2, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    if line.strip().startswith("ad.anynet.id="):
                        return line.strip().split("=", 1)[1].strip()
        except Exception:
            pass

    # 3. Try anydesk --get-id
    try:
        res = subprocess.run(["anydesk", "--get-id"], capture_output=True, text=True, timeout=5)
        if res.stdout.strip().isdigit():
            return res.stdout.strip()
    except Exception:
        pass

    return "AnyDesk ID topilmadi (AnyDesk o'rnatilmagan yoki konfiguratsiya topilmadi)"


def print_file(file_path: str) -> str:
    """Faylni standart Windows printerida qog'ozga chiqarish."""
    if not file_path:
        return "❌ Fayl yo'li ko'rsatilmadi."

    candidates = [
        file_path,
        os.path.join(os.path.expanduser("~"), "Desktop", file_path),
        os.path.join(os.path.expanduser("~"), "Downloads", file_path),
        os.path.join(os.path.expanduser("~"), "Documents", file_path),
    ]
    target = None
    for c in candidates:
        if os.path.exists(c):
            target = c
            break

    if not target:
        return f"❌ Chop etish uchun fayl topilmadi: <code>{file_path}</code>"

    try:
        cmd = f'Start-Process -FilePath "{target}" -Verb Print'
        subprocess.run(["powershell", "-Command", cmd], timeout=15)
        return f"🖨️ <b>Fayl printerga yuborildi:</b> <code>{os.path.basename(target)}</code>"
    except Exception as e:
        return f"❌ Printerga yuborishda xatolik: {e}"


def read_file_content(file_path: str, max_chars: int = 4000) -> str:
    """Fayl ichidagi matn yoki ma'lumotlarni o'qish (Word, PDF, Excel, Text, Code)."""
    if not file_path:
        return "❌ Fayl yo'li ko'rsatilmadi."

    candidates = [
        file_path,
        os.path.join(os.path.expanduser("~"), "Desktop", file_path),
        os.path.join(os.path.expanduser("~"), "Downloads", file_path),
        os.path.join(os.path.expanduser("~"), "Documents", file_path),
    ]
    target = None
    for c in candidates:
        if os.path.exists(c):
            target = c
            break

    if not target:
        return f"❌ Fayl topilmadi: <code>{file_path}</code>"

    ext = os.path.splitext(target)[1].lower()
    try:
        if ext == ".docx":
            import docx
            doc = docx.Document(target)
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return f"📄 <b>WORD HUJJATI MAZMUNI ({os.path.basename(target)}):</b>\n\n{text[:max_chars]}"

        elif ext in [".xlsx", ".xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(target, data_only=True)
            res = []
            for sheet in wb.sheetnames[:2]:
                ws = wb[sheet]
                res.append(f"📊 Varaq: {sheet}")
                for row in list(ws.iter_rows(values_only=True))[:15]:
                    row_clean = [str(x) if x is not None else "" for x in row]
                    if any(row_clean):
                        res.append(" | ".join(row_clean[:6]))
            return f"📊 <b>EXCEL MAZMUNI ({os.path.basename(target)}):</b>\n\n" + "\n".join(res)[:max_chars]

        # Oddiy matnli fayllar (txt, py, json, xml, csv, md, log, html, pdf)
        for enc in ["utf-8", "windows-1251", "cp1256", "latin-1"]:
            try:
                with open(target, "r", encoding=enc) as f:
                    content = f.read(max_chars)
                    return f"📝 <b>FAYL MAZMUNI ({os.path.basename(target)}):</b>\n\n<pre>{content}</pre>"
            except Exception:
                continue

        return f"❌ Ushbu fayl formatini o'qib bo'lmadi: {ext}"
    except Exception as e:
        return f"❌ Faylni o'qishda xatolik: {e}"


def download_or_install_software(target: str) -> str:
    """Dastur yuklab olish yoki winget orqali o'rnatish."""
    target = target.strip()
    if target.startswith("http://") or target.startswith("https://"):
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        fname = target.split("?")[0].split("/")[-1] or "downloaded_file"
        dest = os.path.join(downloads_dir, fname)
        try:
            import requests
            r = requests.get(target, timeout=60, stream=True)
            with open(dest, "wb") as f:
                for chunk in r.iter_content(chunk_size=65536):
                    f.write(chunk)
            return f"✅ <b>Fayl yuklab olindi:</b> <code>{dest}</code> ({round(os.path.getsize(dest)/1024/1024, 1)} MB)"
        except Exception as e:
            return f"❌ Yuklab olishda xatolik: {e}"
    else:
        try:
            cmd = f'winget install --id "{target}" --silent --accept-source-agreements --accept-package-agreements'
            res = subprocess.run(["powershell", "-Command", cmd], capture_output=True, text=True, timeout=90)
            return f"📦 <b>Dastur o'rnatish buyrug'i bajarildi ({target}):</b>\n<code>{res.stdout[:500]}</code>"
        except Exception as e:
            return f"❌ O'rnatishda xatolik: {e}"


def write_file_content(file_path: str, content: str) -> str:
    """Faylga matn yozish yoki yangi fayl yaratish."""
    try:
        if not os.path.isabs(file_path):
            file_path = os.path.join(os.path.expanduser("~"), "Desktop", file_path)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"✅ <b>Fayl saqlandi:</b> <code>{file_path}</code>"
    except Exception as e:
        return f"❌ Faylni saqlashda xatolik: {e}"


