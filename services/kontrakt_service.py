# ============================================================
#  services/kontrakt_service.py
#  Kontraktlar & Bank Debitorkasi Taqqoslash va Yangilash Xizmati
#  Formulalarni 100% buzmasdan saqlash, 3x Ultra HD Screenshotlar & Supabase Storage
# ============================================================

import os
import io
import re
import html
import zipfile
import tempfile
import uuid
from datetime import datetime, timedelta
import openpyxl
from PIL import Image, ImageDraw, ImageFont
from fuzzywuzzy import fuzz

from services.supabase_storage import upload_document_to_supabase
from services.atlas_db import log_audit

# Directory for storing generated contract assets (Serverless safe)
is_serverless = os.environ.get("VERCEL") == "1" or os.environ.get("AWS_LAMBDA_FUNCTION_NAME") is not None or os.path.exists("/tmp")

if is_serverless:
    CONTRACT_STORAGE_DIR = os.path.join(tempfile.gettempdir(), 'saved_documents', 'contracts')
else:
    CONTRACT_STORAGE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'saved_documents', 'contracts')

try:
    os.makedirs(CONTRACT_STORAGE_DIR, exist_ok=True)
except Exception:
    CONTRACT_STORAGE_DIR = os.path.join(tempfile.gettempdir(), 'saved_documents', 'contracts')
    try:
        os.makedirs(CONTRACT_STORAGE_DIR, exist_ok=True)
    except Exception:
        pass



# ============================================================
# 1. NAME NORMALIZATION & HELPERS
# ============================================================

def cyrillic_to_latin(text):
    """Kirill alifbosidagi ismlarni Lotin alifbosiga o'tkazish"""
    if not text:
        return ""
    text = str(text)
    trans = {
        'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'j','з':'z','и':'i',
        'й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
        'у':'u','ф':'f','х':'h','ц':'c','ч':'ch','ш':'sh','щ':'sh','ъ':'','ы':'y','ь':'',
        'э':'e','ю':'yu','я':'ya','ў':'o','ғ':'g','қ':'q','ҳ':'h',
        'А':'a','Б':'b','В':'v','Г':'g','Д':'d','Е':'e','Ё':'yo','Ж':'j','З':'z','И':'i',
        'Й':'y','К':'k','Л':'l','М':'m','Н':'n','О':'o','П':'p','Р':'r','С':'s','Т':'t',
        'У':'u','Ф':'f','Х':'h','Ц':'c','Ч':'ch','Ш':'sh','Щ':'sh','Ъ':'','Ы':'y','Ь':'',
        'Э':'e','Ю':'yu','Я':'ya','Ў':'o','Ғ':'g','Қ':'q','Ҳ':'h'
    }
    for k, v in trans.items():
        text = text.replace(k, v)
    return text


def ismlarni_standartlash(ism):
    """Ismlarni taqqoslash uchun tozalash va standartlash"""
    if not ism:
        return ""
    ism = cyrillic_to_latin(ism).strip().lower()
    ism = ism.replace("`", "").replace("ʻ", "").replace("‘", "").replace("’", "").replace("'", "")
    ism = ism.replace("о‘", "o").replace("o‘", "o").replace("o'", "o").replace("о'", "o")
    ism = ism.replace("g‘", "g").replace("g'", "g").replace("г‘", "g")
    ism = ism.replace("ch", "c").replace("sh", "s").replace("x", "h").replace("ya", "a").replace("yu", "u")
    return "".join(c for c in ism if c.isalpha() or c.isspace())


def fmt_num(val):
    if val is None or val == "" or val == "-":
        return "-"
    try:
        fval = float(val)
        if fval == 0:
            return "0"
        return f"{int(round(fval)):,}".replace(",", " ")
    except Exception:
        return str(val)


def cleanup_old_contract_temp_files(max_age_seconds: int = 600):
    """
    10 daqiqadan (600 soniya) oshgan barcha vaqtinchalik kontrakt screenshotlari,
    rasmlar va ZIP fayllarini avtomatik tozalab xotirani bo'shatadi.
    """
    import shutil
    import time

    now = time.time()
    dirs_to_clean = [
        CONTRACT_STORAGE_DIR,
        tempfile.gettempdir()
    ]

    for base_dir in dirs_to_clean:
        if not os.path.exists(base_dir):
            continue
        try:
            for item in os.listdir(base_dir):
                item_path = os.path.join(base_dir, item)
                if any(k in item for k in ['screenshots_', 'screenshot_', 'xulosa_', 'Guruhlar_Screenshotlari_', 'KONTRAKTLAR', 'debitorka']):
                    try:
                        mtime = os.path.getmtime(item_path)
                        if now - mtime > max_age_seconds:
                            if os.path.isdir(item_path):
                                shutil.rmtree(item_path, ignore_errors=True)
                            else:
                                os.remove(item_path)
                    except Exception:
                        pass
        except Exception:
            pass


def get_font(size, bold=True):
    bundled_tnr = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'TimesNewRomanBold.ttf')
    if os.path.exists(bundled_tnr):
        try:
            return ImageFont.truetype(bundled_tnr, size)
        except Exception:
            pass

    font_paths = [
        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'fonts', 'AppBoldFont.ttf'),
        r'C:\Windows\Fonts\timesbd.ttf' if bold else r'C:\Windows\Fonts\times.ttf',
        r'C:\Windows\Fonts\arialbd.ttf' if bold else r'C:\Windows\Fonts\arial.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf' if bold else '/usr/share/fonts/truetype/freefont/FreeSerif.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf' if bold else '/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf',
    ]
    for path in font_paths:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                pass
    try:
        return ImageFont.load_default(size=size)
    except Exception:
        return ImageFont.load_default()


# ============================================================
# 2. IMAGE TABLE GENERATORS (3X ULTRA HD TIMES NEW ROMAN)
# ============================================================

def generate_group_table_image(group_name, date_str, rows_data, output_path, header_bg_color=(0, 112, 192)):
    """Times New Roman shriftida pixel-perfect HD screenshot hosil qilish"""
    S = 3  # 3x Ultra HD Resolution
    col_w = [int(w * S) for w in [110, 60, 520, 310, 230, 250]]

    headers = [
        'GURUHI',
        '№',
        'Familiiyasi Ismi va Sharfi',
        'Shu vaqtgacha bo\'lishi\nkerak bo\'lgan to\'lov',
        'Jami',
        'Shu vaqtgacha\nqarzi'
    ]

    table_w = sum(col_w)
    title_h = int(55 * S)
    header_h = int(75 * S)
    row_h = int(42 * S)
    summary_h = int(50 * S)

    num_rows = len(rows_data)
    table_h = title_h + header_h + (num_rows * row_h) + summary_h

    margin = int(24 * S)
    img_w = table_w + 2 * margin
    img_h = table_h + 2 * margin

    img = Image.new('RGB', (img_w, img_h), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    font_cell = get_font(int(24 * S), bold=True)
    font_title = get_font(int(30 * S), bold=True)
    font_header = get_font(int(24 * S), bold=True)
    font_summary = get_font(int(25 * S), bold=True)

    grid_col = (0, 0, 0)
    border_w = int(3 * S // 2)

    ox, oy = margin, margin

    # 1. Title Row
    draw.rectangle([ox, oy, ox + table_w, oy + title_h], fill=(255, 255, 255), outline=grid_col, width=border_w)
    title_str = f'Yangilangan sanasi:   {date_str}'
    bbox = font_title.getbbox(title_str)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((ox + (table_w - tw) // 2, oy + (title_h - th) // 2 - bbox[1]), title_str, fill=(0, 0, 0), font=font_title)

    # 2. Header Row
    curr_y = oy + title_h
    curr_x = ox
    for idx, (h_text, w) in enumerate(zip(headers, col_w)):
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + header_h], fill=header_bg_color, outline=grid_col, width=border_w)
        lines = h_text.split('\n')
        total_lines_h = len(lines) * int(28 * S)
        line_y = curr_y + (header_h - total_lines_h) // 2
        for line in lines:
            bbox = font_header.getbbox(line)
            tw = bbox[2] - bbox[0]
            tx = curr_x + (w - tw) // 2
            draw.text((tx, line_y - bbox[1]), line, fill=(255, 255, 255), font=font_header)
            line_y += int(28 * S)
        curr_x += w

    # 3. Data Rows
    curr_y += header_h
    tot_kerak, tot_jami, tot_qarzi_musbat = 0.0, 0.0, 0.0

    for row_idx, rdata in enumerate(rows_data):
        curr_x = ox
        no_val = str(rdata.get('no', row_idx + 1))
        fio_val = str(rdata.get('fio', ''))
        kerak_num = float(rdata.get('kerak', 0.0))
        jami_num = float(rdata.get('jami', 0.0))
        qarzi_num = float(rdata.get('qarzi', 0.0))

        tot_kerak += kerak_num
        tot_jami += jami_num
        if qarzi_num > 0:
            tot_qarzi_musbat += qarzi_num

        cells_info = [
            (group_name, 'center', (255, 255, 255), (0, 0, 0)),
            (no_val, 'center', (255, 255, 255), (0, 0, 0)),
            (fio_val, 'left', (255, 255, 255), (0, 0, 0)),
            (fmt_num(kerak_num), 'right', (255, 255, 255), (0, 0, 0)),
            (fmt_num(jami_num), 'right', (255, 255, 255), (0, 0, 0)),
        ]

        if qarzi_num > 0:
            debt_bg = (255, 199, 206)
            debt_fg = (156, 0, 6)
        else:
            debt_bg = (198, 239, 206)
            debt_fg = (0, 97, 0)

        cells_info.append((fmt_num(qarzi_num), 'right', debt_bg, debt_fg))

        for (c_text, align, bg_col, fg_col), w in zip(cells_info, col_w):
            draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + row_h], fill=bg_col, outline=grid_col, width=border_w)
            bbox = font_cell.getbbox(c_text)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]

            if align == 'center':
                tx = curr_x + (w - tw) // 2
            elif align == 'right':
                tx = curr_x + w - tw - int(16 * S)
            else:
                tx = curr_x + int(16 * S)

            ty = curr_y + (row_h - th) // 2 - bbox[1]
            draw.text((tx, ty), c_text, fill=fg_col, font=font_cell)
            curr_x += w

        curr_y += row_h

    # 4. Summary Row (JAMI)
    curr_x = ox
    jami_w = col_w[0] + col_w[1] + col_w[2]
    draw.rectangle([curr_x, curr_y, curr_x + jami_w, curr_y + summary_h], fill=header_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox('JAMI')
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (jami_w - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), 'JAMI', fill=(255, 255, 255), font=font_summary)
    curr_x += jami_w

    summary_cols = [
        (fmt_num(tot_kerak), col_w[3]),
        (fmt_num(tot_jami), col_w[4]),
        (fmt_num(tot_qarzi_musbat), col_w[5]),
    ]
    for c_text, w in summary_cols:
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + summary_h], fill=header_bg_color, outline=grid_col, width=border_w)
        bbox = font_summary.getbbox(c_text)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        tx = curr_x + w - tw - int(16 * S)
        ty = curr_y + (summary_h - th) // 2 - bbox[1]
        draw.text((tx, ty), c_text, fill=(255, 255, 255), font=font_summary)
        curr_x += w

    img.save(output_path, 'PNG', quality=100)
    return output_path


def generate_xulosa_table_image(xulosa_rows, output_path):
    """Guruh rahbarlari va umumiy jamlanma XULOSA rasm jadvalini yaratish"""
    S = 3  # 3x Ultra HD Resolution
    col_w = [int(w * S) for w in [280, 160, 180, 240]]
    headers = ['Guruh rahbari', 'Guruh', 'Talabalar soni', 'Qarzdorligi']

    table_w = sum(col_w)
    header_h = int(65 * S)
    row_h = int(42 * S)
    summary_h = int(50 * S)

    num_rows = len(xulosa_rows)
    table_h = header_h + (num_rows * row_h) + summary_h

    margin = int(24 * S)
    img_w = table_w + 2 * margin
    img_h = table_h + 2 * margin

    img = Image.new('RGB', (img_w, img_h), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    font_cell = get_font(int(21 * S), bold=True)
    font_header = get_font(int(22 * S), bold=True)
    font_summary = get_font(int(23 * S), bold=True)

    grid_col = (0, 0, 0)
    border_w = int(3 * S // 2)
    ox, oy = margin, margin

    # 1. Header Row (To'q olovrang fon: #ED7D31)
    header_bg_color = (237, 125, 49)
    curr_x = ox
    curr_y = oy
    for idx, (h_text, w) in enumerate(zip(headers, col_w)):
        draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + header_h], fill=header_bg_color, outline=grid_col, width=border_w)
        bbox = font_header.getbbox(h_text)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        tx = curr_x + (w - tw) // 2
        ty = curr_y + (header_h - th) // 2 - bbox[1]
        draw.text((tx, ty), h_text, fill=(255, 255, 255), font=font_header)
        curr_x += w

    # 2. Data Rows
    curr_y += header_h
    tot_talabalar = 0
    tot_qarzdorlik = 0.0

    for rdata in xulosa_rows:
        curr_x = ox
        rahbar = str(rdata.get('rahbar', ''))
        guruh = str(rdata.get('guruh', ''))
        soni = int(rdata.get('soni', 0))
        qarz = float(rdata.get('qarz', 0.0))

        tot_talabalar += soni
        tot_qarzdorlik += qarz

        qarz_bg = (198, 239, 206) if qarz <= 0 else (255, 255, 255)

        cells_info = [
            (rahbar, 'center', (255, 255, 255), (0, 0, 0)),
            (guruh, 'center', (255, 255, 255), (0, 0, 0)),
            (str(soni), 'center', (255, 255, 255), (0, 0, 0)),
            (fmt_num(qarz), 'right', qarz_bg, (0, 0, 0)),
        ]

        for (c_text, align, bg_col, fg_col), w in zip(cells_info, col_w):
            draw.rectangle([curr_x, curr_y, curr_x + w, curr_y + row_h], fill=bg_col, outline=grid_col, width=border_w)
            bbox = font_cell.getbbox(c_text)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]

            if align == 'center':
                tx = curr_x + (w - tw) // 2
            elif align == 'right':
                tx = curr_x + w - tw - int(16 * S)
            else:
                tx = curr_x + int(16 * S)

            ty = curr_y + (row_h - th) // 2 - bbox[1]
            draw.text((tx, ty), c_text, fill=fg_col, font=font_cell)
            curr_x += w

        curr_y += row_h

    # 3. Bottom Summary Row (JAMI - Qizil fon)
    curr_x = ox
    jami_w = col_w[0] + col_w[1]
    summary_bg_color = (255, 0, 0)

    draw.rectangle([curr_x, curr_y, curr_x + jami_w, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox('Jami')
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (jami_w - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), 'Jami', fill=(0, 0, 0), font=font_summary)
    curr_x += jami_w

    w_soni = col_w[2]
    draw.rectangle([curr_x, curr_y, curr_x + w_soni, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox(str(tot_talabalar))
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text((curr_x + (w_soni - tw) // 2, curr_y + (summary_h - th) // 2 - bbox[1]), str(tot_talabalar), fill=(0, 0, 0), font=font_summary)
    curr_x += w_soni

    w_qarz = col_w[3]
    qarz_str = fmt_num(tot_qarzdorlik)
    draw.rectangle([curr_x, curr_y, curr_x + w_qarz, curr_y + summary_h], fill=summary_bg_color, outline=grid_col, width=border_w)
    bbox = font_summary.getbbox(qarz_str)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    tx = curr_x + w_qarz - tw - int(16 * S)
    ty = curr_y + (summary_h - th) // 2 - bbox[1]
    draw.text((tx, ty), qarz_str, fill=(0, 0, 0), font=font_summary)

    img.save(output_path, 'PNG', quality=100)
    return output_path


# ============================================================
# 3. ANALYZE BASE EXCEL FILE
# ============================================================

def analyze_baza_excel(baza_path):
    """Asosiy baza Excel faylini tekshirib, oxirgi yangilangan sana va guruhlarni aniqlash"""
    try:
        wb = openpyxl.load_workbook(baza_path, data_only=True)
        varoq_nomi = 'KONTRAKTLAR' if 'KONTRAKTLAR' in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[varoq_nomi]

        detected_date = None
        for r in range(1, 30):
            for c in range(1, 10):
                val = str(sheet.cell(row=r, column=c).value or "")
                if 'yangilangan sanasi' in val.lower():
                    cell_val = sheet.cell(row=r, column=c + 1).value or sheet.cell(row=r, column=c + 2).value
                    if cell_val:
                        if isinstance(cell_val, datetime):
                            detected_date = cell_val
                        elif isinstance(cell_val, str):
                            try:
                                detected_date = datetime.strptime(cell_val.strip(), "%d.%m.%Y")
                            except ValueError:
                                pass

        if not detected_date:
            match = re.search(r'(\d{2}\.\d{2}\.\d{4})', os.path.basename(baza_path))
            if match:
                try:
                    detected_date = datetime.strptime(match.group(1), "%d.%m.%Y")
                except ValueError:
                    pass

        suggested_start = (detected_date + timedelta(days=1)) if detected_date else datetime.now()

        # Count students and groups
        ism_ustun = 3
        guruh_ustun = 1
        kerak_ustun = 4
        tolov_ustun = 5
        boshlanish_row = 23

        for r in range(1, 30):
            for c in range(1, 15):
                val = str(sheet.cell(row=r, column=c).value or "").lower()
                if 'guruh' in val and 'rahbar' not in val and 'soni' not in val:
                    guruh_ustun = c
                if any(x in val for x in ['familiya', 'f.i.sh', 'ism', 'sharfi']):
                    ism_ustun = c
                    boshlanish_row = r + 1
                if 'bo\'lishi' in val or 'kerak' in val:
                    kerak_ustun = c
                if any(x in val for x in ['jami', 'to\'lagan summasi', 'to\'lov']):
                    tolov_ustun = c

        groups_set = set()
        students_count = 0
        total_debt = 0.0

        for row in range(boshlanish_row, sheet.max_row + 1):
            fio = sheet.cell(row=row, column=ism_ustun).value
            if fio and str(fio).strip() and not str(fio).lower().startswith(('familiya', 'f.i.sh', 'итого', 'jami', 'guruh')):
                students_count += 1
                g_val = sheet.cell(row=row, column=guruh_ustun).value
                if g_val:
                    g_str = str(g_val).strip()
                    if g_str.endswith('.0'): g_str = g_str[:-2]
                    groups_set.add(g_str)

                kerak_val = sheet.cell(row=row, column=kerak_ustun).value or 0
                tolov_val = sheet.cell(row=row, column=tolov_ustun).value or 0
                try: k_num = float(kerak_val)
                except: k_num = 0.0
                try: t_num = float(tolov_val)
                except: t_num = 0.0

                if k_num > t_num:
                    total_debt += (k_num - t_num)

        wb.close()

        return {
            "success": True,
            "detected_date": detected_date.strftime("%d.%m.%Y") if detected_date else "",
            "suggested_start_date": suggested_start.strftime("%d.%m.%Y"),
            "total_students": students_count,
            "groups_count": len(groups_set),
            "groups": sorted(list(groups_set)),
            "total_debt": total_debt
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# ============================================================
# 4. EXECUTE CONTRACT & DEBITORKA UPDATE
# ============================================================

def execute_contract_update(baza_path, deb_path, cheklov_sanasi, session_id=None):
    """
    Baza va debitorka fayllarini taqqoslab, formulalarga tegmasdan yangilaydi,
    xulosa rasmi va to'liq tahliliy hisobotni qaytaradi.
    """
    cleanup_old_contract_temp_files(max_age_seconds=600)

    if not session_id:
        session_id = uuid.uuid4().hex[:12]

    wb_baza_write = openpyxl.load_workbook(baza_path, data_only=False)
    wb_baza_read = openpyxl.load_workbook(baza_path, data_only=True)
    wb_deb = openpyxl.load_workbook(deb_path, data_only=True)

    varoq_nomi = 'KONTRAKTLAR' if 'KONTRAKTLAR' in wb_baza_write.sheetnames else wb_baza_write.sheetnames[0]
    sheet_write = wb_baza_write[varoq_nomi]
    sheet_read = wb_baza_read[varoq_nomi]
    sheet_deb = wb_deb['bank'] if 'bank' in wb_deb.sheetnames else wb_deb.active

    ism_ustun = 3
    kerak_ustun = 4
    tolov_ustun = 5
    guruh_ustun = 1
    boshlanish_row = 23

    for r in range(1, 30):
        for c in range(1, 15):
            val = str(sheet_read.cell(row=r, column=c).value or "").lower()
            if 'guruh' in val and 'rahbar' not in val and 'soni' not in val:
                guruh_ustun = c
            if any(x in val for x in ['familiya', 'f.i.sh', 'ism', 'sharfi']):
                ism_ustun = c
                boshlanish_row = r + 1
            if 'bo\'lishi' in val or 'kerak' in val:
                kerak_ustun = c
            if any(x in val for x in ['jami', 'to\'lagan summasi', 'to\'lov']):
                tolov_ustun = c

    baza_talabalari = []
    for row in range(boshlanish_row, sheet_read.max_row + 1):
        fio = sheet_read.cell(row=row, column=ism_ustun).value
        if fio and str(fio).strip() and not str(fio).lower().startswith(('familiya', 'f.i.sh', 'итого', 'jami', 'guruh')):
            guruh_val = sheet_read.cell(row=row, column=guruh_ustun).value
            guruh_str = str(guruh_val).strip() if guruh_val else "Noma'lum"
            if guruh_str.endswith('.0'): guruh_str = guruh_str[:-2]

            eski_val = sheet_read.cell(row=row, column=tolov_ustun).value
            kerak_val = sheet_read.cell(row=row, column=kerak_ustun).value
            try: eski_sum = float(eski_val) if eski_val else 0.0
            except: eski_sum = 0.0

            try: kerak_sum = float(kerak_val) if kerak_val else 0.0
            except: kerak_sum = 0.0

            baza_talabalari.append({
                "row": row,
                "original_name": str(fio).strip(),
                "clean_name": ismlarni_standartlash(fio),
                "guruh": guruh_str,
                "kerak_summa": kerak_sum,
                "boshlangich_summa": eski_sum,
                "joriy_summa": eski_sum
            })

    yangilangan_talabalar = []
    topilmaganlar = []
    jami_tushgan_pul = 0.0
    oxirgi_to_lov_sanasi = None
    yangilangan_talabalar_set = set()

    max_deb_rows = sheet_deb.max_row
    for row in range(2, max_deb_rows + 1):
        sana_val = sheet_deb.cell(row=row, column=1).value
        if not sana_val: continue

        try:
            if isinstance(sana_val, str):
                if '.' in sana_val:
                    to_lov_sanasi = datetime.strptime(sana_val.strip(), "%d.%m.%y")
                else:
                    to_lov_sanasi = datetime.strptime(sana_val.strip(), "%Y-%m-%d %H:%M:%S")
            elif isinstance(sana_val, datetime):
                to_lov_sanasi = sana_val
            else:
                continue
        except ValueError:
            continue

        if to_lov_sanasi >= cheklov_sanasi:
            summa_val = sheet_deb.cell(row=row, column=7).value
            h_val = sheet_deb.cell(row=row, column=8).value
            deb_fio = sheet_deb.cell(row=row, column=9).value

            deb_fio_str = str(deb_fio).strip() if deb_fio else ""
            h_str = str(h_val).strip() if h_val else ""

            if not deb_fio_str and not h_str: continue
            if not summa_val: continue

            try:
                yangi_summa = float(summa_val)
            except:
                continue

            if oxirgi_to_lov_sanasi is None or to_lov_sanasi > oxirgi_to_lov_sanasi:
                oxirgi_to_lov_sanasi = to_lov_sanasi

            deb_fio_clean = ismlarni_standartlash(deb_fio_str or h_str)
            eng_yaxshi_moslik = None
            eng_yuqori_ball = 0

            for talaba in baza_talabalari:
                s_set = fuzz.token_set_ratio(deb_fio_clean, talaba["clean_name"])
                s_partial = fuzz.partial_ratio(talaba["clean_name"], deb_fio_clean)
                s_sort = fuzz.token_sort_ratio(deb_fio_clean, talaba["clean_name"])
                ball = max(s_set, s_partial, s_sort)

                if ball > eng_yuqori_ball:
                    eng_yuqori_ball = ball
                    eng_yaxshi_moslik = talaba

            if eng_yaxshi_moslik and eng_yuqori_ball >= 70:
                target_row = eng_yaxshi_moslik["row"]
                eski_summa = eng_yaxshi_moslik["joriy_summa"]
                jami_yangi = eski_summa + yangi_summa

                sheet_write.cell(row=target_row, column=tolov_ustun).value = jami_yangi
                eng_yaxshi_moslik["joriy_summa"] = jami_yangi

                jami_tushgan_pul += yangi_summa
                yangilangan_talabalar_set.add(eng_yaxshi_moslik['original_name'])

                yangilangan_talabalar.append({
                    "orig_name": eng_yaxshi_moslik['original_name'],
                    "deb_name": deb_fio_str or h_str,
                    "guruh": eng_yaxshi_moslik['guruh'],
                    "date": to_lov_sanasi.strftime('%d.%m.%Y'),
                    "amount": yangi_summa,
                    "total_paid": jami_yangi,
                    "debt_left": max(0.0, eng_yaxshi_moslik['kerak_summa'] - jami_yangi)
                })
            else:
                disp_name = deb_fio_str or h_str or "Noma'lum"
                topilmaganlar.append({
                    "name": disp_name,
                    "amount": yangi_summa,
                    "date": to_lov_sanasi.strftime('%d.%m.%Y')
                })

    oxirgi_sana_str = oxirgi_to_lov_sanasi.strftime('%d.%m.%Y') if oxirgi_to_lov_sanasi else cheklov_sanasi.strftime('%d.%m.%Y')
    for r in range(1, 30):
        for c in range(1, 10):
            val = str(sheet_write.cell(row=r, column=c).value or "")
            if 'yangilangan sanasi' in val.lower():
                target_c = c + 1
                if not sheet_write.cell(row=r, column=target_c).value:
                    target_c = c + 2
                sheet_write.cell(row=r, column=target_c).value = oxirgi_sana_str

    # Xulosa ma'lumotlari (114, 115, 116 va 0 talabali guruhlar chiqarib tashlanadi)
    xulosa_rows = []
    excluded_groups = {'114', '115', '116', '114.0', '115.0', '116.0'}

    for r in range(1, 20):
        rahbar = sheet_read.cell(row=r, column=3).value
        guruh = sheet_read.cell(row=r, column=4).value
        if rahbar and guruh and str(rahbar).strip() and str(guruh).strip():
            if str(rahbar).lower().startswith(('jami', 'итого', 'guruh rahbari')): continue
            g_str = str(guruh).strip()
            if g_str.endswith('.0'): g_str = g_str[:-2]

            # Foydalanuvchi talabi: Xulosadan 114, 115, 116 ni olib tashlash
            if g_str in excluded_groups:
                continue

            g_students = [t for t in baza_talabalari if t['guruh'] == g_str]
            soni = len(g_students)
            if soni == 0 and int(sheet_read.cell(row=r, column=5).value or 0) == 0:
                continue

            qarz_sum = sum(max(0.0, t['kerak_summa'] - t['joriy_summa']) for t in g_students)

            xulosa_rows.append({
                'rahbar': str(rahbar).strip(),
                'guruh': g_str,
                'soni': soni if soni > 0 else int(sheet_read.cell(row=r, column=5).value or 0),
                'qarz': qarz_sum
            })

    # Save final updated Excel
    out_excel_filename = f"{oxirgi_sana_str}_GACHA_KONTRAKTLAR.xlsx"
    out_excel_path = os.path.join(CONTRACT_STORAGE_DIR, f"{session_id}_{out_excel_filename}")
    wb_baza_write.save(out_excel_path)

    wb_baza_write.close()
    wb_baza_read.close()
    wb_deb.close()

    # Generate Xulosa image
    xulosa_img_filename = f"xulosa_{session_id}_{oxirgi_sana_str}.png"
    xulosa_img_path = os.path.join(CONTRACT_STORAGE_DIR, xulosa_img_filename)
    if xulosa_rows:
        generate_xulosa_table_image(xulosa_rows, xulosa_img_path)
    else:
        # Fallback dummy
        img = Image.new('RGB', (800, 200), (255, 255, 255))
        img.save(xulosa_img_path)

    # Upload to Supabase Storage
    supabase_excel_url = upload_document_to_supabase(out_excel_path, f"contracts/{session_id}_{out_excel_filename}")
    supabase_xulosa_url = upload_document_to_supabase(xulosa_img_path, f"contracts/{xulosa_img_filename}")

    keyingi_sana_dt = (oxirgi_to_lov_sanasi or cheklov_sanasi) + timedelta(days=1)
    keyingi_sana_str = keyingi_sana_dt.strftime('%d.%m.%Y')

    return {
        "success": True,
        "session_id": session_id,
        "excel_filename": out_excel_filename,
        "excel_path": out_excel_path,
        "excel_url": supabase_excel_url or f"/api/contracts/download-excel/{session_id}",
        "xulosa_img_path": xulosa_img_path,
        "xulosa_img_url": supabase_xulosa_url or f"/api/contracts/download-xulosa/{session_id}",
        "metrics": {
            "total_income": jami_tushgan_pul,
            "updated_count": len(yangilangan_talabalar_set),
            "unmatched_count": len(topilmaganlar),
            "start_date": cheklov_sanasi.strftime('%d.%m.%Y'),
            "end_date": oxirgi_sana_str,
            "next_date": keyingi_sana_str
        },
        "updated_students": yangilangan_talabalar,
        "unmatched_records": topilmaganlar,
        "xulosa_rows": xulosa_rows
    }


# ============================================================
# 5. GENERATE GROUP SCREENSHOTS & ZIP BUNDLE
# ============================================================

def execute_group_screenshots(baza_path, session_id=None):
    """Barcha guruhlar bo'yicha HD screenshotlar va ZIP paketini tayyorlash"""
    if not session_id:
        session_id = uuid.uuid4().hex[:12]

    wb = openpyxl.load_workbook(baza_path, data_only=True)
    sheet = wb.active

    header_row = 22
    date_str = datetime.now().strftime("%d.%m.%Y")

    for r in range(1, 30):
        val_a = str(sheet.cell(row=r, column=1).value or "").lower()
        val_c = str(sheet.cell(row=r, column=3).value or "").lower()
        if "guruh" in val_a or "familiy" in val_c:
            header_row = r
            break
        for c in range(1, 10):
            cell_val = str(sheet.cell(row=r, column=c).value or "")
            if "yangilangan sanasi" in cell_val.lower():
                next_cell = sheet.cell(row=r, column=c + 1).value or sheet.cell(row=r, column=c + 2).value
                if next_cell:
                    if isinstance(next_cell, datetime):
                        date_str = next_cell.strftime("%d.%m.%Y")
                    else:
                        date_str = str(next_cell).strip()

    guruhlar = {}
    for r in range(header_row + 1, sheet.max_row + 1):
        guruh_val = sheet.cell(row=r, column=1).value
        fio_val = sheet.cell(row=r, column=3).value

        if not guruh_val or not fio_val: continue

        g_name = str(guruh_val).strip()
        if g_name.endswith('.0'): g_name = g_name[:-2]
        if g_name.lower().startswith(('jami', 'итого', 'guruh')): continue

        no_val = sheet.cell(row=r, column=2).value or len(guruhlar.get(g_name, [])) + 1
        kerak_val = sheet.cell(row=r, column=4).value or 0
        jami_val = sheet.cell(row=r, column=5).value or 0
        qarzi_val = sheet.cell(row=r, column=6).value or 0

        try: kerak_num = float(kerak_val)
        except: kerak_num = 0.0
        try: jami_num = float(jami_val)
        except: jami_num = 0.0
        try: qarzi_num = float(qarzi_val)
        except: qarzi_num = 0.0

        if g_name not in guruhlar:
            guruhlar[g_name] = []

        guruhlar[g_name].append({
            "no": no_val,
            "fio": str(fio_val).strip(),
            "kerak": kerak_num,
            "jami": jami_num,
            "qarzi": qarzi_num
        })

    # 2. Xulosa ma'lumotlarini ajratib olish (114, 115, 116 chiqariladi)
    xulosa_rows = []
    excluded_groups = {'114', '115', '116', '114.0', '115.0', '116.0'}
    for r in range(1, 20):
        rahbar = sheet.cell(row=r, column=3).value
        guruh = sheet.cell(row=r, column=4).value
        if rahbar and guruh and str(rahbar).strip() and str(guruh).strip():
            if str(rahbar).lower().startswith(('jami', 'итого', 'guruh rahbari')): continue
            g_str = str(guruh).strip()
    for row in range(boshlanish_row, sheet.max_row + 1):
        fio = sheet.cell(row=row, column=ism_ustun).value
        if fio and str(fio).strip() and not str(fio).lower().startswith(('familiya', 'f.i.sh', 'итого', 'jami', 'guruh')):
            guruh_val = sheet.cell(row=row, column=guruh_ustun).value
            guruh_str = str(guruh_val).strip() if guruh_val else "Noma'lum"
            if guruh_str.endswith('.0'): guruh_str = guruh_str[:-2]

            try:
                kerak_sum = float(sheet.cell(row=row, column=kerak_ustun).value or 0)
            except Exception:
                kerak_sum = 0.0

            try:
                tolangan_sum = float(sheet.cell(row=row, column=tolov_ustun).value or 0)
            except Exception:
                tolangan_sum = 0.0

            qarz_sum = max(0.0, kerak_sum - tolangan_sum)

            if guruh_str not in guruhlar:
                guruhlar[guruh_str] = []

            guruhlar[guruh_str].append({
                'fio': str(fio).strip(),
                'kerak': kerak_sum,
                'tolandi': tolangan_sum,
                'qarzi': qarz_sum
            })

    wb.close()

    # Build Xulosa summary rows
    xulosa_rows = []
    for g_idx, g_name in enumerate(sorted(guruhlar.keys())):
        rows = guruhlar[g_name]
        tot_kerak = sum(r['kerak'] for r in rows)
        tot_tolangan = sum(r['tolandi'] for r in rows)
        tot_qarz = sum(r['qarzi'] for r in rows if r['qarzi'] > 0)
        
        rahbar_nomi = "Biriktirilmagan"
        try:
            from services.atlas_db import get_student_groups
            all_groups = get_student_groups()
            matched = next((g for g in all_groups if g.get("group_name") == g_name), None)
            if matched and matched.get("rahbar_name"):
                rahbar_nomi = matched.get("rahbar_name")
        except Exception:
            pass

        xulosa_rows.append({
            't_r': g_idx + 1,
            'guruh': g_name,
            'rahbar': rahbar_nomi,
            'soni': len(rows),
            'kerak': tot_kerak,
            'tolandi': tot_tolangan,
            'qarz': tot_qarz
        })

    generated_groups = []
    session_screenshots_dir = os.path.join(CONTRACT_STORAGE_DIR, f"screenshots_{session_id}")
    os.makedirs(session_screenshots_dir, exist_ok=True)

    zip_path = os.path.join(CONTRACT_STORAGE_DIR, f"Guruhlar_Screenshotlari_{session_id}.zip")

    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_f:
        # 1. First, generate XULOSA screenshot and add to zip
        if xulosa_rows:
            xul_filename = "00_XULOSA_Guruh_Rahbarlari.png"
            xul_img_path = os.path.join(session_screenshots_dir, xul_filename)
            generate_xulosa_table_image(xulosa_rows, xul_img_path)
            zip_f.write(xul_img_path, xul_filename)

            # Also save as main xulosa image
            main_xulosa_path = os.path.join(CONTRACT_STORAGE_DIR, f"xulosa_{session_id}_{date_str}.png")
            try:
                generate_xulosa_table_image(xulosa_rows, main_xulosa_path)
            except Exception:
                pass

            tot_xul_debt = sum(x['qarz'] for x in xulosa_rows if x['qarz'] > 0)
            tot_xul_students = sum(x['soni'] for x in xulosa_rows)

            generated_groups.append({
                "group_name": "XULOSA (Guruh Rahbarlari)",
                "is_xulosa": True,
                "student_count": tot_xul_students,
                "debt_total": tot_xul_debt,
                "image_url": f"/api/contracts/download-screenshot/{session_id}/XULOSA",
                "download_url": f"/api/contracts/download-screenshot/{session_id}/XULOSA",
                "local_path": xul_img_path
            })

        # 2. Next, generate each student group screenshot
        for g_name in sorted(guruhlar.keys()):
            rows_data = guruhlar[g_name]
            clean_filename = f"screenshot_{g_name.replace('/', '_').replace(' ', '_')}.png"
            img_path = os.path.join(session_screenshots_dir, clean_filename)

            generate_group_table_image(g_name, date_str, rows_data, img_path)
            zip_f.write(img_path, clean_filename)

            tot_debt = sum(r['qarzi'] for r in rows_data if r['qarzi'] > 0)
            generated_groups.append({
                "group_name": g_name,
                "is_xulosa": False,
                "student_count": len(rows_data),
                "debt_total": tot_debt,
                "image_url": f"/api/contracts/download-screenshot/{session_id}/{g_name}",
                "download_url": f"/api/contracts/download-screenshot/{session_id}/{g_name}",
                "local_path": img_path
            })

    # Upload zip to Supabase Storage
    sb_zip_url = upload_document_to_supabase(zip_path, f"contracts/zips/Guruhlar_Screenshotlari_{session_id}.zip")

    return {
        "success": True,
        "session_id": session_id,
        "date_str": date_str,
        "total_groups": len(generated_groups),
        "groups": generated_groups,
        "zip_url": sb_zip_url or f"/api/contracts/download-all-screenshots-zip/{session_id}",
        "zip_path": zip_path
    }


# ============================================================
# 6. FORWARD TO TELEGRAM GROUPS / CHANNELS
# ============================================================

def forward_to_telegram(chat_ids, caption_text, excel_path=None, xulosa_img_path=None, group_img_paths=None):
    """Tanlangan Telegram guruhlari yoki shaxsiy chatga bot orqali xabar, rasm va fayllarni jo'natish (Serverless safe)"""
    try:
        import telebot
        import requests
        token = os.environ.get("BOT_TOKEN") or "7737397731:AAFFwV8G6v0aE2E72q8vEaA7Jc-w3jYn7v8"
        bot = telebot.TeleBot(token)

        results = []
        if isinstance(chat_ids, (str, int)):
            chat_ids = [chat_ids]

        for cid in chat_ids:
            try:
                cid_str = str(cid).strip()
                if not cid_str: continue

                # 1. Send Caption / Summary Text
                if caption_text:
                    bot.send_message(cid_str, caption_text, parse_mode="HTML")

                # 2. Send Xulosa Image (Local file or Remote Supabase URL)
                if xulosa_img_path:
                    try:
                        if isinstance(xulosa_img_path, str) and xulosa_img_path.startswith(('http://', 'https://')):
                            resp = requests.get(xulosa_img_path, timeout=20)
                            if resp.status_code == 200:
                                xf = io.BytesIO(resp.content)
                                xf.name = "Xulosa_Hisoboti.png"
                                bot.send_photo(cid_str, photo=xf, caption="📊 <b>Guruh rahbarlari bo'yicha XULOSA hisoboti</b>", parse_mode="HTML")
                        elif isinstance(xulosa_img_path, str) and os.path.exists(xulosa_img_path):
                            with open(xulosa_img_path, 'rb') as xf:
                                bot.send_photo(cid_str, photo=xf, caption="📊 <b>Guruh rahbarlari bo'yicha XULOSA hisoboti</b>", parse_mode="HTML")
                    except Exception as img_err:
                        print(f"Xulosa send error: {img_err}")

                # 3. Send Excel Document (Local file or Remote Supabase URL) — sessiya raqamisiz toza fayl nomi
                if excel_path:
                    try:
                        clean_filename = "Yangilangan_Kontraktlar_Bazasi.xlsx"
                        if isinstance(excel_path, str):
                            raw_name = os.path.basename(excel_path.split("?")[0])
                            # Sessiya prefiksini olib tashlash (masalan: 8c1c18cb2407_11.08.2026_GACHA_KONTRAKTLAR.xlsx -> 11.08.2026_GACHA_KONTRAKTLAR.xlsx)
                            if "_" in raw_name:
                                parts = raw_name.split("_", 1)
                                if len(parts[0]) >= 8:
                                    clean_filename = parts[1]
                                else:
                                    clean_filename = raw_name
                            else:
                                clean_filename = raw_name

                        if isinstance(excel_path, str) and excel_path.startswith(('http://', 'https://')):
                            resp = requests.get(excel_path, timeout=25)
                            if resp.status_code == 200:
                                ef = io.BytesIO(resp.content)
                                ef.name = clean_filename
                                bot.send_document(cid_str, document=ef, caption="📄 <b>Yangilangan Kontraktlar Bazasi (.xlsx)</b>", parse_mode="HTML")
                        elif isinstance(excel_path, str) and os.path.exists(excel_path):
                            with open(excel_path, 'rb') as ef_file:
                                ef_bytes = io.BytesIO(ef_file.read())
                                ef_bytes.name = clean_filename
                                bot.send_document(cid_str, document=ef_bytes, caption="📄 <b>Yangilangan Kontraktlar Bazasi (.xlsx)</b>", parse_mode="HTML")
                    except Exception as doc_err:
                        print(f"Excel send error: {doc_err}")

                # 4. Send Group Screenshots
                if group_img_paths:
                    for g_img in group_img_paths:
                        try:
                            if isinstance(g_img, str) and g_img.startswith(('http://', 'https://')):
                                resp = requests.get(g_img, timeout=15)
                                if resp.status_code == 200:
                                    gf = io.BytesIO(resp.content)
                                    gf.name = "guruh_screenshot.png"
                                    bot.send_photo(cid_str, photo=gf)
                            elif isinstance(g_img, str) and os.path.exists(g_img):
                                with open(g_img, 'rb') as gf:
                                    bot.send_photo(cid_str, photo=gf)
                        except Exception as ss_err:
                            print(f"Group screenshot send error: {ss_err}")

                results.append({"chat_id": cid_str, "status": "success"})
            except Exception as err:
                print(f"Error forwarding to {cid}: {err}")
                results.append({"chat_id": str(cid), "status": "error", "error": str(err)})

        return {"success": True, "results": results}
    except Exception as e:
        return {"success": False, "error": str(e)}

